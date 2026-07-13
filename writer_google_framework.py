"""
writer_google_framework.py
Writes Framework pillar results to CoE_Google_Framework_Analysis_Templates.xlsm.

Template tab layout:
  Framework_Reference  — agent writes to: D(STATUS), H(What We Saw), I(Why It Matters), J(WYSD)
  Framework_Analysis   — agent writes to: A1(title), B3(account), B4(window), B5(download date)
  Logic and Calculation — formulas pull STATUS from Framework_Reference automatically

Column map in Framework_Reference:
  A=Block, B=ControlID, C=Name, D=STATUS, E=What, F=Why, G=How,
  H=What We Saw, I=Why It Matters, J=What You Should Do,
  K=Data Source, L=Impact(formula), M=Importance, N=Priority(formula), O=Notes
"""
from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from config import STATUS_OK, STATUS_FLAG, ControlResult
from config_google_framework import SCORING_EXCLUDED, PRIORITY_POINTS, IMPORTANCE
from reader_databricks_google import GoogleContext, clean_text


def _safe_write(ws, cell_addr: str, value) -> None:
    """Write to a cell, resolving merged-cell top-left anchor if needed."""
    from openpyxl.cell import MergedCell
    cell = ws[cell_addr]
    if isinstance(cell, MergedCell):
        for merge_range in ws.merged_cells.ranges:
            if cell_addr in merge_range:
                ws.cell(merge_range.min_row, merge_range.min_col).value = value
                return
        return
    cell.value = value


def _compute_score(results: dict) -> tuple[int, str]:
    """Binary scoring: FLAG = full penalty, OK = 0 penalty. No PARTIAL."""
    score = 100
    for cid, res in results.items():
        if cid in SCORING_EXCLUDED:
            continue
        if res.status == STATUS_FLAG:
            imp = IMPORTANCE.get(cid, 5)
            score += PRIORITY_POINTS.get(imp, -5)
    score = max(0, min(100, score))
    if score >= 75:
        grade = "Is Compliant"
    elif score >= 40:
        grade = "Need Improvement"
    else:
        grade = "Non-Compliant"
    return score, grade


def write_framework_output(
    template_path: str,
    output_path: str,
    results: dict,
    ctx: GoogleContext,
) -> None:
    wb = load_workbook(template_path, keep_vba=True)
    ws_main = wb["Framework_Analysis"]
    ws_ref  = wb["Framework_Reference"]

    score, grade = _compute_score(results)

    # ── Framework_Analysis header block ──────────────────────────────────────
    _safe_write(ws_main, "A1",
        f"{ctx.hash_name} — Google Framework Analysis")
    _safe_write(ws_main, "B3",
        f"Account: {ctx.hash_name} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.account_id}")
    if ctx.window_start and ctx.window_end and ctx.window_days:
        _safe_write(ws_main, "B4",
            f"{ctx.window_start} to {ctx.window_end} ({ctx.window_days} days)")
    if ctx.downloaded:
        ws_main["B5"] = ctx.downloaded
        ws_main["B5"].number_format = "yyyy-mm-dd hh:mm:ss"

    # ── Build control ID → row mapping from Framework_Reference ──────────────
    cid_to_row: dict[str, int] = {}
    for r in range(2, ws_ref.max_row + 1):
        raw = ws_ref[f"B{r}"].value
        cid = clean_text(raw).upper() if raw else ""
        if cid.startswith("F"):
            cid_to_row[cid] = r

    # ── Write results into Framework_Reference ────────────────────────────────
    for cid, res in results.items():
        cid_upper = cid.upper()
        if cid_upper not in cid_to_row:
            print(f"[writer_framework] WARNING: {cid} not found in reference tab — skipping.")
            continue

        rr = cid_to_row[cid_upper]

        # D = STATUS
        ws_ref[f"D{rr}"].value = res.status

        # H = What We Saw
        ws_ref[f"H{rr}"].value = res.what
        ws_ref[f"H{rr}"].alignment = Alignment(wrap_text=True, vertical="top")

        # I = Why It Matters
        ws_ref[f"I{rr}"].value = res.why
        ws_ref[f"I{rr}"].alignment = Alignment(wrap_text=True, vertical="top")

        # J = What You Should Do (wysd stored on ControlResult — access safely)
        wysd = getattr(res, "wysd", None) or ""
        if wysd:
            ws_ref[f"J{rr}"].value = wysd
            ws_ref[f"J{rr}"].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)
    print(
        f"[writer_framework] Saved: {output_path} | "
        f"Score: {score} | Grade: {grade} | "
        f"FLAG: {sum(1 for r in results.values() if r.status == STATUS_FLAG)} | "
        f"OK: {sum(1 for r in results.values() if r.status == STATUS_OK)}"
    )
