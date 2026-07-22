from __future__ import annotations

import math
import re
import warnings
from calendar import monthrange
from dataclasses import dataclass
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Per-sheet header row overrides.
# Default is row index 5 (0-based), matching the original behaviour.
# Add overrides here if any sheet has its header on a different row.
# ---------------------------------------------------------------------------
SHEET_HEADER_ROW: dict[str, int] = {
    # sheet_name: 0-based row index of the header
    # e.g. '01_Advertiser_Name': 0,
}
DEFAULT_HEADER_ROW = 5


@dataclass
class DatabricksContext:
    path: str
    hash_name: str = ''
    tenant_id: str = ''
    account_id: str = ''
    window_start: object = None
    window_end: object = None
    downloaded: object = None
    window_days: Optional[int] = None
    ref_date: object = None
    ay: str = ''
    am: str = ''
    bn: str = ''
    au: str = ''
    bw: str = ''
    o7: object = None
    ax7: object = None
    journey_h7: object = None
    proj_h: object = None
    proj_i: object = None
    proj_j: object = None
    proj_k: object = None
    proj_cs_notes: str = ''
    df02: Optional[pd.DataFrame] = None
    df04: Optional[pd.DataFrame] = None
    df05: Optional[pd.DataFrame] = None
    df07: Optional[pd.DataFrame] = None
    df14: Optional[pd.DataFrame] = None
    df37: Optional[pd.DataFrame] = None
    # ── Last complete month KPIs (tab 04) ─────────────────────────────────────
    lm_label: str = ''                  # e.g. 'May 2026'
    lm_total_sales: Optional[float] = None
    lm_ad_sales: Optional[float] = None
    lm_ad_spend: Optional[float] = None
    lm_acos: Optional[float] = None
    lm_tacos: Optional[float] = None
    lm_organic_sales: Optional[float] = None
    # ── MoM: last complete month vs prior month (tab 04) ──────────────────────
    mom_label: str = ''                 # e.g. 'May vs Apr 2026'
    mom_total_sales_chg: Optional[float] = None   # (lm - pm) / pm
    mom_ad_spend_chg: Optional[float] = None
    mom_acos_chg: Optional[float] = None          # absolute pp change
    mom_tacos_chg: Optional[float] = None
    # ── QoQ: L3M vs P3M aggregates (tab 04) ──────────────────────────────────
    l3m_label: str = ''                 # e.g. 'Mar–May 2026'
    p3m_label: str = ''                 # e.g. 'Dec 2025–Feb 2026'
    l3m_total_sales: Optional[float] = None
    p3m_total_sales: Optional[float] = None
    l3m_ad_spend: Optional[float] = None
    p3m_ad_spend: Optional[float] = None
    l3m_acos: Optional[float] = None   # weighted average
    p3m_acos: Optional[float] = None
    l3m_tacos: Optional[float] = None
    p3m_tacos: Optional[float] = None
    qoq_total_sales_chg: Optional[float] = None  # (l3m - p3m) / p3m
    qoq_ad_spend_chg: Optional[float] = None
    qoq_acos_chg: Optional[float] = None         # absolute pp change
    qoq_tacos_chg: Optional[float] = None
    # ── YoY: last complete month this year vs same month last year (tab 05) ───
    yoy_label: str = ''                 # e.g. 'May 2026 vs May 2025'
    yoy_total_sales_chg: Optional[float] = None
    yoy_ad_spend_chg: Optional[float] = None
    yoy_acos_chg: Optional[float] = None          # absolute pp change
    yoy_tacos_chg: Optional[float] = None
    yoy_lm_total_sales: Optional[float] = None    # this year last month
    yoy_py_total_sales: Optional[float] = None    # prior year same month
    df26: Optional[pd.DataFrame] = None
    df27: Optional[pd.DataFrame] = None
    df28: Optional[pd.DataFrame] = None
    df29: Optional[pd.DataFrame] = None
    df31: Optional[pd.DataFrame] = None
    df32: Optional[pd.DataFrame] = None
    df33: Optional[pd.DataFrame] = None
    df34: Optional[pd.DataFrame] = None
    df35: Optional[pd.DataFrame] = None
    metrics: dict = None
    parent_count: Optional[int] = None
    top1: Optional[float] = None
    top3: Optional[float] = None
    top5: Optional[float] = None
    tags: list = None
    gap: Optional[int] = None
    last_call: object = None
    prev_call: object = None
    # ── Tab 55: Salesforce Consolidated (CSP structured fields) ──────────────
    sf_primary_objective: str = ''
    sf_primary_objective_context: str = ''
    sf_near_term: str = ''
    sf_near_term_conflict: str = ''       # 'Yes', 'No', or '' (not assessed)
    sf_current_challenges: str = ''
    sf_primary_spend_kpi: str = ''        # 'ACOS', 'ROAS', 'TACOS', or ''
    sf_acos_constraint: object = None     # numeric or None
    sf_tacos_constraint: object = None    # numeric or None
    sf_daily_target_spend: object = None  # numeric or None
    sf_target_roas: object = None         # numeric or None
    sf_sales_concentration: str = ''      # 'Low Concentration' | 'Medium Concentration' | 'High Concentration'
    sf_commodity_or_brand: str = ''       # CSP field: Commodity or Brand designation (1.3.5)
    sf_reseller: str = ''                 # CSP field: Reseller designation (1.3.11)
    sf_top_priority: str = ''             # CSP field: Top Priority for Upcoming Quarter (1.5.3)
    sf_second_priority: str = ''          # CSP field: Second Priority for Upcoming Quarter (1.5.4)
    sf_expansion_opportunity: str = ''    # CSP field: Biggest Expansion Opportunity (1.5.5)
    # ── CJM stage data (from tab 55 Salesforce Consolidated) ─────────────────
    cjm_id: str = ''
    cjm_name: str = ''              # CJM record name from tab 39
    cjm_status: list = None               # [StatusS1..S4] — None-padded to length 4
    cjm_strategy: list = None             # [StrategyS1..S4]
    cjm_adoption: list = None             # [AdoptionOrUpsellS1..S4]
    cjm_intro_date: list = None           # [IntroductionDateS1..S4]
    cjm_exec_date: list = None            # [ExecutionDateS1..S4]
    cjm_actual_completion: list = None    # [ActualCompletionDateStage1..S4]
    cjm_modified_date: object = None
    cjm_reviewed_date: object = None
    # ── Tab 37: Gong Call Insights (operational constraints field) ────────────
    sf_operational_constraints: str = ''    # Operational_Constraints__c — free text or None


def clean_text(v) -> str:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return ''
    return str(v).replace('&#39;', "'").strip()


def to_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not pd.isna(v):
        return float(v)
    s = clean_text(v)
    if not s or s.lower() in {'nan', 'none', 'null', '-'}:
        return None
    s = s.replace('$', '').replace(',', '').strip()
    if s.endswith('%'):
        try:
            return float(s[:-1]) / 100.0
        except Exception:
            return None
    m = re.match(r'^([0-9]*\.?[0-9]+)k$', s, re.I)
    if m:
        return float(m.group(1)) * 1000.0
    try:
        return float(s)
    except Exception:
        return None


def norm_pct(v) -> Optional[float]:
    x = to_float(v)
    if x is None:
        return None
    return x if x <= 1 else x / 100.0


def pct_str(v: Optional[float], decimals: int = 1) -> str:
    if v is None:
        return 'Not documented'
    return f'{v * 100:.{decimals}f}%'


def money_str(v: Optional[float]) -> str:
    if v is None:
        return 'Not documented'
    return f'${v:,.0f}'


def trim(s: str, n: int = 260) -> str:
    s = re.sub(r'\s+', ' ', s or '').strip()
    return s if len(s) <= n else s[: n - 1].rstrip() + '…'


def _parse_header(ws) -> dict:
    """Extract account metadata from sheet 01_Advertiser_Name."""
    a1 = clean_text(ws['A1'].value)
    a2 = clean_text(ws['A2'].value)
    a3 = clean_text(ws['A3'].value)
    a4 = clean_text(ws['A4'].value)
    hash_name = re.sub(r'\s*-\s*Advertiser[_\s]*Name\s*$', '', a1, flags=re.I).strip()
    m2 = re.search(r'Tenant ID:\s*(.*?)\s*\|\s*Account ID:\s*(.*)$', a2)
    m3 = re.search(r'Date Range:\s*([0-9\-]+)\s*to\s*([0-9\-]+)', a3)
    tenant = m2.group(1).strip() if m2 else ''
    account_id = m2.group(2).strip() if m2 else ''
    start = datetime.strptime(m3.group(1), '%Y-%m-%d').date() if m3 else None
    end = datetime.strptime(m3.group(2), '%Y-%m-%d').date() if m3 else None
    downloaded = datetime.strptime(a4.replace('Downloaded:', '').strip(), '%Y-%m-%d %H:%M:%S') if a4 else None
    return {
        'hash_name': hash_name,
        'tenant_id': tenant,
        'account_id': account_id,
        'window_start': start,
        'window_end': end,
        'downloaded': downloaded,
        'window_days': (end - start).days + 1 if start and end else None,
        'ref_date': downloaded.date() if downloaded else end,
    }


def _find_sheet(wb, prefix: str) -> Optional[str]:
    """Return the first sheet name that starts with prefix, or None."""
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return name
    return None


def _get_ws(wb, prefix: str):
    """Return the worksheet whose name starts with prefix, or None."""
    name = _find_sheet(wb, prefix)
    return wb[name] if name else None


def _get_df_from_xl(xl: pd.ExcelFile, sheet_prefix: str) -> Optional[pd.DataFrame]:
    """Read a sheet from an open pd.ExcelFile handle into a DataFrame using prefix matching.
    Uses xl.parse() so the file is never re-opened — critical for large accounts."""
    sheet = next((s for s in xl.sheet_names if s.startswith(sheet_prefix)), None)
    if sheet is None:
        return None
    try:
        header_row = SHEET_HEADER_ROW.get(sheet_prefix, DEFAULT_HEADER_ROW)
        df = xl.parse(sheet, header=header_row)
        return df
    except Exception:
        return None


def _ws_to_df(ws) -> Optional[pd.DataFrame]:
    """Convert an already-open openpyxl worksheet to a DataFrame.
    Only used for sheets that need single-cell access (38, 39, 54).
    For bulk DataFrame sheets use _get_df_from_xl instead."""
    try:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= DEFAULT_HEADER_ROW:
            return None
        headers = [str(c) if c is not None else f'Unnamed_{i}' for i, c in enumerate(rows[DEFAULT_HEADER_ROW])]
        data = rows[DEFAULT_HEADER_ROW + 1:]
        df = pd.DataFrame(data, columns=headers)
        df = df.dropna(how='all')
        return df if not df.empty else None
    except Exception:
        return None


def _latest_row_by_modstamp(df: pd.DataFrame) -> pd.Series:
    """
    Return the row with the most recent SystemModstamp value.
    Falls back to the first row if the column is absent or unparseable.
    """
    modstamp_col = next(
        (c for c in df.columns if 'systemmod' in str(c).lower() or 'modstamp' in str(c).lower()),
        None
    )
    if modstamp_col:
        try:
            df = df.copy()
            df['_ts'] = pd.to_datetime(df[modstamp_col], errors='coerce')
            valid = df.dropna(subset=['_ts'])
            if not valid.empty:
                return valid.loc[valid['_ts'].idxmax()].drop(labels=['_ts'])
        except Exception as _e:
            # Surfaced (was silent): a failure here silently returns the wrong row
            # (first instead of most-recent). Still degrades gracefully to iloc[0].
            print(f"[reader_mastery] most-recent-row selection failed, using first row: "
                  f"{type(_e).__name__}: {_e}")
    return df.iloc[0]


def _find_col_name(df: pd.DataFrame, *candidates: str) -> Optional[str]:
    """Return the first column name that matches any candidate (case-insensitive, ignoring _ and spaces)."""
    norm = {re.sub(r'[\s_]', '', str(c)).lower(): c for c in df.columns}
    for cand in candidates:
        key = re.sub(r'[\s_]', '', cand).lower()
        if key in norm:
            return norm[key]
    return None


def latest_gap_days(call_df: Optional[pd.DataFrame]):
    if call_df is None or call_df.empty or 'Gong__Call_End__c' not in call_df.columns:
        return None, None, None
    dates = pd.to_datetime(call_df['Gong__Call_End__c'], errors='coerce').dropna().sort_values()
    if len(dates) == 0:
        return None, None, None
    if len(dates) == 1:
        return None, dates.iloc[-1], None
    return int((dates.iloc[-1] - dates.iloc[-2]).days), dates.iloc[-1], dates.iloc[-2]


def load_databricks_context(path: str) -> DatabricksContext:
    # Two-pass strategy to minimise memory on large accounts:
    #   Pass 1 — pd.ExcelFile for all bulk DataFrame sheets (pandas engine,
    #             memory-efficient, never re-opens the file)
    #   Pass 2 — openpyxl read_only=True for header + single-cell reads
    #             (sheets 38, 39, 54 and account header from 01)

    # --- Pass 1: bulk DataFrames via pandas ---
    with pd.ExcelFile(path, engine="calamine") as xl:
        df02 = _get_df_from_xl(xl, '02_Date_Range_KPIs__Date_Range_')
        df04 = _get_df_from_xl(xl, '04_L24M_Monthly_Performance_Sum')
        df05 = _get_df_from_xl(xl, '05_Monthly_Sales_YoY_Comparison')
        df07 = _get_df_from_xl(xl, '07_KPIs_by_Parent_ASIN_by_Month')
        df14 = _get_df_from_xl(xl, '14_Campaign_Performance_by_Adve')
        df37 = _get_df_from_xl(xl, '37_Gong_Call_Insights_for_Sales')
        df26 = _get_df_from_xl(xl, '26_Unmanaged_ASIN')
        df27 = _get_df_from_xl(xl, '27_Timeframe_Boost')
        df28 = _get_df_from_xl(xl, '28_Unmanaged_Budget')
        df29 = _get_df_from_xl(xl, '29_Negative_Keywords__Global')
        df31 = _get_df_from_xl(xl, '31_Unmanaged_campaigns')
        df32 = _get_df_from_xl(xl, '32_Unmanaged_Campaigns_Budget_O')
        df33 = _get_df_from_xl(xl, '33_RBO_Configuration_Insights')
        df34 = _get_df_from_xl(xl, '34_Product_Level_ACoS')
        df35 = _get_df_from_xl(xl, '35_Campaign_Level_ACoS')

    # --- Pass 2: openpyxl for header + single-cell sheets ---
    wb = load_workbook(path, data_only=True, read_only=True)

    try:
        ws01 = _get_ws(wb, '01_Advertiser_Name')
        if ws01 is None:
            raise ValueError('Sheet starting with 01_Advertiser_Name not found in export.')
        h = _parse_header(ws01)
        ctx = DatabricksContext(path=path, **h)

        # Attach DataFrames loaded in Pass 1
        ctx.df02 = df02
        ctx.df04 = df04
        ctx.df05 = df05
        ctx.df07 = df07
        ctx.df14 = df14
        ctx.df37 = df37
        ctx.df26 = df26
        ctx.df27 = df27
        ctx.df28 = df28
        ctx.df29 = df29
        ctx.df31 = df31
        ctx.df32 = df32
        ctx.df33 = df33
        ctx.df34 = df34
        ctx.df35 = df35

        # --- Tab 38: Client Success Insights — latest row by SystemModstamp ---
        ws38 = _get_ws(wb, '38_Client_Success_Insights_Repo')
        if ws38 is None:
            raise ValueError('Sheet starting with 38_Client_Success_Insights_Repo not found in export.')
        df38 = _ws_to_df(ws38)
        if df38 is None or df38.empty:
            import warnings
            warnings.warn(
                f'38_Client_Success_Insights_Repo has no data rows for {path}. '
                'Client success fields will be treated as missing.'
            )
            row38 = pd.Series([None] * 200)
        else:
            row38 = _latest_row_by_modstamp(df38)
        # Map by column position (headers may vary); fall back to positional index.
        # Original cells: AY7, AM7, BN7, AL7, AU7, BW7, O7, AX7
        # Column letters are 0-based index: A=0, O=14, AL=37, AM=38, AU=46,
        # AX=49, AY=50, BN=65, BW=75
        def _pos(letter: str):
            letter = letter.upper()
            idx = 0
            for ch in letter:
                idx = idx * 26 + (ord(ch) - ord('A') + 1)
            return idx - 1  # 0-based

        def _cell_val(row: pd.Series, letter: str):
            i = _pos(letter)
            return row.iloc[i] if i < len(row) else None

        ctx.ay  = clean_text(_cell_val(row38, 'AY'))
        ctx.am  = clean_text(_cell_val(row38, 'AM'))
        ctx.bn  = clean_text(_cell_val(row38, 'BN'))
        # Near_Term_and_Primary_Objective_Conflict__c — read by name from tab 38 (same field as tab 55)
        _conflict_38 = df38[df38.columns[df38.columns.str.lower() == 'near_term_and_primary_objective_conflict__c'].tolist()[0]].iloc[0] if any(df38.columns.str.lower() == 'near_term_and_primary_objective_conflict__c') else None
        ctx.sf_near_term_conflict = clean_text(_conflict_38) if _conflict_38 is not None and str(_conflict_38) != 'nan' else ''
        ctx.au  = clean_text(_cell_val(row38, 'AU'))
        ctx.bw  = clean_text(_cell_val(row38, 'BW')).upper()
        ctx.o7  = _cell_val(row38, 'O')
        ctx.ax7 = _cell_val(row38, 'AX')

        # --- Tab 39: Client Journey Insights — SSOT for all CJM fields ---
        ws39 = _get_ws(wb, '39_Client_Journey_Insights_Data')
        ctx.journey_h7 = ws39['H7'].value if ws39 is not None else None

        if ws39 is not None:
            df39 = _ws_to_df(ws39)
            if df39 is not None and not df39.empty:
                row39 = df39.iloc[0]  # single CJM record per account

                def _s39(field_name: str) -> str:
                    """Return cleaned string from row39 by column name (case-insensitive)."""
                    if field_name in row39.index:
                        return clean_text(row39[field_name])
                    for col in row39.index:
                        if str(col).lower() == field_name.lower():
                            return clean_text(row39[col])
                    return ''

                def _s39_raw(field_name: str):
                    """Return raw value from row39 (dates, numerics)."""
                    if field_name in row39.index:
                        v = row39[field_name]
                        return None if pd.isna(v) else v
                    for col in row39.index:
                        if str(col).lower() == field_name.lower():
                            v = row39[col]
                            return None if pd.isna(v) else v
                    return None

                ctx.cjm_id            = _s39('Id')
                ctx.cjm_name          = _s39('Name')
                ctx.cjm_modified_date = _s39_raw('LastModifiedDate')
                ctx.cjm_reviewed_date = _s39_raw('CGM_Last_Reviewed_Date__c')
                ctx.cjm_status        = [_s39(f'StatusS{i}__c') or None for i in range(1, 5)]
                ctx.cjm_strategy      = [_s39(f'StrategyS{i}__c') or None for i in range(1, 5)]
                ctx.cjm_adoption      = [_s39(f'AdoptionOrUpsellS{i}__c') or None for i in range(1, 5)]
                ctx.cjm_intro_date    = [_s39_raw(f'IntroductionDateS{i}__c') for i in range(1, 5)]
                ctx.cjm_exec_date     = [_s39_raw(f'ExecutionDateS{i}__c') for i in range(1, 5)]
                ctx.cjm_actual_completion = [
                    _s39_raw(f'ActualCompletionDateStage{i}__c') for i in range(1, 5)
                ]

        # --- Tab 54: Project Dataset — filter by Advertiser_ID, latest by Modstamp ---
        ws54 = _get_ws(wb, '54_Project_Dataset_on_SF')
        if ws54 is None:
            raise ValueError('Sheet starting with 54_Project_Dataset_on_SF not found in export.')
        df54 = _ws_to_df(ws54)
        if df54 is None or df54.empty:
            import warnings
            warnings.warn(
                f'54_Project_Dataset_on_SF has no data rows for {path}. '
                'Project dataset fields will be treated as missing.'
            )
            row54 = pd.Series([None] * 200)
        else:
            # Filter to rows matching this export's Advertiser_ID
            adv_col = _find_col_name(df54, 'Advertiser_ID_c', 'Advertiser_ID', 'AdvertiserID')
            if adv_col and ctx.account_id:
                matched = df54[df54[adv_col].astype(str).str.strip() == str(ctx.account_id).strip()]
                df54_filtered = matched if not matched.empty else df54
            else:
                df54_filtered = df54

            # Pick latest row by SystemModstamp
            row54 = _latest_row_by_modstamp(df54_filtered)

        def _col54(letter: str):
            i = _pos(letter)
            return row54.iloc[i] if i < len(row54) else None

        ctx.proj_h        = _col54('H')
        ctx.proj_i        = _col54('I')
        ctx.proj_j        = _col54('J')
        ctx.proj_k        = _col54('K')
        ctx.proj_cs_notes = clean_text(_col54('T'))

        # --- Tab 55: Salesforce Consolidated — CSP structured fields + CJM stages ---
        ws55 = _get_ws(wb, '55_Salesforce_Consolidated')
        if ws55 is not None:
            df55 = _ws_to_df(ws55)
            if df55 is not None and not df55.empty:
                row55 = _latest_row_by_modstamp(df55)

                def _s55(field_name: str) -> str:
                    """Return a cleaned string value from row55 by column name."""
                    if field_name in row55.index:
                        return clean_text(row55[field_name])
                    # case-insensitive fallback
                    for col in row55.index:
                        if str(col).lower() == field_name.lower():
                            return clean_text(row55[col])
                    return ''

                def _s55_raw(field_name: str):
                    """Return raw value from row55 (for dates, numerics)."""
                    if field_name in row55.index:
                        v = row55[field_name]
                        return None if pd.isna(v) else v
                    for col in row55.index:
                        if str(col).lower() == field_name.lower():
                            v = row55[col]
                            return None if pd.isna(v) else v
                    return None

                # CSP structured fields
                ctx.sf_primary_objective       = _s55('Primary_Objective__c')
                ctx.sf_primary_objective_context = _s55('Primary_Objective_Additional_Context__c')
                ctx.sf_near_term               = _s55('Near_Term_3_Month_Considerations__c')
                # Only use tab 55 conflict field if tab 38 didn't already provide it
                if not ctx.sf_near_term_conflict:
                    ctx.sf_near_term_conflict = _s55('Near_Term_and_Primary_Objective_Conflict__c')
                ctx.sf_current_challenges      = _s55('Current_Challenges__c')
                ctx.sf_primary_spend_kpi       = _s55('Primary_Spend_KPI__c').upper()
                ctx.sf_acos_constraint         = _s55_raw('ACOS_Constraint__c')
                ctx.sf_tacos_constraint        = _s55_raw('TACOS_Constraint__c')
                ctx.sf_daily_target_spend      = _s55_raw('daily_target_spend__c')
                ctx.sf_target_roas             = _s55_raw('Target_ROAS__c')
                ctx.sf_sales_concentration     = _s55('Sales_Concentration__c')
                ctx.sf_commodity_or_brand      = _s55('Commodity_or_Brand__c')       # not confirmed in export — may be blank
                ctx.sf_reseller                = _s55('Reseller__c')                  # not confirmed in export — may be blank
                ctx.sf_top_priority            = _s55('Top_Priority__c')              # confirmed column name in tab 55
                ctx.sf_second_priority         = _s55('Second_Priority__c')           # confirmed column name in tab 55
                ctx.sf_expansion_opportunity   = _s55('Biggest_Expansion_Opportunity__c')

                # CJM modified date fallback from tab 55 (if tab 39 not loaded)
                # Tab 39 is SSOT — this is only used if tab 39 failed to load
                if ctx.cjm_modified_date is None:
                    ctx.cjm_modified_date = _s55_raw('CJM_LastModifiedDate')
            else:
                warnings.warn(
                    f'55_Salesforce_Consolidated_PreA has no data rows for {path}. '
                    'Salesforce CSP/CJM structured fields will be unavailable.',
                )
        else:
            warnings.warn(
                f'Sheet starting with 55_Salesforce_Consolidated not found in {path}. '
                'Salesforce CSP/CJM structured fields will be unavailable.',
            )

    finally:
        try:
            wb.close()
        except Exception:
            pass

    # --- Derived metrics (no workbook needed) ---
    ctx.metrics = {}
    if ctx.df02 is not None and not ctx.df02.empty:
        row = ctx.df02.iloc[0]
        ctx.metrics = {k: row.get(k) for k in ['AdSales', 'TotalSales', 'AdSpend', 'TACoS', 'ACoS', 'Clicks', 'Revenue']}
        acos_val = row.get('ACoS')
        ctx.metrics['ROAS'] = (1 / float(acos_val)) if pd.notna(acos_val) and float(acos_val) != 0 else None

    if ctx.df07 is not None and not ctx.df07.empty and 'ParentASIN' in ctx.df07.columns and 'ThisYearTotalSales' in ctx.df07.columns:
        grp = ctx.df07.groupby('ParentASIN', dropna=True)['ThisYearTotalSales'].sum().sort_values(ascending=False)
        total = float(grp.sum()) if not grp.empty else 0.0
        ctx.parent_count = int(grp.index.nunique()) if not grp.empty else 0
        if total > 0:
            ctx.top1 = float(grp.head(1).sum() / total)
            ctx.top3 = float(grp.head(3).sum() / total)
            ctx.top5 = float(grp.head(5).sum() / total)

    ctx.tags = []
    if ctx.df14 is not None and not ctx.df14.empty:
        for c in ['Tag1', 'Tag2', 'Tag3', 'Tag4', 'Tag5']:
            if c in ctx.df14.columns:
                ctx.tags.extend([clean_text(x) for x in ctx.df14[c].dropna().tolist() if clean_text(x)])

    ctx.gap, ctx.last_call, ctx.prev_call = latest_gap_days(ctx.df37)

    # --- Tab 37: Operational Constraints field ---
    if ctx.df37 is not None and not ctx.df37.empty:
        constraints_col = _find_col_name(
            ctx.df37,
            'Operational_Constraints__c',
            'OperationalConstraints__c',
            'OperationalConstraints',
            'Operational Constraints',
        )
        if constraints_col:
            raw_val = ctx.df37[constraints_col].dropna()
            if not raw_val.empty:
                ctx.sf_operational_constraints = clean_text(raw_val.iloc[-1])

    # --- Monthly KPI derivation (tab 04 + tab 05) ---
    # Last complete month = last month whose period is strictly before window_end month.
    # e.g. window_end = 2026-06-23 → last complete month = May 2026 (period 2026-05).
    if ctx.df04 is not None and not ctx.df04.empty and ctx.window_end is not None:
        try:
            df04_work = ctx.df04.copy()
            df04_work['Month'] = pd.to_datetime(df04_work['Month'], errors='coerce')
            df04_work = df04_work.dropna(subset=['Month']).sort_values('Month')
            current_period = pd.Period(ctx.window_end, 'M')
            complete = df04_work[df04_work['Month'].dt.to_period('M') < current_period]

            if len(complete) >= 1:
                lm_row = complete.iloc[-1]
                ctx.lm_label = lm_row['Month'].strftime('%b %Y')
                ctx.lm_total_sales   = lm_row.get('TotalSales')
                ctx.lm_ad_sales      = lm_row.get('AdSales')
                ctx.lm_ad_spend      = lm_row.get('AdSpend')
                ctx.lm_acos          = lm_row.get('ACoS')
                ctx.lm_tacos         = lm_row.get('TACoS')
                ctx.lm_organic_sales = lm_row.get('OrganicSales')

                # MoM: last month vs the one before it
                if len(complete) >= 2:
                    pm_row = complete.iloc[-2]
                    pm_label = pm_row['Month'].strftime('%b %Y')
                    ctx.mom_label = f"{ctx.lm_label} vs {pm_label}"

                    def _chg(a, b):
                        try:
                            a, b = float(a), float(b)
                            return (a - b) / b if b != 0 else None
                        except Exception:
                            return None

                    ctx.mom_total_sales_chg = _chg(lm_row.get('TotalSales'), pm_row.get('TotalSales'))
                    ctx.mom_ad_spend_chg    = _chg(lm_row.get('AdSpend'), pm_row.get('AdSpend'))
                    # ACoS/TACoS: absolute pp change (both already 0–1 decimals)
                    try:
                        ctx.mom_acos_chg  = float(lm_row.get('ACoS'))  - float(pm_row.get('ACoS'))
                        ctx.mom_tacos_chg = float(lm_row.get('TACoS')) - float(pm_row.get('TACoS'))
                    except Exception:
                        pass

            # QoQ: L3M (last 3 complete months) vs P3M (prior 3 months)
            if len(complete) >= 6:
                l3m = complete.iloc[-3:]
                p3m = complete.iloc[-6:-3]
                ctx.l3m_label = f"{l3m.iloc[0]['Month'].strftime('%b')}–{l3m.iloc[-1]['Month'].strftime('%b %Y')}"
                ctx.p3m_label = f"{p3m.iloc[0]['Month'].strftime('%b')}–{p3m.iloc[-1]['Month'].strftime('%b %Y')}"

                def _safe_sum(rows, col):
                    try:
                        return float(rows[col].sum())
                    except Exception:
                        return None

                def _wavg_ratio(rows, num_col, den_col):
                    """Weighted average of a ratio: sum(num) / sum(den)."""
                    try:
                        n = float(rows[num_col].sum())
                        d = float(rows[den_col].sum())
                        return n / d if d != 0 else None
                    except Exception:
                        return None

                ctx.l3m_total_sales = _safe_sum(l3m, 'TotalSales')
                ctx.p3m_total_sales = _safe_sum(p3m, 'TotalSales')
                ctx.l3m_ad_spend    = _safe_sum(l3m, 'AdSpend')
                ctx.p3m_ad_spend    = _safe_sum(p3m, 'AdSpend')
                ctx.l3m_acos        = _wavg_ratio(l3m, 'AdSpend', 'AdSales')
                ctx.p3m_acos        = _wavg_ratio(p3m, 'AdSpend', 'AdSales')
                ctx.l3m_tacos       = _wavg_ratio(l3m, 'AdSpend', 'TotalSales')
                ctx.p3m_tacos       = _wavg_ratio(p3m, 'AdSpend', 'TotalSales')

                def _chg(a, b):
                    try:
                        a, b = float(a), float(b)
                        return (a - b) / b if b != 0 else None
                    except Exception:
                        return None

                ctx.qoq_total_sales_chg = _chg(ctx.l3m_total_sales, ctx.p3m_total_sales)
                ctx.qoq_ad_spend_chg    = _chg(ctx.l3m_ad_spend, ctx.p3m_ad_spend)
                try:
                    ctx.qoq_acos_chg  = ctx.l3m_acos  - ctx.p3m_acos  if ctx.l3m_acos  and ctx.p3m_acos  else None
                    ctx.qoq_tacos_chg = ctx.l3m_tacos - ctx.p3m_tacos if ctx.l3m_tacos and ctx.p3m_tacos else None
                except Exception:
                    pass
        except Exception as _e:
            warnings.warn(f'Monthly KPI derivation (tab 04) failed for {path}: {_e}')

    # YoY: last complete month this year vs same month last year (tab 05)
    if ctx.df05 is not None and not ctx.df05.empty and ctx.lm_label and ctx.window_end is not None:
        try:
            df05_work = ctx.df05.copy()
            df05_work['Month'] = pd.to_datetime(df05_work['Month'], errors='coerce')
            df05_work = df05_work.dropna(subset=['Month'])
            # Match row where year == window_end.year and month == last complete month
            lm_month = (pd.Period(ctx.window_end, 'M') - 1).month
            lm_year  = (pd.Period(ctx.window_end, 'M') - 1).year
            yoy_row_mask = (
                (df05_work['Month'].dt.month == lm_month) &
                (df05_work['Month'].dt.year  == lm_year)
            )
            yoy_rows = df05_work[yoy_row_mask]
            if not yoy_rows.empty:
                yr = yoy_rows.iloc[0]
                py_label = f"{yr['Month'].strftime('%b')} {lm_year - 1}"
                ctx.yoy_label = f"{ctx.lm_label} vs {py_label}"
                ctx.yoy_lm_total_sales = yr.get('ThisYearTotalSales')
                ctx.yoy_py_total_sales = yr.get('LastYearTotalSales')

                def _yoy_chg(col_ty, col_ly):
                    try:
                        ty = float(yr.get(col_ty))
                        ly = float(yr.get(col_ly))
                        return (ty - ly) / ly if ly != 0 else None
                    except Exception:
                        return None

                ctx.yoy_total_sales_chg = _yoy_chg('ThisYearTotalSales', 'LastYearTotalSales')
                ctx.yoy_ad_spend_chg    = _yoy_chg('ThisYearAdSpend', 'LastYearAdSpend')
                # ACoS/TACoS: absolute pp change
                try:
                    ctx.yoy_acos_chg  = float(yr.get('ThisYearACoS'))  - float(yr.get('LastYearACoS'))
                    ctx.yoy_tacos_chg = float(yr.get('ThisYearTACoS')) - float(yr.get('LastYearTACoS'))
                except Exception:
                    pass
        except Exception as _e:
            warnings.warn(f'Monthly KPI derivation (tab 05 YoY) failed for {path}: {_e}')

    return ctx


def monthly_budget_from_daily(ctx: DatabricksContext) -> Optional[float]:
    """
    Estimates monthly budget from the daily budget target in proj_h.
    Uses the month of window_end. Returns None with a warning if the
    window spans more than one calendar month, since the estimate would
    be misleading.
    """
    if ctx.window_end is None:
        return None
    daily = to_float(ctx.proj_h)
    if daily is None:
        return None
    if ctx.window_start is not None and ctx.window_start.month != ctx.window_end.month:
        warnings.warn(
            f"monthly_budget_from_daily: window spans "
            f"{ctx.window_start} to {ctx.window_end} (multiple months). "
            f"Budget estimate uses {ctx.window_end.strftime('%B %Y')} only and may be inaccurate.",
            stacklevel=2,
        )
    return daily * monthrange(ctx.window_end.year, ctx.window_end.month)[1]
