"""
writer_google_implementation.py
Writes Implementation results to CoE_Google_Account_Implementation_Analysis_Templates.xlsm.

Tab structure (post-cleanup):
  Account Implement_Analysis       — header, grade, key findings (auto-populated by LET formula)
  Account Implement_Reference  — 12 controls, agent writes D/H/I/J per row
  Logic and Calculation            — scoring engine, references (U) tab

Writer targets the (U) tab exclusively. The old reference tab has been deleted from the template.

Column map for Account Implement_Reference:
  A=Block  B=ControlID  C=Name  D=STATUS  E=What(def)  F=Why  G=How
  H=What We Saw  I=Why It Matters  J=What You Should Do
  K=Data Source  L=Impact  M=Importance  N=Priority(formula)  O=Notes

Agent writes: D (STATUS), H (What We Saw), I (Why It Matters), J (What You Should Do)
"""
from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment

from reader_databricks_google import GoogleContext, clean_text
from config_google_implementation import SCORING_EXCLUDED, PRIORITY_POINTS, IMPORTANCE
from config import STATUS_OK, STATUS_FLAG, STATUS_PARTIAL


def _safe_write(ws, row: int, col: int, value):
    """Write to a cell, resolving merged cell anchors transparently."""
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for merge_range in ws.merged_cells.ranges:
            if cell in merge_range:
                ws.cell(merge_range.min_row, merge_range.min_col).value = value
                return
        return
    cell.value = value


def _compute_score(results: dict) -> tuple:
    score = 100
    for cid, res in results.items():
        if cid in SCORING_EXCLUDED:
            continue
        if res.status in (STATUS_FLAG, STATUS_PARTIAL):
            imp = IMPORTANCE.get(cid, 5)
            pts = PRIORITY_POINTS.get(imp, -15)
            if res.status == STATUS_PARTIAL:
                pts = pts // 2
            score += pts
    score = max(0, min(100, score))
    if score >= 75:   grade = "Healthy"
    elif score >= 40: grade = "Needs Attention"
    else:             grade = "At Risk"
    return score, grade


def write_implementation_output(
    template_path: str,
    output_path: str,
    results: dict,
    ctx: GoogleContext,
) -> None:
    wb = load_workbook(template_path, keep_vba=True)

    # ── Validate tabs exist ───────────────────────────────────────────────────
    REF_TAB = "Account Implement_Reference"
    if REF_TAB not in wb.sheetnames:
        raise RuntimeError(f"Expected tab '{REF_TAB}' not found in template. "
                           f"Available: {wb.sheetnames}")

    ws_main = wb["Account Implement_Analysis"]
    ws_ref  = wb[REF_TAB]

    score, grade = _compute_score(results)

    # ── Analysis tab — header block ───────────────────────────────────────────
    _safe_write(ws_main, 1, 1, f"{ctx.hash_name} — Google Implementation Analysis")
    _safe_write(ws_main, 3, 2,
        f"Account: {ctx.hash_name} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.account_id}")
    if ctx.window_start and ctx.window_end and ctx.window_days:
        _safe_write(ws_main, 4, 2,
            f"{ctx.window_start} to {ctx.window_end} ({ctx.window_days} days)")
    if ctx.downloaded:
        cell = ws_main.cell(5, 2)
        if not isinstance(cell, MergedCell):
            cell.value = ctx.downloaded
            cell.number_format = "yyyy-mm-dd hh:mm:ss"

    # ── Reference (U) tab — build Control ID → row index ─────────────────────
    # Control ID is always column B (index 2); header is row 1, data starts row 2
    cid_to_row: dict = {}
    for r in range(2, ws_ref.max_row + 1):
        raw = ws_ref.cell(r, 2).value
        cid = clean_text(raw).upper() if raw else ""
        if cid.startswith("I"):
            cid_to_row[cid] = r

    # ── Write per-control results ─────────────────────────────────────────────
    # Columns: D=4 STATUS | H=8 What We Saw | I=9 Why It Matters | J=10 What You Should Do
    for cid, res in results.items():
        rr = cid_to_row.get(cid)
        if rr is None:
            print(f"[writer_implementation] WARNING: {cid} not found in '{REF_TAB}' — skipping.")
            continue

        _safe_write(ws_ref, rr, 4, res.status)   # D — STATUS
        _safe_write(ws_ref, rr, 8, res.what)      # H — What We Saw
        _safe_write(ws_ref, rr, 9, res.why)       # I — Why It Matters
        _safe_write(ws_ref, rr, 10, res.why)      # J — What You Should Do (same as why; editable post-run)

        for col in (8, 9, 10):
            cell = ws_ref.cell(rr, col)
            if not isinstance(cell, MergedCell):
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)
    print(f"[writer_implementation] Saved: {output_path} | Score: {score} | Grade: {grade}")
