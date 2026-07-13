"""
reader_databricks_google.py
Shared reader for all Google CoE pillars.
Loads all 40 tabs from the Google Databricks export.
Header row is always index 5 (Excel row 6) — hard locked.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


# ── Public helpers ────────────────────────────────────────────────────────────

def to_float(v) -> Optional[float]:
    if pd.isna(v):
        return None
    try:
        return float(str(v).replace(",", "").replace("$", "").replace("%", "").strip())
    except (ValueError, TypeError):
        return None


def to_str(v) -> str:
    if pd.isna(v):
        return ""
    return str(v).strip()


def pct_str(v: Optional[float], decimals: int = 1) -> str:
    if v is None:
        return "N/A"
    return f"{v * 100:.{decimals}f}%"


def money_str(v: Optional[float]) -> str:
    if v is None:
        return "N/A"
    return f"${v:,.2f}"


def num_str(v: Optional[float], decimals: int = 0) -> str:
    if v is None:
        return "N/A"
    return f"{v:,.{decimals}f}"


def clean_text(v) -> str:
    if pd.isna(v):
        return ""
    return str(v).strip().replace("\n", " ").replace("\r", " ")


def _parse_date(v) -> Optional[date]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    dt = pd.to_datetime(s, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.date()


def _parse_datetime(v) -> Optional[datetime]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    dt = pd.to_datetime(s, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.to_pydatetime()


# ── Tab name registry ─────────────────────────────────────────────────────────
# Maps logical key → list of tab prefixes, checked in order.
# Supports two export generations:
#   OLD (Equip Supply style): tabs 13-40 at their original numbers
#   NEW (current): all tabs shifted down by 1 (old 13→new 12, old 22→new 21, etc.)
#                  plus a new tab 40_DPL_Tag_Coverage at the end
# The reader tries every prefix in the list and uses the first match found.

TAB_KEYS = {
    # ── Stable tabs (01-09): identical across all export generations ──────────
    "ADVERTISER_NAME":         ["01_Advertiser_Name"],
    "DATE_RANGE_KPIS":         ["02_Date_Range_KPIs"],
    "YEARLY_KPIS":             ["03_Yearly_KPIs"],
    "L24M_MONTHLY":            ["04_L24M_Monthly"],
    "MONTHLY_SALES_YOY":       ["05_Monthly_Sales"],
    "CAMPAIGN_REPORT":         ["06_Campaign_Report"],
    "CHANNEL_TYPE":            ["07_Campaigns_by_Channel"],
    "PRODUCT_SHOPPING":        ["08_Product_Shopping"],
    "KEYWORD_REPORT":          ["09_Keyword_Report"],
    # ── Shifted tabs: gen1=old number, gen2=mid number, gen3(48-tab)=new number
    # Each list is checked in order; first match wins.
    # Search Terms:       gen1=11, gen2=10
    "SEARCH_TERMS":            ["10_Search_Terms_Report",    "11_Search_Terms_Report"],
    # Search Classifier:  gen1=12, gen2=11
    "SEARCH_CLASSIFIER":       ["11_Search_Terms_Classifier","12_Search_Terms_Classifier"],
    # Campaign Gold:      gen1=13, gen2=12
    "CAMPAIGN_GOLD":           ["12_Campaign_Gold",          "13_Campaign_Gold"],
    # Campaign Metadata:  gen1=14, gen2=13
    "CAMPAIGN_METADATA":       ["13_Campaign_Metadata",      "14_Campaign_Metadata"],
    # Stripe:             gen1=15, gen2=14
    "STRIPE_INFO":             ["14_Stripe_and_Account",     "15_Stripe_and_Account"],
    # Device:             gen1=16, gen2=15
    "DEVICE_BREAKDOWN":        ["15_Device_Breakdown",       "16_Device_Breakdown"],
    # Product Monthly:    gen1=17, gen2=16
    "PRODUCT_MONTHLY_KPIS":    ["16_Product_Monthly",        "17_Product_Monthly"],
    # PMAX Channels:      gen1=18, gen2=17
    "PMAX_CHANNELS":           ["17_PMAX_Channels",          "18_PMAX_Channels"],
    # MultiChannel:       gen1=19, gen2=18
    "MULTICHANNEL_PRODUCTS":   ["18_MultiChannel_Products",  "19_MultiChannel_Products"],
    # Price Competitiveness: gen1=20, gen2=19
    "PRICE_COMPETITIVENESS":   ["19_Price_Competitiveness",  "20_Price_Competitiveness"],
    # Campaign Month CDM: gen1=21, gen2=20
    "CAMPAIGN_MONTH_CDM":      ["20_Campaign_Month",         "21_Campaign_Month"],
    # Client Success:     gen1=22, gen2=21
    "CLIENT_SUCCESS":          ["21_Client_Success",         "22_Client_Success"],
    # Campaign Perf CDM:  gen1=23, gen2=22
    "CAMPAIGN_PERF_CDM":       ["22_Campaign_Performance",   "23_Campaign_Performance"],
    # PLA Summary:        gen1=24, gen2=23
    "PLA_SUMMARY":             ["23_PLA_Summary",            "24_PLA_Summary"],
    # KPIs CDM:           gen1=25, gen2=24
    "KPIS_CDM":                ["24_KPIs_CDM",               "25_KPIs_CDM"],
    # Location Perf:      gen1=26, gen2=25, gen3=23
    "LOCATION_PERF":           ["23_Location_Performance",   "25_Location_Performance",  "26_Location_Performance"],
    # Amazon Product:     gen1=27, gen2=26, gen3=24
    "AMAZON_PRODUCT":          ["24_Amazon_Product",         "26_Amazon_Product",        "27_Amazon_Product"],
    # DPL Performance:    gen1=28, gen2=27, gen3=25
    "DPL_PERFORMANCE":         ["25_DPL_Performance",        "27_DPL_Performance",       "28_DPL_Performance"],
    # Search Terms CDM:   gen1=29, gen2=28, gen3=26
    "SEARCH_TERMS_CDM":        ["26_Search_Terms_Performance","28_Search_Terms_Performance","29_Search_Terms_Performance"],
    # Feed Products:      gen1=30, gen2=29, gen3=27
    "FEED_PRODUCTS":           ["27_Feed_Products",          "29_Feed_Products",         "30_Feed_Products"],
    # Negative Keywords:  gen1=31, gen2=30, gen3=28
    "NEGATIVE_KEYWORDS":       ["28_Negative_Keywords",      "30_Negative_Keywords",     "31_Negative_Keywords"],
    # Asset Groups:       gen1=32, gen2=31, gen3=29
    "ASSET_GROUPS":            ["29_Google_Asset_Groups",    "31_Google_Asset_Groups",   "32_Google_Asset_Groups"],
    # Assets Extensions:  gen1=33, gen2=32, gen3=30
    "ASSETS_EXTENSIONS":       ["30_Google_Assets_Extensions","32_Google_Assets_Extensions","33_Google_Assets_Extensions"],
    # Advertiser Details: gen1=34, gen2=33, gen3=31
    "ADVERTISER_DETAILS":      ["31_Google_Advertiser_Details","33_Google_Advertiser_Details","34_Google_Advertiser_Details"],
    # Campaigns V2:       gen1=35, gen2=34, gen3=32
    "CAMPAIGNS_V2_ENRICHED":   ["32_Google_Campaigns_V2_Enriched","34_Google_Campaigns_V2_Enriched","35_Google_Campaigns_V2_Enriched"],
    # Product Groups:     gen1=36, gen2=35, gen3=33
    "PRODUCT_GROUPS":          ["33_Google_Product_Groups_Listin","35_Google_Product_Groups","36_Google_Product_Groups"],
    # Ad Group Ads:       gen1=37, gen2=36, gen3=34
    "AD_GROUP_ADS":            ["34_Google_Ad_Group_Ads",    "36_Google_Ad_Group_Ads",   "37_Google_Ad_Group_Ads"],
    # Campaign Settings:  gen1=38, gen2=37, gen3=35
    "CAMPAIGN_SETTINGS":       ["35_Google_Campaign_Settings","37_Google_Campaign_Settings","38_Google_Campaign_Settings"],
    # Ad Groups:          gen1=39, gen2=38, gen3=36
    "AD_GROUPS":               ["36_Google_Ad_Groups",       "38_Google_Ad_Groups",      "39_Google_Ad_Groups"],
    # Account Links:      gen1=40, gen2=39, gen3=37
    "ACCOUNT_LINKS":           ["37_Google_Account_Links",   "39_Google_Account_Links",  "40_Google_Account_Links"],
    # DPL Tag Coverage:   gen2=40, gen3=38
    "DPL_TAG_COVERAGE":        ["38_DPL_Tag_Coverage",       "40_DPL_Tag_Coverage"],
    # ── New tabs — 48-tab export format (gen3) only ───────────────────────────
    # Product Issue Summary: GMC product-level issues with servability flags
    "PRODUCT_ISSUE_SUMMARY":   ["41_Product_Issue_Summary"],
    # Conversion Actions: full conversion action config — Primary/Secondary, Category, Status
    "CONVERSION_ACTIONS":      ["42_Conversion_Actions"],
    # Product Health Summary: direct DisapprovalRate, TotalProducts, DisapprovedProducts
    "PRODUCT_HEALTH_SUMMARY":  ["43_Product_Health_Summary"],
    # Platform Connections: Shopify/e-commerce platform connection status
    "PLATFORM_CONNECTIONS":    ["44_Platform_Connections"],
    # Feed Duplicate Groups: GMC duplicate feed detection
    "FEED_DUPLICATE_GROUPS":   ["45_Feed_Duplicate_Groups"],
    # GMC Feed Status: feed health from GMC side
    "GMC_FEED_STATUS":         ["46_GMC_Feed_Status"],
    # Portal Feed Monitoring: QT Portal feed export status, HoursSinceLastUpdate, FreshPortalFeedFlag
    "PORTAL_FEED_MONITORING":  ["47_Portal_Feed_Monitoring"],
    # Salesforce Google Targets: SF-linked ROAS/budget targets for Google
    "SALESFORCE_GOOGLE_TARGETS": ["48_Salesforce_Google_Targets"],
}


# ── Context dataclass ─────────────────────────────────────────────────────────

@dataclass
class GoogleContext:
    workbook_path: str

    # Identity
    hash_name: str
    tenant_id: str
    account_id: str
    downloaded: Optional[datetime]

    # Eval window
    window_start: Optional[date]
    window_end: Optional[date]
    window_days: Optional[int]
    window_str: str

    # All sheets — keyed by logical TAB_KEY
    sheets: Dict[str, pd.DataFrame] = field(default_factory=dict)


# ── Header extraction ─────────────────────────────────────────────────────────

def _extract_header(path: str) -> Tuple[str, str, str, Optional[date], Optional[date], Optional[datetime]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = None
    for s in wb.sheetnames:
        if str(s).strip().lower().startswith("01_"):
            sheet = s
            break
    if sheet is None:
        wb.close()
        return "", "", "", None, None, None

    ws = wb[sheet]
    a1 = ws["A1"].value
    hash_name = re.sub(r"\s*-\s*Advertiser_Name\s*$", "", str(a1 or "").strip(), flags=re.IGNORECASE).strip()

    tenant_id = account_id = ""
    start = end = None
    downloaded_dt = None

    for r in range(1, 25):
        cells = [str(ws.cell(r, c).value or "") for c in range(1, 15)]
        line = " ".join(cells).strip()
        low = line.lower()

        if "tenant id" in low and "advertiser id" in low:
            mt = re.search(r"Tenant\s*ID:\s*([0-9a-fA-F-]{8,})", line)
            ma = re.search(r"Advertiser\s*ID:\s*([0-9]{6,})", line)
            if mt:
                tenant_id = mt.group(1).strip()
            if ma:
                account_id = ma.group(1).strip()

        if "date range" in low:
            m = re.search(r"(\d{4}-\d{2}-\d{2})\s*to\s*(\d{4}-\d{2}-\d{2})", line)
            if m:
                start = _parse_date(m.group(1))
                end = _parse_date(m.group(2))

        if "downloaded" in low and downloaded_dt is None:
            m = re.search(r"(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})", line)
            if m:
                downloaded_dt = _parse_datetime(m.group(1))

    wb.close()
    return hash_name, tenant_id, account_id, start, end, downloaded_dt


# ── Main loader ───────────────────────────────────────────────────────────────

def load_google_export(path: str) -> GoogleContext:
    """
    Load a Google Databricks export workbook.
    Returns a GoogleContext with all tabs loaded into sheets dict keyed by logical TAB_KEY.

    Handles two export generations transparently:
      - Old format: tabs numbered 13-40 (e.g. Equip Supply style)
      - New format: tabs shifted down by 1 (old 13→12, old 22→21, etc.)
                    plus new 40_DPL_Tag_Coverage tab
    TAB_KEYS maps each logical key to a list of candidate prefixes tried in order.
    Tabs returning NO DATA or empty are stored as empty DataFrames.
    """
    xls = pd.ExcelFile(path, engine="calamine")

    # Build reverse map: prefix → logical key (first-match wins within each key's list)
    prefix_to_key: Dict[str, str] = {}
    for logical_key, prefixes in TAB_KEYS.items():
        for prefix in prefixes:
            if prefix not in prefix_to_key:
                prefix_to_key[prefix] = logical_key

    def _match_key(sheet_name: str) -> Optional[str]:
        for prefix, key in prefix_to_key.items():
            if sheet_name.startswith(prefix):
                return key
        return None

    sheets: Dict[str, pd.DataFrame] = {}

    for sname in xls.sheet_names:
        key = _match_key(sname)
        if key is None:
            continue
        # Don't overwrite if this key was already loaded from a higher-priority prefix
        if key in sheets and not sheets[key].empty:
            continue
        try:
            df = pd.read_excel(xls, sheet_name=sname, header=5)
            # Drop Unnamed columns from merged title rows
            df = df.loc[:, ~df.columns.astype(str).str.match(r"^Unnamed:\s*\d+$", na=False)].copy()
            # Detect NO DATA tabs
            if len(df.columns) == 1 and "NO DATA" in str(df.columns[0]).upper():
                df = pd.DataFrame()
            sheets[key] = df
        except Exception as e:
            print(f"[reader_google] WARNING: could not load tab {sname}: {e}")
            sheets[key] = pd.DataFrame()

    hash_name, tenant_id, account_id, h_start, h_end, downloaded_dt = _extract_header(path)

    if h_start and h_end:
        window_days = (h_end - h_start).days + 1
        window_str = f"{h_start} to {h_end} ({window_days} days)"
    else:
        window_days = None
        window_str = "UNKNOWN WINDOW"

    return GoogleContext(
        workbook_path=path,
        hash_name=hash_name,
        tenant_id=tenant_id,
        account_id=account_id,
        downloaded=downloaded_dt,
        window_start=h_start,
        window_end=h_end,
        window_days=window_days,
        window_str=window_str,
        sheets=sheets,
    )


def get_sheet(ctx: GoogleContext, key: str) -> pd.DataFrame:
    """
    Return a sheet by TAB_KEY. Returns empty DataFrame if missing or no data.
    Never raises — callers must check len(df) == 0.
    """
    return ctx.sheets.get(key, pd.DataFrame())


def find_col(df: pd.DataFrame, candidates: list) -> Optional[str]:
    """
    Case-insensitive, underscore-insensitive column lookup.
    Returns the actual column name or None.
    """
    norm = {str(c).strip().lower().replace(" ", "").replace("_", ""): c for c in df.columns}
    for cand in candidates:
        key = str(cand).strip().lower().replace(" ", "").replace("_", "")
        if key in norm:
            return norm[key]
    return None


def get_active_campaigns(ctx: "GoogleContext") -> pd.DataFrame:
    """
    Return enabled campaigns from CAMPAIGN_SETTINGS (Tab 37/38).
    Both old and new exports now use the same campaigns_v2 source with:
      State (lowercase: enabled/paused/removed) and IsEnabled (bool).
    Falls back to CAMPAIGNS_V2_ENRICHED if CAMPAIGN_SETTINGS is empty.
    Always normalises AdvertisingChannelType to UPPERCASE before returning.
    """
    df = get_sheet(ctx, "CAMPAIGN_SETTINGS")
    if df.empty:
        df = get_sheet(ctx, "CAMPAIGNS_V2_ENRICHED")
    if df.empty:
        return pd.DataFrame()

    # Normalise State/IsEnabled — field names are lowercase in campaigns_v2
    state_col = find_col(df, ["State", "state"])
    enabled_col = find_col(df, ["IsEnabled", "isenabled"])

    if state_col:
        mask = df[state_col].astype(str).str.lower() == "enabled"
        df = df[mask].copy()
    elif enabled_col:
        mask = df[enabled_col].astype(str).str.lower().isin(["true", "1"])
        df = df[mask].copy()

    # Normalise AdvertisingChannelType to uppercase so all rules can compare consistently
    ch_col = find_col(df, ["AdvertisingChannelType"])
    if ch_col:
        df[ch_col] = df[ch_col].astype(str).str.upper()

    return df.copy()
