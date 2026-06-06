"""
writer_strategy.py
──────────────────
Writes the Strategy Analysis output workbook.

What it does
────────────
1.  Reads the pre-analysis Databricks workbook via reader_databricks_strategy.
2.  Writes the Questionnaire Survey - AMZ tab (header fields, Salesforce data, Gong notes).
3.  Runs the strategy auto-flag logic and writes OK / PARTIAL / FLAG to column G
    of the STRATEGY OVERVIEW tab for every row that has a data signal.
    Rows with no automatable signal are left as-is (default OK in template).
4.  Reads back every row from STRATEGY OVERVIEW and copies the ones that are
    FLAG or PARTIAL into the Account Strategy _Analysis findings table.
    OK rows that belong to a section with at least one FLAG/PARTIAL are also
    included as context (marked OK).
5.  Writes header, grade, and interpretation to the Analysis tab.
6.  Writes the ChildASIN View tab.
7.  Saves the output .xlsm file.

Auto-flag logic (column G, STRATEGY OVERVIEW)
─────────────────────────────────────────────
Row  3  ACoS target above constraint by >10pp                             → FLAG
Row  3  ACoS target above constraint by 5–10pp                            → PARTIAL
Row  4  ACoS consistently decreasing, changes ≥2 in 30 days              → PARTIAL
Row  6  ACoS being loosened (direction = increasing)                      → FLAG
Row  7  NB ACoS ≥3× branded AND branded spend ≥25% AND target above branded → FLAG
Row  8  Account has OOB events in period                                  → FLAG
Row 10  Slow mover ASINs (<3 orders) have BA spend AND no ATM/WATM       → FLAG
Row 10  Slow mover ASINs in BA but WATM exists                           → PARTIAL
Row 12  ATM+BA overlap on ASIN >80 orders AND CPC >$1.20                 → FLAG
Row 13  ATM+BA overlap on ASIN >80 orders (general)                      → PARTIAL
Row 17  Single parent ASIN but bulk multi-ASIN structures active          → FLAG
Row 18  CPC increased YoY by >20%                                         → FLAG
Row 18  CPC increased YoY by 10–20%                                       → PARTIAL
Row 20  Account has OOB events in period (budget constraint note)         → FLAG
Row 29  Framework compliance gap: Imported+NonQT > 40% of spend          → FLAG
Row 29  Framework compliance gap: Imported+NonQT 20–40%                  → PARTIAL
Row 30  SPT present AND SPT campaign ACoS above constraint                → PARTIAL
Row 31  SPT present AND any Tier 100 ASIN has SPT spend                  → PARTIAL
Row 32  ATM < 3% of spend (severely underweighted)                       → FLAG
Row 32  ATM 3–8% of spend                                                → PARTIAL
Row 37  Any ASIN with <3 orders has BA spend                             → FLAG
Row 37  BA active AND ACoS above constraint (no slow movers detected)    → PARTIAL
Row 39  BA active but fewer than 2 campaigns (no category segmentation)  → FLAG
Row 41  >80 campaigns with only 1–3 orders in period                     → FLAG
Row 41  >40 campaigns with only 1–3 orders in period                     → PARTIAL
Row 45  No SB spend at all                                                → FLAG
Row 47  Imported spend > 0 (import kickoff needed)                       → FLAG
Row 62  No CAT_SP AND OP outperforming by ≥20% AND account has headroom  → FLAG
Row 62  No CAT_SP AND OP outperforming AND ACoS above constraint         → PARTIAL
Row 63  No SBV campaigns                                                  → FLAG
Row 67  No OP / product-target campaigns detected                         → PARTIAL
Row 71  Multiple WATM campaigns (>1) active                               → PARTIAL
Row 76  Missing WATM or CatchAll (coverage gap)                          → FLAG
Row 82  No SD campaigns and SD impressions = 0                            → FLAG
Row 83  No SD_ATC / ProSuite ATC campaign                                 → PARTIAL
Row 84  SD active but no SD_PRD campaigns                                 → PARTIAL
Row 86  Portfolios present but 0 managed                                  → PARTIAL
Row 86  0 portfolios with budget cap and >3 portfolios                    → FLAG
Row 78  BAK campaign >15% of spend AND ACoS >100% of constraint          → FLAG
Row 78  BAK campaign >15% of spend AND ACoS >50% of constraint           → PARTIAL
Row 87  Non-QT + Imported > 50% of spend                                 → FLAG
Row 88  Campaign-level ACoS overrides detected                            → PARTIAL
Row 89  Product-level ACoS overrides detected                             → PARTIAL
Row 93  SBV campaigns present but not named with SBV_ convention         → PARTIAL
Row 94  Campaigns not in portfolio > 0                                    → FLAG
Row 95  Campaigns not in portfolio > 0 (rename/unmanaged note)           → FLAG
Row 104 SB active (impressions > 0) but SBV spend = 0                   → FLAG
Row 111 YoY ad sales is negative                                          → FLAG
Row 120 GGS committed AND SD spend < 5% of total                         → FLAG
Row 121 GGS committed AND SD spend < 5% of total (display coverage note) → FLAG
Row 126 SD active but no SD_FLEX / remarketing campaigns                 → PARTIAL
Row 127 SD active but no ATC retargeting (duplicate of row 83 at GGS)   → PARTIAL
"""

from __future__ import annotations

import math
import os
import re
import sys
from datetime import datetime

import openpyxl
from openpyxl.formatting.rule import (
    ColorScaleRule,
    IconSetRule,
    Rule,
)
from openpyxl.styles import Font
from openpyxl.styles.differential import DifferentialStyle

from reader_databricks_strategy import read_strategy_context, StrategyContext


# ── helpers ───────────────────────────────────────────────────────────────────

def _safe(val, default=''):
    return default if val is None else val


def _read_header(ws):
    account_str = date_range = downloaded = ''
    for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                if 'Account:' in cell:
                    account_str = cell
                elif 'Date Range:' in cell:
                    date_range = cell.replace('Date Range: ', '').strip()
                elif 'Downloaded:' in cell:
                    downloaded = cell.replace('Downloaded: ', '').strip()
    return account_str, date_range, downloaded


def _find_header_row(ws, max_scan=10):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        if len([c for c in row if c is not None]) > 3:
            return i
    return None


def _tab_to_dict(ws):
    hr = _find_header_row(ws)
    if hr is None:
        return {}
    rows = list(ws.iter_rows(min_row=hr, max_row=hr + 1, values_only=True))
    if len(rows) < 2:
        return {}
    return {h: rows[1][i] for i, h in enumerate(rows[0]) if h is not None and i < len(rows[1])}


def _tab_to_records_full(ws):
    hr = _find_header_row(ws)
    if hr is None:
        return []
    headers = None
    records = []
    for row in ws.iter_rows(min_row=hr, values_only=True):
        if headers is None:
            headers = list(row)
            continue
        if not any(row):
            continue
        records.append({headers[j]: row[j] for j in range(len(headers)) if headers[j] is not None})
    return records


def _tab_to_records(ws):
    return _tab_to_records_full(ws)


def _latest_record(records):
    if not records:
        return {}
    if len(records) == 1:
        return records[0]
    col = next(
        (k for k in records[0] if re.sub(r'[\s_]', '', k).lower() in ('systemmodstamp', 'modstamp')),
        None
    )
    if col:
        try:
            import pandas as _pd
            return max(records, key=lambda r: _pd.to_datetime(r.get(col), errors='coerce') or _pd.NaT)
        except Exception:
            pass
    return records[0]


def _filter_by_advertiser(records, account_id):
    if not account_id or not records:
        return records
    def _norm(s): return re.sub(r'[\s_]', '', str(s)).lower()
    adv_col = next(
        (k for k in records[0] if 'advertiserid' in _norm(k)),
        None
    )
    if not adv_col:
        return records
    filtered = [r for r in records if str(r.get(adv_col, '')).strip() == str(account_id).strip()]
    return filtered if filtered else records


# ── auto-flag engine ──────────────────────────────────────────────────────────

# ── Control ID → template row lookup (New Strategy Overview tab) ─────────────
_SID_TO_ROW: dict[str, int] = {
    'S001': 2,   'S002': 3,   'S003': 4,   'S004': 5,   'S005': 6,
    'S006': 7,   'S007': 8,   'S008': 9,   'S009': 10,  'S010': 11,
    'S011': 12,  'S012': 13,  'S013': 14,  'S014': 15,  'S015': 16,
    'S016': 17,  'S017': 18,  'S018': 19,  'S019': 20,  'S020': 21,
    'S021': 22,  'S022': 23,  'S023': 24,  'S024': 25,  'S025': 26,
    'S026': 27,  'S027': 28,  'S028': 29,  'S029': 30,  'S030': 31,
    'S031': 32,  'S032': 33,  'S033': 34,  'S034': 35,  'S035': 36,
    'S036': 37,  'S037': 38,  'S038': 39,  'S039': 40,  'S040': 41,
    'S041': 42,  'S042': 43,  'S043': 44,  'S044': 45,  'S045': 46,
    'S046': 47,  'S047': 48,  'S048': 49,  'S049': 50,  'S050': 51,
    'S051': 52,  'S052': 53,  'S053': 54,  'S054': 55,  'S055': 56,
    'S056': 57,  'S057': 58,  'S058': 59,  'S059': 60,  'S060': 61,
    'S061': 62,  'S062': 63,  'S063': 64,  'S064': 65,  'S065': 66,
    'S066': 67,  'S067': 68,  'S068': 69,  'S069': 70,  'S070': 71,
    'S071': 72,  'S072': 73,  'S073': 74,  'S074': 75,  'S075': 76,
    'S076': 77,  'S077': 78,  'S078': 79,  'S079': 80,  'S080': 81,
    'S081': 82,  'S082': 83,  'S083': 84,  'S084': 85,  'S085': 86,
    'S086': 87,  'S087': 88,  'S088': 89,  'S089': 90,  'S090': 91,
    'S091': 92,  'S092': 93,  'S093': 94,  'S094': 95,  'S095': 96,
    'S096': 97,  'S097': 98,  'S098': 99,  'S099': 100, 'S100': 101,
    'S101': 102, 'S102': 103, 'S103': 104, 'S104': 105,
    'S105': 106, 'S106': 107, 'S107': 108, 'S108': 109, 'S109': 110,
    'S110': 111, 'S111': 112, 'S112': 113, 'S113': 114, 'S114': 115,
    'S115': 116, 'S116': 117, 'S117': 118, 'S118': 119, 'S119': 120,
    'S120': 121, 'S121': 122, 'S122': 123, 'S123': 124, 'S124': 125,
    'S125': 126,
    'S126': 127, 'S127': 128, 'S128': 129, 'S129': 130, 'S130': 131,
    'S131': 132, 'S132': 133, 'S133': 134, 'S134': 135, 'S135': 136,
    'S136': 137,
}


def _compute_flags(ctx: StrategyContext) -> dict[str, str]:
    """
    Performance-gated strategy suggestions.

    Design principle:
      Every flag requires a KPI condition — not just structural presence/absence.
      Framework pillar owns structural presence checks.
      Strategy owns TIMING and PRIORITY based on where the account is:
      ACoS vs constraint, TACoS trend, YoY growth, objective, spend efficiency.

    Returns {control_id: 'FLAG'|'PARTIAL'} keyed by S-IDs.
    The writer translates to row numbers using _SID_TO_ROW.
    """
    flags: dict[str, str] = {}

    def flag(sid: str, level: str):
        if flags.get(sid) == 'FLAG':   # never downgrade
            return
        flags[sid] = level

    # ── normalised comparisons ────────────────────────────────────────────────
    # acos_actual / tacos_actual are decimal (0.60 = 60%)
    # acos_constraint / tacos_constraint are integer pp (60 = 60%)
    acos_pp        = ctx.acos_actual  * 100
    tacos_pp       = ctx.tacos_actual * 100
    constraint     = ctx.acos_constraint
    tacos_con      = ctx.tacos_constraint
    has_constraint = constraint > 0
    has_tacos_con  = tacos_con  > 0

    # Fallback when no constraint is documented:
    # use current ACoS + 5pp as the working constraint so rules that depend
    # on constraint arithmetic don't error or silently misfire.
    # has_constraint stays False so rules gated on documented constraints don't fire.
    if not has_constraint and ctx.acos_actual > 0:
        constraint = acos_pp + 5.0
    if not has_tacos_con and ctx.tacos_actual > 0:
        tacos_con = tacos_pp + 5.0
    above_acos     = has_constraint and acos_pp  > constraint
    above_acos_10  = has_constraint and acos_pp  > constraint * 1.10
    above_tacos    = has_tacos_con  and tacos_pp > tacos_con
    above_tacos_10 = has_tacos_con  and tacos_pp > tacos_con  * 1.10
    non_qt_total   = ctx.pct_imported + ctx.pct_non_quartile
    declining_yoy  = ctx.yoy_ad_sales < -0.05
    growing_yoy    = ctx.yoy_ad_sales >  0.10
    tacos_rising   = ctx.tacos_trend == 'increasing' and ctx.tacos_trend_pp > 1.5
    spend_rising   = ctx.mom_spend_change > 0.10
    has_atc        = any(
        re.search(r'\bATC\b|SD_FLEX', n, re.IGNORECASE)
        for n in ctx.campaign_names
    )

    # ── account state gates ───────────────────────────────────────────────────
    # Used to suppress rules that only apply at certain scale or maturity levels
    at_scale        = ctx.total_spend >= 1500          # meaningful spend volume
    base_built      = ctx.pct_ba > 0 and ctx.pct_bak > 0  # core keyword layers exist
    advanced_ready  = base_built and not above_acos_10 and at_scale  # ready for upper funnel
    efficiency_ok   = not above_acos and not declining_yoy

    # ── primary objective booleans (from Salesforce via tab 38) ──────────────
    obj_growth      = ctx.primary_objective == 'Growth'
    obj_expansion   = ctx.primary_objective == 'Expansion'
    obj_brand       = ctx.primary_objective == 'Brand Building'
    obj_profit      = 'Profit Maximization' in ctx.primary_objective
    obj_recovery    = ctx.primary_objective == 'Recovery/Stabilization'
    obj_maintenance = ctx.primary_objective == 'Maintenance (holding steady)'
    obj_ntb         = 'Aquisition' in ctx.primary_objective or 'Acquisition' in ctx.primary_objective
    repeat_high     = ctx.repeat_purchase == 'High'
    repeat_low      = ctx.repeat_purchase == 'Low'
    is_commodity    = ctx.commodity_or_brand == 'Commodity'
    high_concentration = 'High' in ctx.sales_concentration

    # ── ACOS AND TARGET ───────────────────────────────────────────────────────

    # S002 — ACoS target above constraint.
    # FLAG: target > constraint + 10pp  PARTIAL: target > constraint + 5pp
    if has_constraint and ctx.acos_gap_to_constraint > 10:
        flag('S002', 'FLAG')
    elif has_constraint and ctx.acos_gap_to_constraint > 5:
        flag('S002', 'PARTIAL')

    # S003 — TACoS alignment. We reduce ACoS to indirectly reduce TACoS.
    # FLAG: TACoS actual > constraint + 5pp
    # PARTIAL: TACoS actual > constraint + 2pp
    if has_tacos_con and tacos_pp > tacos_con + 5:
        flag('S003', 'FLAG')
    elif has_tacos_con and tacos_pp > tacos_con + 2:
        flag('S003', 'PARTIAL')

    # S004 — ACoS reduction cadence.
    # FLAG: 0 changes in 30 days AND ACoS above constraint (stuck, not acting)
    # PARTIAL: changes happening but ACoS still above constraint (moving but gap remains)
    if above_acos and ctx.acos_changes_30d == 0:
        flag('S004', 'FLAG')
    elif above_acos and ctx.acos_changes_30d > 0 and ctx.acos_gap_to_constraint > 0:
        flag('S004', 'PARTIAL')

    # S006 — ACoS being loosened while performance is already below target.
    if ctx.acos_direction == 'increasing' and (above_acos or declining_yoy):
        flag('S006', 'FLAG')

    # S007 — Branded vs Non-Branded ACoS imbalance.
    # FLAG: NB ACoS ≥ 3× branded AND portal target above branded ACoS level
    # AND branded spend is at least 25% of total (imbalance only matters at scale)
    if (ctx.branded_acos > 0
            and ctx.non_branded_acos > 0
            and ctx.non_branded_acos / ctx.branded_acos >= 3.0
            and ctx.acos_current_target > ctx.branded_acos * 100
            and ctx.branded_spend_pct >= 0.25):
        flag('S007', 'FLAG')

    # S008 — OOB + ACoS not being reduced.
    # FLAG: account has OOB AND ACoS direction NOT decreasing AND above constraint
    # PARTIAL: OOB AND ACoS decreasing but still above constraint
    if ctx.has_oob and ctx.acos_direction != 'decreasing' and above_acos:
        flag('S008', 'FLAG')
    elif ctx.has_oob and ctx.acos_direction == 'decreasing' and above_acos:
        flag('S008', 'PARTIAL')

    # S009 — Framework compliance review.
    # Counts 4 key LP-transition structural gaps: no SB, no SPT, no WATM/CatchAll, no SD_SPT.
    _gaps = sum([
        ctx.spend_sb      == 0,
        ctx.spend_spt     == 0,
        ctx.watm_campaign_count == 0 and not ctx.has_catchall,
        not any(re.search(r'SD_SPT', n, re.IGNORECASE) for n in ctx.campaign_names),
    ])
    if _gaps >= 3 and (above_acos or declining_yoy):
        flag('S009', 'FLAG')
    elif _gaps == 2 and (above_acos or declining_yoy):
        flag('S009', 'PARTIAL')

    # ── OVERALL STRUCTURE ─────────────────────────────────────────────────────

    # S005 — Portfolio completion — flag when 50–80% already assigned (migration in progress).
    if at_scale and 0.50 <= ctx.campaigns_in_portfolio_pct < 0.80 and ctx.total_campaign_count > 5:
        flag('S005', 'PARTIAL')

    # S010 — Spend concentration: slow movers (<3 orders) in BA.
    # Correct spend model: ATM for best sellers (Tier 30), BA for mid sellers, WATM for slow movers.
    # FLAG: slow movers with BA spend AND (no ATM on any top-seller OR no WATM)
    # PARTIAL: slow movers in BA but WATM exists (partial coverage)
    if ctx.slow_movers_with_ba > 0 and (ctx.tier1_with_atm == 0 or ctx.watm_campaign_count == 0):
        flag('S010', 'FLAG')
    elif ctx.slow_movers_with_ba > 0 and ctx.watm_campaign_count > 0:
        flag('S010', 'PARTIAL')

    # S011 — Same issue for accounts with no ATM-qualifying ASIN (no >1.5 orders/day).
    no_atm_qualifying = ctx.tier1_with_atm == 0 and ctx.tier1_asin_count == 0
    if ctx.slow_movers_with_ba > 0 and no_atm_qualifying and ctx.watm_campaign_count == 0:
        flag('S011', 'FLAG')
    elif ctx.slow_movers_with_ba > 0 and no_atm_qualifying:
        flag('S011', 'PARTIAL')

    # S012 — ATM and BA spend on the same ASIN with high velocity.
    # FLAG: ATM spend > 0 AND BA spend > 0 AND CPC > $1.20 AND at least one ASIN has >80 orders
    # No PARTIAL — all three conditions must be met for this to matter
    if ctx.atm_ba_overlap_count > 0 and ctx.cpc_current > 1.20:
        flag('S012', 'FLAG')

    # S013 — ATM + BA overlap on high-velocity ASINs (>80 orders/30d) — general overlap signal.
    if ctx.atm_ba_overlap_count > 0:
        flag('S013', 'PARTIAL')

    # S014 — Bulk campaign structure development.
    # If no ASIN has >45 orders/period, the account needs full bulk structure deployed.
    # Also fires when BA exists but BAK harvest layer is missing.
    bulk_missing = sum([
        ctx.pct_ba       == 0,
        ctx.pct_bak      == 0,
        not ctx.has_cat_sp,
        ctx.spend_spt    == 0,
        ctx.watm_campaign_count == 0 and not ctx.has_catchall,
    ])
    no_top_seller = ctx.max_asin_orders_30d < 45 and ctx.total_spend > 500
    if no_top_seller and bulk_missing >= 3:
        flag('S014', 'FLAG')
    elif no_top_seller and bulk_missing >= 2:
        flag('S014', 'PARTIAL')
    elif ctx.pct_ba > 0 and ctx.pct_bak == 0 and ctx.total_spend > 500 and not above_tacos_10:
        flag('S014', 'FLAG')
    elif ctx.pct_ba > 0 and 0 < ctx.pct_bak < ctx.pct_ba * 0.20:
        flag('S014', 'PARTIAL')

    # S017 — Single ASIN with multi-ASIN bulk structures.
    if ctx.parent_asin_count == 1 and (ctx.has_catchall or ctx.spend_spt > 0):
        flag('S017', 'FLAG')

    # ── OVERALL PARAMETERS AND KPIs ───────────────────────────────────────────

    # S018 — CPC rise YoY. Only surfaces as a strategic concern when the
    # account is NOT growing cleanly. A rising CPC on a growing, efficient
    # account is expected market competition — not a strategy problem.
    if ctx.cpc_yoy_change_pct > 0.20 and above_acos:
        flag('S018', 'FLAG')
    elif ctx.cpc_yoy_change_pct > 0.20 and not growing_yoy:
        flag('S018', 'PARTIAL')  # CPC rising, sales not growing to offset it
    elif ctx.cpc_yoy_change_pct > 0.10 and above_acos:
        flag('S018', 'PARTIAL')

    # S019 — TACoS increasing trend approaching constraint.
    if tacos_rising and ctx.tacos_trend_pp > 0.70 and above_tacos and ctx.total_spend >= 1000:
        flag('S019', 'FLAG')
    elif tacos_rising and ctx.tacos_trend_pp > 0.70 and ctx.total_spend >= 1000:
        flag('S019', 'PARTIAL')

    # S020/S023 — OOB signals.
    # FLAG: OOB AND (ACoS above constraint OR TACoS above constraint)
    # PARTIAL: OOB AND last-3-month total sales declining (mom_sales_change < 0)
    # Silent: OOB + growing + clean (budget expansion opportunity)
    if ctx.has_oob and (above_acos or above_tacos):
        flag('S020', 'FLAG')
        flag('S023', 'FLAG')
    elif ctx.has_oob and ctx.mom_sales_change < 0:
        flag('S020', 'PARTIAL')
        flag('S023', 'PARTIAL')

    # S021 — TACoS high in absolute terms (not just vs constraint).
    # FLAG: TACoS actual > 50%  PARTIAL: TACoS actual > 30%
    if ctx.tacos_actual > 0.50:
        flag('S021', 'FLAG')
    elif ctx.tacos_actual > 0.30:
        flag('S021', 'PARTIAL')

    # S022 — Catalog activation scope. Gate on meaningful catalog size.
    if ctx.catalog_asin_count >= 10:
        spend_coverage = ctx.spending_asin_count / ctx.catalog_asin_count
        if spend_coverage < 0.10:
            flag('S022', 'FLAG')
        elif spend_coverage < 0.20:
            flag('S022', 'PARTIAL')

    # ── BASIC STRATEGY ────────────────────────────────────────────────────────

    # S029 — Unmanaged spend. More urgent when ACoS is already above constraint.
    if non_qt_total > 0.40 or (non_qt_total > 0.20 and above_acos_10):
        flag('S029', 'FLAG')
    elif non_qt_total > 0.20:
        flag('S029', 'PARTIAL')

    # S030 — SPT structure review gated on SPT campaign ACoS, not account average.
    # PARTIAL: SPT is active AND SPT campaign avg ACoS is above the account constraint
    if ctx.spend_spt > 0 and has_constraint and ctx.spt_avg_acos > 0 and ctx.spt_avg_acos > constraint / 100:
        flag('S030', 'PARTIAL')

    # S031 — SPT covering Tier 100 ASINs (slow movers should not be in SPT).
    # PARTIAL: SPT spend > 0 AND any Tier 100 ASIN has SPT spend
    if ctx.spend_spt > 0 and len(ctx.tier100_with_spt_asins) > 0:
        flag('S031', 'PARTIAL')

    # S032 — ATM expansion. Only when account has headroom and bulk-heavy
    # accounts with BAK already running are a lower priority for ATM expansion.
    bulk_heavy = (ctx.pct_ba + ctx.pct_bak + ctx.pct_spt) > 0.60 and ctx.pct_bak > 0
    if ctx.pct_atm < 0.03 and not above_tacos_10 and not bulk_heavy and not (ctx.has_oob and above_acos_10):
        flag('S032', 'FLAG')
    elif ctx.pct_atm < 0.03 and bulk_heavy and declining_yoy:
        flag('S032', 'PARTIAL')  # bulk-heavy but declining — still worth flagging
    elif ctx.pct_atm < 0.08 and not above_tacos and growing_yoy and not bulk_heavy:
        flag('S032', 'PARTIAL')

    # S037 — BA covering slow movers (<3 orders in period).
    # FLAG: any ASIN with <3 orders has BA spend
    # PARTIAL: BA active AND ACoS above constraint but no slow movers detected
    if ctx.slow_movers_with_ba > 0:
        flag('S037', 'FLAG')
    elif ctx.spend_ba > 0 and above_acos:
        flag('S037', 'PARTIAL')

    # S038 — BAK expansion from BA learnings.
    # FLAG: BA > 30% of total spend AND BAK spend = 0 (no harvest layer at all)
    # Note: if BAK already exists with same name pattern as BA → suppress
    if ctx.pct_ba > 0.30 and ctx.pct_bak == 0 and not ctx.bak_name_overlaps_ba:
        flag('S038', 'FLAG')
    elif ctx.pct_ba > 0.30 and ctx.pct_bak == 0 and ctx.bak_name_overlaps_ba:
        flag('S038', 'PARTIAL')  # BAK pattern exists but no current spend

    # S039 — BA not segmented. Only meaningful with enough spend and catalog size.
    if 0 < ctx.ba_campaign_count < 2 and ctx.total_spend > 1500 and ctx.catalog_asin_count >= 5:
        flag('S039', 'FLAG')

    # S041 — Low-order campaign consolidation. Gate on at_scale — fragmentation only matters with volume.
    if at_scale and ctx.low_order_campaign_count > 80:
        flag('S041', 'FLAG')
    elif at_scale and ctx.low_order_campaign_count > 40:
        flag('S041', 'PARTIAL')

    # S045 — SB missing. Gate on base_built — keyword structure must be in place first.
    has_product_targeting_base = (ctx.has_op and ctx.pct_op > 0) or (ctx.has_cat_sp and ctx.total_spend > 500)
    if base_built and ctx.spend_sb == 0 and has_product_targeting_base:
        if declining_yoy:
            flag('S045', 'FLAG')
        elif above_acos_10:
            flag('S045', 'PARTIAL')  # fix ACoS first, then SB
        else:
            flag('S045', 'FLAG')

    # S047 — Import kickoff. Based on non-managed campaigns only.
    # FLAG: imported > 30% of total  PARTIAL: 15–30%  Silent: ≤15%
    if ctx.total_spend > 0 and ctx.pct_imported > 0.30:
        flag('S047', 'FLAG')
    elif ctx.total_spend > 0 and ctx.pct_imported > 0.15:
        flag('S047', 'PARTIAL')

    # ── NEW DEPLOYS ───────────────────────────────────────────────────────────

    # S053 — Keyword Focus Narrowing.
    # PARTIAL: any BAK campaign has spend>$200 AND ACoS>1.5x constraint AND orders<5
    # Gate: above_acos AND at_scale — only surfaces when account is already struggling
    if ctx.inefficient_bak_count > 0 and above_acos and at_scale:
        flag('S053', 'PARTIAL')

    # S054/S055/S056 — Campaign type outperforming average.
    # These are positive suggestions: if a specific type is 20%+ better than account avg, flag it.
    # FLAG = "this is working well, consider expanding it" (not a problem)
    acct_acos = ctx.acos_actual  # decimal
    if acct_acos > 0:
        if ctx.atm_avg_acos > 0 and ctx.atm_avg_acos < acct_acos * 0.80:
            flag('S054', 'FLAG')   # ATM outperforming by >20%
        if ctx.br_avg_acos > 0 and ctx.br_avg_acos < acct_acos * 0.80:
            flag('S055', 'FLAG')   # BR_ outperforming by >20%
        if ctx.ph_avg_acos > 0 and ctx.ph_avg_acos < acct_acos * 0.80:
            flag('S056', 'FLAG')   # PH_ outperforming by >20%

    # S057 — Keyword Strategy Too Broad.
    # PARTIAL: BR consuming >15% of spend AND BR avg ACoS > 1.5x constraint AND above ACoS AND at_scale
    if ctx.br_inefficiency_flag and at_scale:
        flag('S057', 'PARTIAL')

    # S058 — Product targeting (OP_) outperforming account average by >20%.
    if ctx.op_avg_acos > 0 and ctx.acos_actual > 0 and ctx.op_avg_acos < ctx.acos_actual * 0.80:
        flag('S058', 'FLAG')

    # S034 — SD_AUDI campaigns outperforming account average by >20% (positive suggestion).
    if ctx.sd_audi_avg_acos > 0 and ctx.acos_actual > 0 and ctx.sd_audi_avg_acos < ctx.acos_actual * 0.80:
        flag('S034', 'FLAG')

    # S035 — Best-Seller Campaigns Paused.
    # FLAG: any Tier 10-30 ASIN has zero enabled ATM AND zero enabled BA spend
    # Gate: tier1_asin_count > 0 AND at_scale
    if ctx.tier1_asin_count > 0 and at_scale and ctx.top_seller_type_gaps > 0:
        flag('S035', 'FLAG')

    # S042 — SB_ campaigns outperforming account average by >20%.
    if ctx.sb_avg_acos > 0 and ctx.acos_actual > 0 and ctx.sb_avg_acos < ctx.acos_actual * 0.80:
        flag('S042', 'FLAG')

    # S043 — SBV_ campaigns outperforming account average by >20%.
    if ctx.sbv_avg_acos > 0 and ctx.acos_actual > 0 and ctx.sbv_avg_acos < ctx.acos_actual * 0.80:
        flag('S043', 'FLAG')

    # S060 — SD_FLEX_ campaigns outperforming account average by >20%.
    if ctx.sd_flex_avg_acos > 0 and ctx.acos_actual > 0 and ctx.sd_flex_avg_acos < ctx.acos_actual * 0.80:
        flag('S060', 'FLAG')

    # S062 — CAT_SP missing. Surface when OP campaigns are already outperforming account avg
    # by ≥20% — meaning product targeting works and CAT_SP is the next logical step.
    # FLAG: no CAT_SP AND spend > $500 AND (growing YoY OR ACoS ≤ constraint)
    #        AND OP campaigns outperforming by ≥20% (or no OP data — structural absence)
    # PARTIAL: no CAT_SP AND spend > $500 AND ACoS above constraint
    op_outperforming = (
        ctx.op_avg_acos > 0
        and ctx.acos_actual > 0
        and ctx.op_avg_acos < ctx.acos_actual * 0.80
    )
    if not ctx.has_cat_sp and ctx.total_spend > 500:
        if op_outperforming and (growing_yoy or not above_acos):
            flag('S062', 'FLAG')
        elif op_outperforming and above_acos:
            flag('S062', 'PARTIAL')
        elif not ctx.has_op and (growing_yoy or not above_acos):
            # No OP at all — still suggest CAT_SP when account has headroom
            flag('S062', 'FLAG')
        elif not ctx.has_op:
            flag('S062', 'PARTIAL')

    # S063 — SBV missing. Gate on base_built — SB must be well established first.
    sb_well_established = ctx.pct_sb > 0.05 and not above_acos
    if base_built and not ctx.has_sbv and ctx.spend_sbv == 0 and sb_well_established and has_product_targeting_base:
        flag('S063', 'FLAG')

    # S064 — Paused SB Campaign Rebuild.
    # PARTIAL: any SB campaign is paused with historical spend AND SB is now $0 AND not above ACoS AND at_scale
    if ctx.paused_sb_count > 0 and ctx.spend_sb == 0 and not above_acos and at_scale:
        flag('S064', 'PARTIAL')

    # S065 — SD_FLEX campaigns outperforming account average by >20%.
    if ctx.sd_flex_avg_acos > 0 and ctx.acos_actual > 0 and ctx.sd_flex_avg_acos < ctx.acos_actual * 0.80:
        flag('S065', 'FLAG')

    # S066 — SBV Campaign Reactivation.
    # PARTIAL: any SBV campaign is paused with historical spend AND SBV is now $0
    #          AND SB is active (SBV reactivation only makes sense when SB is live) AND not above ACoS
    if ctx.paused_sbv_count > 0 and ctx.spend_sbv == 0 and ctx.spend_sb > 0 and not above_acos and base_built:
        flag('S066', 'PARTIAL')

    # S067 — CAT_SP Launch (was OP Target Expansion — corrected to match template).
    op_outperforming = (
        ctx.op_avg_acos > 0
        and ctx.acos_actual > 0
        and ctx.op_avg_acos < ctx.acos_actual * 0.80
    )
    if not ctx.has_cat_sp and ctx.total_spend > 500:
        if op_outperforming and (growing_yoy or not above_acos):
            flag('S067', 'FLAG')
        elif op_outperforming and above_acos:
            flag('S067', 'PARTIAL')
        elif not ctx.has_op and (growing_yoy or not above_acos):
            flag('S067', 'FLAG')
        elif not ctx.has_op:
            flag('S067', 'PARTIAL')

    # S068 — SBV Product Targeting Launch.
    sb_well_established2 = ctx.pct_sb > 0.05 and not above_acos
    has_product_targeting_base2 = (ctx.has_op and ctx.pct_op > 0) or (ctx.has_cat_sp and ctx.total_spend > 500)
    if base_built and not ctx.has_sbv and ctx.spend_sbv == 0 and sb_well_established2 and has_product_targeting_base2:
        flag('S068', 'FLAG')

    # S069 — Broad Match Graduation Signal: BR_ outperforms OW_, suggest expanding BR_.
    if (ctx.br_campaign_count > 30 and ctx.ow_campaign_count > 30
            and ctx.ph_campaign_count < 10
            and ctx.br_avg_acos > 0 and ctx.ow_avg_acos > 0
            and ctx.br_avg_acos < ctx.ow_avg_acos):
        flag('S069', 'FLAG')

    # S071 — Exact Match Graduation Signal: OW_ outperforms BR_, suggest expanding OW_.
    if (ctx.br_campaign_count > 30 and ctx.ow_campaign_count > 30
            and ctx.ph_campaign_count < 10
            and ctx.br_avg_acos > 0 and ctx.ow_avg_acos > 0
            and ctx.ow_avg_acos < ctx.br_avg_acos):
        flag('S071', 'FLAG')

    # S072 — OP Target Expansion Opportunity.
    kw_total = ctx.br_campaign_count + ctx.ow_campaign_count + ctx.ph_campaign_count
    if ctx.op_campaign_count < 10 and kw_total > 50:
        flag('S072', 'FLAG')
    elif ctx.op_campaign_count < 10 and kw_total > 30:
        flag('S072', 'PARTIAL')

    # S073 — CatchAll Keyword Graduation. Add BAK underweight gate for precision.
    if ctx.catchall_orders > 100 and ctx.pct_bak < 0.10:
        flag('S073', 'FLAG')
    elif ctx.catchall_orders > 100:
        flag('S073', 'PARTIAL')  # CatchAll active but BAK already exists
    elif ctx.catchall_orders > 50:
        flag('S073', 'PARTIAL')

    # S074 — CatchAll Graduation Overdue.
    if ctx.catchall_orders > 100 and ctx.pct_bak < 0.10 and at_scale:
        flag('S074', 'FLAG')
    elif ctx.catchall_orders > 50 and ctx.pct_bak < 0.10 and at_scale:
        flag('S074', 'PARTIAL')

    # S075 — CAT_SP Above ACoS Target.
    if ctx.catsp_avg_acos > 0 and has_constraint and ctx.catsp_avg_acos * 100 > constraint:
        flag('S075', 'FLAG')
    elif ctx.catsp_avg_acos > 0 and has_constraint and ctx.catsp_avg_acos * 100 > constraint * 0.85:
        flag('S075', 'PARTIAL')

    # S077 — Multiple WATM. Gate on WATM consuming meaningful budget.
    if ctx.watm_campaign_count > 2 and ctx.spend_watm >= ctx.total_spend * 0.02:
        flag('S077', 'PARTIAL')

    # S082 — WATM/CatchAll catalog coverage. Gate on meaningful catalog and spend.
    if (ctx.watm_campaign_count > 0 or ctx.has_catchall) and ctx.catalog_asin_count >= 8 and ctx.total_spend >= 1000:
        coverage = ctx.spending_asin_count / ctx.catalog_asin_count
        if coverage < 0.60:
            flag('S082', 'PARTIAL')

    # S083 — WATM and CatchAll active simultaneously (redundant overlap).
    has_watm_active     = ctx.watm_campaign_count > 0
    has_catchall_active = ctx.has_catchall
    if has_watm_active and has_catchall_active:
        flag('S083', 'FLAG')

    # S084 — WATM spend < 3% of total (WATM not getting meaningful budget).
    if ctx.watm_campaign_count > 0 and ctx.pct_watm < 0.03:
        flag('S084', 'FLAG')

    # S085 — BAK campaign with high spend and ACoS above constraint.
    if has_constraint and ctx.bak_campaigns:
        for bak in ctx.bak_campaigns:
            if bak['pct_of_total'] > 0.15 and bak['acos'] > constraint / 100:
                flag('S085', 'FLAG')
                break
            elif bak['pct_of_total'] > 0.15 and bak['acos'] > (constraint / 100) * 0.50:
                flag('S085', 'PARTIAL')

    # S088 — SD Remarketing — Product View. Gate on base_built.
    if base_built and not ctx.has_sd and ctx.spend_sd == 0 and ctx.total_spend > 500:
        if (growing_yoy or ctx.spend_sb > 0) and ctx.max_asin_orders_30d >= 50:
            flag('S088', 'FLAG')
        elif not above_acos and ctx.max_asin_orders_30d >= 50:
            flag('S088', 'PARTIAL')

    # S089 — SD ATC Retargeting — ProSuite. Gate on SD being a real commitment.
    if not has_atc and ctx.has_prosuite_audiences and ctx.pct_sd >= 0.03:
        flag('S089', 'PARTIAL')

    # S090 / S091 are MANUAL — no automated flag.

    # S092 — OW Own-Page Coverage Missing.
    if ctx.has_op and not ctx.has_ow and advanced_ready:
        if not obj_recovery and not obj_maintenance:
            flag('S092', 'PARTIAL')

    # S093 — SD Suggested — PDP Maturity Too Low.
    if (not ctx.has_sd and ctx.total_spend > 500 and base_built
            and (growing_yoy or ctx.spend_sb > 0)
            and ctx.max_asin_orders_30d < 50 and ctx.max_asin_orders_30d > 0):
        flag('S093', 'PARTIAL')

    # ── GOVERNANCE ────────────────────────────────────────────────────────────

    # S094 — Portfolio governance.
    using_portfolios = ctx.campaigns_in_portfolio_pct > 0.10
    if using_portfolios and ctx.portfolio_count > 3 and ctx.portfolios_with_budget_cap == 0:
        flag('S094', 'FLAG')
    elif using_portfolios and ctx.portfolio_count > 0 and ctx.managed_portfolio_count == 0:
        flag('S094', 'PARTIAL')

    # S095 — Campaign-Level ACoS Overrides Active.
    if ctx.has_campaign_acos_overrides and above_acos:
        flag('S095', 'PARTIAL')

    # S096 — Product-Level ACoS Overrides Active.
    if ctx.has_product_acos_overrides and above_acos:
        flag('S096', 'PARTIAL')

    # S097 — VCPM overuse.
    if ctx.vcpm_spend_pct > 0.10:
        flag('S097', 'FLAG')
    elif ctx.vcpm_spend_pct > 0.05:
        flag('S097', 'PARTIAL')

    # S098 — Tagging/segmentation gap.
    if at_scale and ctx.spend_ba > 0 and ctx.spend_spt > 0 and ctx.spend_atm > 0 and not ctx.has_op:
        flag('S098', 'PARTIAL')

    # S099 — RBO Weekend Bid Management. MANUAL — no automated check.

    # S100 — SBV naming convention.
    if ctx.has_sbv and not ctx.sbv_naming_compliant:
        flag('S100', 'PARTIAL')

    # S101 — Campaigns outside portfolio. Gate: ≥50% already assigned.
    using_portfolios_50 = ctx.campaigns_in_portfolio_pct >= 0.50
    if using_portfolios_50 and ctx.campaigns_not_in_portfolio > 5:
        flag('S101', 'FLAG')
    elif using_portfolios_50 and ctx.campaigns_not_in_portfolio > 0:
        flag('S101', 'PARTIAL')

    # S102 — Best-Seller Visibility.
    # FLAG: any Tier 10-30 ASIN missing ≥2 of (ATM, BAK, OP) from tab 15
    if ctx.tier1_asin_count > 0 and ctx.top_seller_type_gaps > 0 and at_scale and not obj_expansion:
        flag('S102', 'FLAG')

    # S103 / S104 are MANUAL — no automated flag.

    # S105 — Proven Targets Not Utilised.
    # PARTIAL: ≥3 top-30 search terms with orders≥3 AND CVR≥10% AND pct_bak < 0.20 AND at_scale
    # (proxy: tab 17 terms are high-quality but BAK is underfunded — terms likely not captured)
    if ctx.unconverted_top_terms >= 3 and ctx.pct_bak < 0.20 and at_scale and base_built:
        flag('S105', 'PARTIAL')

    # S106 — ProSuite AMC Audience Testing. Gated on advanced_ready.
    if not ctx.has_prosuite_audiences and advanced_ready and growing_yoy:
        flag('S106', 'FLAG')
    elif not ctx.has_prosuite_audiences and advanced_ready:
        flag('S106', 'PARTIAL')

    # S107 — Inefficient ASIN Spend Reduction.
    # FLAG: any ASIN with spend>$100 AND ACoS>1.5x constraint AND orders<5
    if ctx.inefficient_asin_count > 0 and at_scale and not obj_expansion:
        flag('S107', 'FLAG')

    # S108 — SB Active — SBV Missing. Gate on base_built.
    if base_built and ctx.spend_sbv == 0 and ctx.pct_sb > 0.05 and not above_acos:
        flag('S108', 'FLAG')

    # S109 / S110 are MANUAL — no automated flag.

    # S111 — Recurring Sales Strategy. Suppress for low repeat-purchase and NTB objective.
    if not ctx.has_sns_active and not repeat_low and not obj_ntb:
        if declining_yoy:
            flag('S111', 'FLAG')
        elif not growing_yoy:
            flag('S111', 'PARTIAL')

    # S112 — Sales Declining While Spend Growing.
    if declining_yoy and spend_rising:
        flag('S112', 'FLAG')

    # S113 — Budget Constraint Alignment.
    # FLAG: documented monthly budget exists AND actual spend deviates >20%
    # PARTIAL: deviation >10%
    if ctx.monthly_budget > 0:
        deviation = abs(ctx.total_spend - ctx.monthly_budget) / ctx.monthly_budget
        if deviation > 0.20:
            flag('S113', 'FLAG')
        elif deviation > 0.10:
            flag('S113', 'PARTIAL')

    # S114 / S115 / S116 are MANUAL — no automated flag.

    # S117 — Subscribe & Save — Not Active. Suppress for low repeat-purchase and NTB objective.
    if not ctx.has_sns_active and not repeat_low and not obj_ntb:
        if declining_yoy:
            flag('S117', 'FLAG')
        else:
            flag('S117', 'PARTIAL')

    # S118 / S119 are MANUAL — no automated flag.

    # ── PROMO AND GGS ─────────────────────────────────────────────────────────

    # S120 — Promo Portfolio Budget Pacing. Gate on meaningful promo cost rate.
    has_named_promo_portfolio = any(
        'PROMO' in str(n).upper() for n in ctx.portfolio_names
    )
    if ctx.has_active_promo and has_named_promo_portfolio and ctx.promo_cost_rate > 0.05:
        flag('S120', 'PARTIAL')

    # S121 is MANUAL — no automated flag.

    # S122 — SD GGS Compliance — No SD Spend.
    has_sd_commitment = ctx.has_promo_portfolio or ctx.ggs_status == 'Yes'
    sd_below_threshold = ctx.spend_sd == 0 or ctx.pct_sd < 0.05
    if has_sd_commitment and sd_below_threshold:
        flag('S122', 'FLAG')
    elif ctx.ggs_status == 'No' and ctx.spend_sd == 0 and ctx.spend_sb > 0:
        flag('S122', 'PARTIAL')

    # S123 — SD Remarketing Missing.
    has_remarketing = any(
        re.search(r'SD_FLEX|SD_AUDI|remarketing', n, re.IGNORECASE)
        for n in ctx.campaign_names
    )
    if ctx.spend_sd > 0 and not has_remarketing:
        flag('S123', 'PARTIAL')

    # S124 — SD ATC Retargeting — GGS Section.
    if ctx.spend_sd > 0 and not has_atc and ctx.has_prosuite_audiences:
        flag('S124', 'PARTIAL')

    # S125 is MANUAL — no automated flag.

    # S126 — Promo Management as Channel Expansion. Only at scale and when account is healthy.
    if not ctx.has_active_promo and at_scale and efficiency_ok and not obj_recovery:
        flag('S126', 'PARTIAL')

    # ── STRUCTURAL SIGNALS (new controls, inserted in correct blocks by ID) ────

    # S018 — Discovery-Performance Mix in BA.
    if ctx.branded_nb_mixed_in_ba and ctx.pct_ba > 0 and at_scale:
        if not obj_ntb and not obj_brand:
            flag('S018', 'PARTIAL')

    # S019 — Auto-to-Manual Conversion Ratio.
    if ctx.auto_spend_pct > 0.50 and ctx.manual_exact_pct < 0.15 and at_scale:
        if not obj_growth and not obj_expansion and not obj_brand:
            flag('S019', 'FLAG')
        elif ctx.auto_spend_pct > 0.65:
            flag('S019', 'PARTIAL')

    # S025 — TACoS/ACoS Divergence.
    if (ctx.tacos_trend == 'increasing' and ctx.tacos_trend_pp > 1.5
            and ctx.acos_direction == 'decreasing' and ctx.total_spend >= 1000):
        flag('S025', 'FLAG')

    # S048 — BAK Harvest Stalled.
    if ctx.bak_underfed and at_scale and not obj_expansion:
        flag('S048', 'PARTIAL')

    # S049 — Own Product Page Undefended.
    if ctx.pct_bak > 0.15 and not ctx.has_op and at_scale:
        if not obj_recovery and not obj_maintenance:
            flag('S049', 'FLAG')

    # S050 — BR Discovery Layer Missing.
    if not ctx.has_br and ctx.pct_bak > 0.10 and at_scale:
        if not obj_ntb and not obj_recovery:
            flag('S050', 'PARTIAL')

    # S074 — CatchAll Graduation Overdue.
    if ctx.catchall_orders > 100 and ctx.pct_bak < 0.10 and at_scale:
        flag('S074', 'FLAG')
    elif ctx.catchall_orders > 50 and ctx.pct_bak < 0.10 and at_scale:
        flag('S074', 'PARTIAL')

    # S080 — CAT_SP No Qualifying Categories.
    if ctx.has_cat_sp and ctx.qualifying_category_count == 0 and ctx.total_spend > 500:
        flag('S080', 'PARTIAL')

    # S081 — BAK Branded and Non-Branded Mixed.
    if ctx.bak_branded_nb_mixed and not is_commodity and at_scale:
        flag('S081', 'PARTIAL')

    # S092 — OW Own-Page Coverage Missing.
    if ctx.has_op and not ctx.has_ow and advanced_ready:
        if not obj_recovery and not obj_maintenance:
            flag('S092', 'PARTIAL')

    # S093 — SD Suggested — PDP Maturity Too Low.
    if (not ctx.has_sd and ctx.total_spend > 500 and base_built
            and (growing_yoy or ctx.spend_sb > 0)
            and ctx.max_asin_orders_30d < 50 and ctx.max_asin_orders_30d > 0):
        flag('S093', 'PARTIAL')

    return flags


# ── dynamic What We Saw text ──────────────────────────────────────────────────

def _build_what_we_saw(ctx: StrategyContext, flags: dict[str, str]) -> dict[str, str]:
    """
    Returns {control_id: text} for every control that fired.
    Each block matches its flag condition. All numbers come from ctx.
    Short sentences only. No jargon. Written for non-native English speakers.
    SIDs with no flag rule (S024/S033/S036/S040/S044/S052/S059/S061/S127-S130)
    are MANUAL or have no automatable signal — no block needed.
    """
    texts: dict[str, str] = {}

    constraint = ctx.acos_constraint if ctx.acos_constraint > 0 else (ctx.acos_actual * 100 + 5.0)

    def pct(v: float) -> str:
        return f'{v:.0%}'

    def dollar(v: float) -> str:
        return f'${v:,.0f}'

    # S002 — ACoS target above constraint
    if 'S002' in flags:
        texts['S002'] = (
            f'The current ACoS target is {ctx.acos_current_target:.0f}%. '
            f'The account constraint is {ctx.acos_constraint:.0f}%. '
            f'The gap is +{ctx.acos_gap_to_constraint:.0f} percentage points. '
            f'The target needs to come down to align with the client objective.'
        )

    # S003 — TACoS actual above constraint
    if 'S003' in flags:
        tacos_gap = ctx.tacos_actual * 100 - ctx.tacos_constraint
        texts['S003'] = (
            f'TACoS is {ctx.tacos_actual:.0%}. '
            f'The TACoS constraint is {ctx.tacos_constraint:.0f}%. '
            f'The gap is +{tacos_gap:.1f} percentage points. '
            f'Reducing ACoS indirectly reduces TACoS — the ACoS target needs to come down.'
        )

    # S004 — ACoS reduction cadence stalled or insufficient
    if 'S004' in flags:
        texts['S004'] = (
            f'ACoS is {ctx.acos_actual:.0%} vs constraint {ctx.acos_constraint:.0f}%. '
            f'{ctx.acos_changes_30d} ACoS target change(s) in the last 30 days. '
            + ('No changes have been made. The gap is not closing.'
               if ctx.acos_changes_30d == 0
               else f'Changes are happening but ACoS is still {ctx.acos_gap_to_constraint:.0f}pp above constraint.')
        )

    # S005 — Portfolio migration in progress
    if 'S005' in flags:
        in_port = round(ctx.campaigns_in_portfolio_pct * ctx.total_campaign_count)
        not_in = ctx.total_campaign_count - in_port
        texts['S005'] = (
            f'{in_port} of {ctx.total_campaign_count} campaigns ({ctx.campaigns_in_portfolio_pct:.0%}) are already in portfolios. '
            f'{not_in} campaign(s) remain outside. Complete the portfolio assignment.'
        )

    # S006 — ACoS target being loosened
    if 'S006' in flags:
        texts['S006'] = (
            f'ACoS target changed {ctx.acos_changes_30d} time(s) in the last 30 days — direction: increasing. '
            f'Current target: {ctx.acos_current_target:.0f}%. '
            f'Spend growth is being driven by loosening efficiency, not by structural improvements.'
        )

    # S007 — Branded vs non-branded ACoS imbalance
    if 'S007' in flags:
        texts['S007'] = (
            f'Branded spend is {ctx.branded_spend_pct:.0%} of total at {ctx.branded_acos:.0%} ACoS. '
            f'Non-branded is at {ctx.non_branded_acos:.0%} ACoS vs portal target {ctx.acos_current_target:.0f}%. '
            f'The target is calibrated to branded performance, leaving non-branded campaigns overspending.'
        )

    # S008 — OOB with ACoS above constraint — reduce ACoS target
    if 'S008' in flags:
        texts['S008'] = (
            f'Account hit daily budget limits. ACoS target: {ctx.acos_current_target:.0f}% vs constraint {ctx.acos_constraint:.0f}%. '
            f'Reducing the ACoS target lowers CPC pressure and eases out-of-budget events.'
        )

    # S009 — Framework compliance gaps
    if 'S009' in flags:
        gap_labels = []
        if ctx.spend_sb == 0: gap_labels.append('no SB campaigns')
        if not ctx.has_cat_sp: gap_labels.append('no CAT_SP campaigns')
        if not ctx.has_sbv and ctx.spend_sbv == 0: gap_labels.append('no SBV campaigns')
        if not ctx.has_sd and ctx.spend_sd == 0: gap_labels.append('no SD campaigns')
        if ctx.spend_spt > 0 and ctx.pct_atm < 0.03: gap_labels.append('SPT active but ATM < 3%')
        if ctx.campaigns_not_in_portfolio > 5: gap_labels.append(f'{ctx.campaigns_not_in_portfolio} campaigns outside portfolios')
        if (ctx.pct_imported + ctx.pct_non_quartile) > 0.40: gap_labels.append(f'{pct(ctx.pct_imported + ctx.pct_non_quartile)} spend outside framework')
        if not ctx.has_op and ctx.pct_op == 0: gap_labels.append('no OP / product-target campaigns')
        if not any(re.search(r'\\bATC\\b|SD_FLEX', n, re.IGNORECASE) for n in ctx.campaign_names): gap_labels.append('no ATC retargeting')
        n_gaps = len(gap_labels)
        gaps_str = ', '.join(gap_labels[:5])
        suffix = f' (+{n_gaps - 5} more)' if n_gaps > 5 else ''
        texts['S009'] = (
            f'{n_gaps} structural framework gap(s) detected: {gaps_str}{suffix}. '
            f'A structured framework review is needed before the next QR.'
        )

    # S010 — Slow movers with BA spend
    if 'S010' in flags:
        asin_list = ', '.join(ctx.slow_mover_asins_with_ba[:5]) if ctx.slow_mover_asins_with_ba else ''
        suffix = f' ASINs: {asin_list}.' if asin_list else ''
        texts['S010'] = (
            f'{ctx.slow_movers_with_ba} ASIN(s) with fewer than 3 orders in the period have BA spend.{suffix} '
            f'Slow movers should be in WATM only — not in BA campaigns.'
        )

    # S011 — Slow movers in BA, no ATM-qualifying ASIN
    if 'S011' in flags:
        texts['S011'] = (
            f'{ctx.slow_movers_with_ba} ASIN(s) with fewer than 3 orders are in BA campaigns and no ATM-qualifying ASIN exists. '
            f'No ASIN has sufficient velocity for ATM. '
            f'WATM is the correct structure for this account — concentrate spend there.'
        )

    # S012 — ATM + BA overlap on high-velocity ASINs (>80 orders, high CPC)
    if 'S012' in flags:
        asin_list = ', '.join(ctx.atm_ba_overlap_asins[:5]) if ctx.atm_ba_overlap_asins else 'see tab 14'
        texts['S012'] = (
            f'{ctx.atm_ba_overlap_count} ASIN(s) have both ATM and BA spend with >80 orders. '
            f'CPC: ${ctx.cpc_current:.2f}. '
            f'ASINs: {asin_list}. '
            f'ATM already covers these high-velocity ASINs — BA spend is redundant and expensive.'
        )

    # S013 — ATM + BA overlap (general)
    if 'S013' in flags:
        texts['S013'] = (
            f'{ctx.atm_ba_overlap_count} ASIN(s) have spend in both ATM and BA campaigns. '
            f'Overlap can limit bid efficiency and reduce control over spend allocation. '
            f'Review whether BA campaigns are needed alongside ATM for these ASINs.'
        )

    # S014 — Bulk campaign structure incomplete
    if 'S014' in flags:
        texts['S014'] = (
            f'BA campaigns are active ({pct(ctx.pct_ba)} of spend / {dollar(ctx.spend_ba)}) '
            f'but no BAK harvest layer exists. '
            f'Discovery data is not being converted into manual precision targets.'
        )

    # S017 — Single parent ASIN with multi-ASIN bulk structures
    if 'S017' in flags:
        texts['S017'] = (
            f'The account has {ctx.parent_asin_count} parent ASIN. '
            f'Multi-ASIN bulk structures add complexity without value at this catalogue size.'
        )

    # S018 — Branded and non-branded mixed in BA auto layer
    if 'S018' in flags:
        texts['S018'] = (
            f'Branded spend is {ctx.branded_spend_pct:.0%} and non-branded is {ctx.non_branded_spend_pct:.0%} '
            f'of total search term spend — both significant inside the same auto campaign layer. '
            f'Branded and non-branded traffic should be in separate campaigns for independent bid control.'
        )

    # S019 — Auto spend dominant, BAK manual layer too thin
    if 'S019' in flags:
        texts['S019'] = (
            f'Auto campaigns (BA + ATM + WATM) account for {ctx.auto_spend_pct:.0%} of total spend. '
            f'BAK (manual exact) is only {ctx.manual_exact_pct:.0%}. '
            f'Discovery is generating learnings that are not being converted into precision manual campaigns.'
        )

    # S020 — OOB with ACoS or TACoS above constraint
    if 'S020' in flags:
        texts['S020'] = (
            f'Account ran out of budget at least once. Total spend: {dollar(ctx.total_spend)}. '
            f'ACoS: {ctx.acos_actual:.0%} vs constraint {ctx.acos_constraint:.0f}%. '
            f'Budget expansion or scope reduction should be reviewed with the client.'
        )

    # S021 — TACoS increasing trend approaching constraint
    if 'S021' in flags:
        texts['S021'] = (
            f'TACoS has been {ctx.tacos_trend} for the last 3 months ({ctx.tacos_trend_pp:+.1f}pp). '
            f'Current TACoS: {ctx.tacos_actual:.0%}. '
            f'Profitability is eroding — review ACoS target strategy to slow the TACoS increase.'
        )

    # S022 — Catalogue activation scope too low
    if 'S022' in flags:
        spend_cov = (ctx.spending_asin_count / ctx.catalog_asin_count) if ctx.catalog_asin_count > 0 else 0
        texts['S022'] = (
            f'{ctx.spending_asin_count} of {ctx.catalog_asin_count} catalogue ASINs ({spend_cov:.0%}) have active spend. '
            f'A large portion of the catalogue has no advertising coverage. '
            f'Review whether key ASINs are missing from campaign structures.'
        )

    # S023 — OOB with budget/scope concern
    if 'S023' in flags:
        texts['S023'] = (
            f'Account hit OOB. Spend: {dollar(ctx.total_spend)}. '
            f'Budget should be expanded, or product scope reduced to concentrate on top ASINs.'
        )

    # S025 — TACoS rising while ACoS improving (divergence)
    if 'S025' in flags:
        texts['S025'] = (
            f'ACoS is trending {ctx.acos_direction} while TACoS has risen '
            f'{ctx.tacos_trend_pp:+.1f}pp over the last 3 months. '
            f'When ACoS improves but TACoS rises, organic sales are likely declining '
            f'or promotional activity is distorting the total sales denominator.'
        )

    # S029 — Non-Quartile or imported spend above threshold
    if 'S029' in flags:
        non_qt = ctx.pct_imported + ctx.pct_non_quartile
        texts['S029'] = (
            f'{pct(non_qt)} of spend is in Imported or Non-Quartile campaigns '
            f'({pct(ctx.pct_imported)} Imported, {pct(ctx.pct_non_quartile)} Non-Quartile). '
            f'The account is not fully operating within the Quartile framework.'
        )

    # S030 — SPT campaign ACoS above constraint
    if 'S030' in flags:
        texts['S030'] = (
            f'SPT is active ({dollar(ctx.spend_spt)}, {pct(ctx.pct_spt)} of spend). '
            f'SPT avg ACoS: {ctx.spt_avg_acos:.0%} vs constraint {ctx.acos_constraint:.0f}%. '
            f'Defensive structure should be split by category or brand segment.'
        )

    # S031 — SPT covering Tier 100 slow-mover ASINs
    if 'S031' in flags:
        asin_list = ', '.join(ctx.tier100_with_spt_asins[:5]) if ctx.tier100_with_spt_asins else ''
        suffix = f' ASINs: {asin_list}.' if asin_list else ''
        texts['S031'] = (
            f'{len(ctx.tier100_with_spt_asins)} Tier 100 ASIN(s) have SPT spend. '
            f'SPT spend: {dollar(ctx.spend_spt)}.{suffix} '
            f'Tier 100 ASINs are slow movers and should not be in SPT campaigns.'
        )

    # S032 — ATM underweighted (< 3% of spend)
    if 'S032' in flags:
        texts['S032'] = (
            f'ATM campaigns represent {pct(ctx.pct_atm)} of spend ({dollar(ctx.spend_atm)}). '
            + ('No ATM spend detected. ' if ctx.pct_atm == 0 else '')
            + f'Automatic targeting on best-selling ASINs should be expanded.'
        )

    # S034 — SD_AUDI investment outperforming
    if 'S034' in flags:
        texts['S034'] = (
            f'SD_AUDI campaigns avg ACoS: {pct(ctx.sd_audi_avg_acos)} vs account avg {pct(ctx.acos_actual)}. '
            f'SD audience campaigns are outperforming — consider expanding SD_AUDI coverage.'
        )

    # S035 — Top-seller ASINs missing key campaign types
    if 'S035' in flags:
        texts['S035'] = (
            f'{ctx.top_seller_type_gaps} of {ctx.tier1_asin_count} top-selling ASIN(s) (Tier 10–30) '
            f'are missing ≥2 key campaign types (ATM, BAK, OP). '
            f'Best-seller campaigns have likely been paused or were never fully deployed.'
        )

    # S037 — BA covering slow-mover ASINs
    if 'S037' in flags:
        asin_list = ', '.join(ctx.slow_mover_asins_with_ba[:5]) if ctx.slow_mover_asins_with_ba else ''
        suffix = f' ASINs with <3 orders in BA: {asin_list}.' if asin_list else ''
        texts['S037'] = (
            f'BA campaigns: {dollar(ctx.spend_ba)} ({pct(ctx.pct_ba)} of spend). '
            f'{ctx.slow_movers_with_ba} ASIN(s) with fewer than 3 orders in the period have BA spend.{suffix} '
            f'Remove slow movers from BA and redirect spend to best sellers.'
        )

    # S038 — BAK harvest layer missing while BA > 30% of spend
    if 'S038' in flags:
        texts['S038'] = (
            f'BA campaigns account for {pct(ctx.pct_ba)} of total spend ({dollar(ctx.spend_ba)}). '
            f'No BAK harvest layer exists. '
            f'Discovery data from BA is not being converted into manual exact match campaigns. '
            f'Launch BAK campaigns to capture proven search terms from BA.'
        )

    # S039 — BA not segmented by category (< 2 campaigns)
    if 'S039' in flags:
        texts['S039'] = (
            f'Only {ctx.ba_campaign_count} BA campaign(s) detected. '
            f'Structure is not segmented by category — new BA campaigns by category are needed.'
        )

    # S041 — Low-order campaign consolidation needed
    if 'S041' in flags:
        severity = 'severe fragmentation' if ctx.low_order_campaign_count > 80 else 'high fragmentation'
        texts['S041'] = (
            f'{ctx.low_order_campaign_count} campaigns have only 1–3 orders in the period ({severity}). '
            f'Consolidate converting terms into BAK campaigns by parent ASIN. '
            f'Remove or archive campaigns with persistent low volume.'
        )

    # S042 — SB outperforming by >20%
    if 'S042' in flags:
        texts['S042'] = (
            f'SB campaigns avg ACoS: {pct(ctx.sb_avg_acos)} vs account avg {pct(ctx.acos_actual)}. '
            f'Sponsored Brands is outperforming — consider increasing SB investment or adding new campaigns.'
        )

    # S043 — SBV outperforming by >20%
    if 'S043' in flags:
        texts['S043'] = (
            f'SBV campaigns avg ACoS: {pct(ctx.sbv_avg_acos)} vs account avg {pct(ctx.acos_actual)}. '
            f'Sponsored Brands Video is outperforming — consider expanding SBV category targets.'
        )

    # S045 — SB missing, base is built
    if 'S045' in flags:
        acos_pp = ctx.acos_actual * 100
        declining = ctx.yoy_ad_sales < -0.05
        acos_high = ctx.acos_constraint > 0 and acos_pp > ctx.acos_constraint * 1.2
        prefix = 'No Sponsored Brands spend detected. '
        prefix += f'SBV is active ({dollar(ctx.spend_sbv)}) but SB is absent. ' if ctx.spend_sbv > 0 else ''
        if declining:
            suffix = f'Ad sales down {pct(abs(ctx.yoy_ad_sales))} YoY — SB is a direct lever for upper-funnel recovery.'
        elif acos_high:
            suffix = f'ACoS is {acos_pp:.0f}% vs {ctx.acos_constraint:.0f}% constraint — address efficiency before launching SB.'
        else:
            suffix = 'SB campaigns should be launched to build upper-funnel coverage.'
        texts['S045'] = prefix + suffix

    # S047 — Import kickoff needed
    if 'S047' in flags:
        texts['S047'] = (
            f'Imported campaigns: {dollar(ctx.spend_imported)} ({pct(ctx.pct_imported)} of spend). '
            f'These run outside the Quartile system. An import kickoff CoE ticket is needed.'
        )

    # S048 — BAK harvest stalled (BAK < 10% of BA spend)
    if 'S048' in flags:
        texts['S048'] = (
            f'BAK spend is {pct(ctx.pct_bak)} vs BA spend at {pct(ctx.pct_ba)}. '
            f'BAK exists but is receiving less than 10% of its BA feeder spend. '
            f'The harvest cycle has stalled — review BA search term report and promote converting terms to BAK.'
        )

    # S049 — Own product page undefended (BAK exists, no OP)
    if 'S049' in flags:
        texts['S049'] = (
            f'BAK represents {pct(ctx.pct_bak)} of total spend ({dollar(ctx.spend_bak)}) '
            f'but no OP (own product targeting) campaigns are active. '
            f'Own product pages are undefended — competitors can place ads on your listings.'
        )

    # S050 — BR discovery layer missing (BAK exists, no BR feeder)
    if 'S050' in flags:
        texts['S050'] = (
            f'No BR (broad match research) campaigns detected. '
            f'BAK has {pct(ctx.pct_bak)} of spend but no broad match discovery feeder. '
            f'The keyword harvest is static — no new terms are being tested.'
        )

    # S053 — Inefficient BAK campaigns
    if 'S053' in flags:
        texts['S053'] = (
            f'{ctx.inefficient_bak_count} BAK campaign(s) have spend >$200, ACoS >{ctx.acos_constraint * 1.5:.0f}%, '
            f'and fewer than 5 orders in the period. '
            f'Keyword focus should be narrowed — consolidate to the top-performing terms and pause low-conversion keywords.'
        )

    # S054 — ATM outperforming by >20%
    if 'S054' in flags:
        texts['S054'] = (
            f'ATM (automatic targeting) campaigns avg ACoS: {pct(ctx.atm_avg_acos)} vs account avg {pct(ctx.acos_actual)}. '
            f'Automatic targeting is outperforming — consider expanding ATM coverage across more ASINs.'
        )

    # S055 — BR campaigns outperforming by >20%
    if 'S055' in flags:
        texts['S055'] = (
            f'BR (broad match) campaigns avg ACoS: {pct(ctx.br_avg_acos)} vs account avg {pct(ctx.acos_actual)}. '
            f'Broad match targeting is outperforming — consider expanding BR coverage or launching new BR campaigns.'
        )

    # S056 — PH_ (phrase match) campaigns outperforming by >20%
    if 'S056' in flags:
        texts['S056'] = (
            f'PH (phrase match) campaigns avg ACoS: {pct(ctx.ph_avg_acos)} vs account avg {pct(ctx.acos_actual)}. '
            f'Phrase match targeting is outperforming — consider expanding PH coverage or launching new PH campaigns.'
        )

    # S057 — Keyword strategy too broad (BR >15% at poor efficiency)
    if 'S057' in flags:
        texts['S057'] = (
            f'BR broad match campaigns represent {pct(ctx.pct_br)} of total spend '
            f'at {pct(ctx.br_avg_acos)} avg ACoS vs {ctx.acos_constraint:.0f}% constraint. '
            f'Broad match is consuming significant budget at poor efficiency. '
            f'Narrow keyword targeting to proven high-intent terms and reduce broad match coverage.'
        )

    # S058 — OP campaigns outperforming account average by >20%
    if 'S058' in flags:
        texts['S058'] = (
            f'OP (product targeting) campaigns avg ACoS: {pct(ctx.op_avg_acos)} vs account avg {pct(ctx.acos_actual)}. '
            f'Product-targeting campaigns are outperforming — consider expanding OP coverage.'
        )

    # S060 — SD_FLEX campaigns outperforming account average by >20%
    if 'S060' in flags:
        texts['S060'] = (
            f'SD_FLEX campaigns avg ACoS: {pct(ctx.sd_flex_avg_acos)} vs account avg {pct(ctx.acos_actual)}. '
            f'SD_FLEX remarketing is outperforming — consider expanding SD_FLEX audience coverage.'
        )

    # S062 — No CAT_SP, OP outperforming or no OP at all
    if 'S062' in flags:
        op_note = ''
        if ctx.op_avg_acos > 0 and ctx.acos_actual > 0:
            op_note = (
                f'OP campaigns avg ACoS: {ctx.op_avg_acos:.0%} vs account avg {ctx.acos_actual:.0%} '
                f'({(1 - ctx.op_avg_acos / ctx.acos_actual):.0%} more efficient). '
            )
        texts['S062'] = (
            f'No CAT_SP campaigns detected. {op_note}'
            f'Category-targeted SP campaigns should be launched for key product categories.'
        )

    # S063 — OP campaigns outperforming keyword campaigns
    if 'S063' in flags:
        texts['S063'] = (
            f'OP (product targeting) campaigns avg ACoS: {pct(ctx.op_avg_acos)} vs account avg {pct(ctx.acos_actual)}. '
            f'Product-targeting is outperforming keyword campaigns. '
            f'Expand OP coverage and consider launching SD_PRD to strengthen product-page presence.'
        )

    # S064 — Paused SB campaigns with historical spend
    if 'S064' in flags:
        texts['S064'] = (
            f'{ctx.paused_sb_count} SB campaign(s) are paused with historical spend. '
            f'Current SB spend: {dollar(ctx.spend_sb)}. '
            f'Paused SB campaigns should be reviewed and rebuilt with updated keyword structures.'
        )

    # S065 — SD_FLEX campaigns outperforming by >20%
    if 'S065' in flags:
        texts['S065'] = (
            f'SD_FLEX campaigns avg ACoS: {pct(ctx.sd_flex_avg_acos)} vs account avg {pct(ctx.acos_actual)}. '
            f'SD_FLEX remarketing is outperforming — consider expanding SD_FLEX coverage with new audiences.'
        )

    # S066 — Paused SBV campaigns with historical spend
    if 'S066' in flags:
        texts['S066'] = (
            f'{ctx.paused_sbv_count} SBV campaign(s) are paused with historical spend. '
            f'SB is active ({dollar(ctx.spend_sb)}). '
            f'SBV should run alongside SB to capture video inventory — evaluate reactivation.'
        )

    # S067 — CAT_SP launch opportunity (no CAT_SP, OP works or no OP)
    if 'S067' in flags:
        texts['S067'] = (
            f'No CAT_SP campaigns detected. Total spend: {dollar(ctx.total_spend)}. '
            + (f'OP campaigns avg ACoS: {pct(ctx.op_avg_acos)} vs account avg {pct(ctx.acos_actual)} — '
               f'product targeting is working. ' if ctx.op_avg_acos > 0 else '')
            + f'Launch CAT_SP campaigns for categories with ≥30 ASINs and ≥5% of sales.'
        )

    # S068 — SBV product targeting launch (SBV missing, SB well established)
    if 'S068' in flags:
        texts['S068'] = (
            f'No SBV (Sponsored Brands Video) campaigns detected. '
            f'SB is active ({dollar(ctx.spend_sb)}, {pct(ctx.pct_sb)} of spend). '
            f'SBV product targeting is the natural next step when SB is established. '
            f'Launch SBV campaigns on the same category targets as existing SB.'
        )

    # S069 — Broad Match Graduation Signal: BR outperforms OW
    if 'S069' in flags:
        texts['S069'] = (
            f'BR (broad match) campaigns avg ACoS: {pct(ctx.br_avg_acos)} — '
            f'outperforming OW (exact match) at {pct(ctx.ow_avg_acos)}. '
            f'{ctx.br_campaign_count} BR campaigns vs {ctx.ow_campaign_count} OW campaigns. '
            f'Broad match is proving more efficient — graduate proven BR learnings into phrase match (PH_) campaigns.'
        )

    # S071 — Exact Match Graduation Signal: OW outperforms BR
    if 'S071' in flags:
        texts['S071'] = (
            f'OW (exact match) campaigns avg ACoS: {pct(ctx.ow_avg_acos)} — '
            f'outperforming BR (broad match) at {pct(ctx.br_avg_acos)}. '
            f'{ctx.ow_campaign_count} OW campaigns vs {ctx.br_campaign_count} BR campaigns. '
            f'Exact match is more efficient — graduate proven BR learnings into OW exact match campaigns.'
        )

    # S072 — OP target expansion: many keyword campaigns, few OP
    if 'S072' in flags:
        kw_total = ctx.br_campaign_count + ctx.ow_campaign_count + ctx.ph_campaign_count
        texts['S072'] = (
            f'{ctx.op_campaign_count} OP (product target) campaign(s) vs {kw_total} keyword campaigns (OW + BR + PH). '
            f'Product-targeting coverage is thin relative to keyword volume. '
            f'Expand OP_ campaigns using top ASINs as targets — product-page coverage is underdeveloped.'
        )

    # S073 — CatchAll keyword graduation (high orders in CatchAll, BAK thin)
    if 'S073' in flags:
        texts['S073'] = (
            f'CatchAll campaigns generated {ctx.catchall_orders:.0f} orders in the period. '
            f'BAK (manual exact) is only {pct(ctx.pct_bak)} of total spend. '
            f'High-converting search terms in CatchAll should be graduated to BAK campaigns for better bid control.'
        )

    # S074 — CatchAll graduation overdue (at scale, BAK still thin)
    if 'S074' in flags:
        texts['S074'] = (
            f'CatchAll campaigns have {ctx.catchall_orders:.0f} orders but BAK is only '
            f'{pct(ctx.pct_bak)} of total spend. '
            f'High-converting search terms should be graduated to BAK campaigns '
            f'for tighter bid control and better efficiency.'
        )

    # S075 — CAT_SP avg ACoS above constraint
    if 'S075' in flags:
        texts['S075'] = (
            f'CAT_SP campaigns avg ACoS: {pct(ctx.catsp_avg_acos)} vs constraint {constraint:.0f}%. '
            f'Category-targeted SP campaigns are above the ACoS threshold. '
            f'Review targeting scope and remove underperforming category nodes.'
        )

    # S077 — Multiple WATM campaigns without structural need
    if 'S077' in flags:
        texts['S077'] = (
            f'{ctx.watm_campaign_count} WATM campaigns are active. '
            f'WATM spend: {pct(ctx.pct_watm)} of total ({dollar(ctx.spend_watm)}). '
            f'Multiple WATM campaigns add fragmentation. '
            f'One WATM per account is standard unless portfolio-level separation is intentional.'
        )

    # S080 — CAT_SP active but no qualifying categories
    if 'S080' in flags:
        texts['S080'] = (
            f'CAT_SP campaigns are active but no product category qualifies under the CoE threshold '
            f'(AsinCount ≥ 30 AND category share ≥ 5% of sales). '
            f'CAT_SP may be targeting categories that are too narrow or not contributing meaningfully to sales.'
        )

    # S081 — BAK branded and non-branded mixed
    if 'S081' in flags:
        texts['S081'] = (
            f'Branded search terms represent {ctx.branded_spend_pct:.0%} of spend and '
            f'non-branded {ctx.non_branded_spend_pct:.0%} — both significant. '
            f'These are likely in the same BAK bucket. '
            f'Branded and non-branded terms require separate bid strategies and should be in separate campaigns.'
        )

    # S082 — WATM / CatchAll catalogue coverage below 60%
    if 'S082' in flags:
        spend_cov = (ctx.spending_asin_count / ctx.catalog_asin_count) if ctx.catalog_asin_count > 0 else 0
        texts['S082'] = (
            f'WATM or CatchAll is active but only {spend_cov:.0%} of the catalogue '
            f'({ctx.spending_asin_count} of {ctx.catalog_asin_count} ASINs) has ad spend. '
            f'Coverage is below 60%. Some ASINs are missing from current campaign structures.'
        )

    # S083 — WATM and CatchAll both active simultaneously (redundant)
    if 'S083' in flags:
        texts['S083'] = (
            f'{ctx.watm_campaign_count} WATM campaign(s) and at least 1 CatchAll campaign are both active. '
            f'WATM and CatchAll serve the same coverage purpose. '
            f'Running both creates redundant overlap — review which structure is preferred and deactivate the other.'
        )

    # S084 — WATM spend underweighted (< 3% of total)
    if 'S084' in flags:
        texts['S084'] = (
            f'WATM campaigns exist ({ctx.watm_campaign_count} active) but represent only {pct(ctx.pct_watm)} of total spend. '
            f'WATM is not receiving enough budget to serve its coverage purpose. '
            f'Review WATM budgets and bid levels.'
        )

    # S085 — BAK high-spend with ACoS above constraint
    if 'S085' in flags:
        over = [b for b in ctx.bak_campaigns if b['pct_of_total'] > 0.15 and b['acos'] > (constraint / 100) * 0.50]
        camp_lines = '; '.join(
            f"{b['name']} ({pct(b['pct_of_total'])}, {b['acos']:.0%} ACoS)"
            for b in over[:3]
        )
        texts['S085'] = (
            f'{len(over)} BAK campaign(s) exceed 15% of total spend with ACoS above the constraint threshold. '
            + (f'Campaigns: {camp_lines}. ' if camp_lines else '')
            + f'Add negatives for wasteful keywords and review top spend terms.'
        )

    # S088 — No SD, account is ready for remarketing
    if 'S088' in flags:
        texts['S088'] = (
            f'No Sponsored Display campaigns active. SD spend: $0. Total spend: {dollar(ctx.total_spend)}. '
            f'The account has sufficient scale ({ctx.max_asin_orders_30d:.0f} orders on top ASIN) for SD retargeting. '
            f'Product-view and ATC remarketing should be launched.'
        )

    # S089 — ProSuite active, no ATC retargeting deployed
    if 'S089' in flags:
        texts['S089'] = (
            f'SD is active ({pct(ctx.pct_sd)} of spend) and ProSuite audiences are available. '
            f'No ATC (add-to-cart) retargeting campaigns detected. '
            f'SD_FLEX_ATC should be deployed to retarget high-intent shoppers who added to cart but did not purchase.'
        )

    # S092 — OW own-page coverage missing
    if 'S092' in flags:
        texts['S092'] = (
            f'OP product-targeting campaigns are active but no OW own-waterfall auto campaigns found. '
            f'Own listing pages have OP coverage but no auto-targeting restricted to own pages.'
        )

    # S093 — SD suggested but PDP maturity too low
    if 'S093' in flags:
        texts['S093'] = (
            f'SD expansion signal is present but top-selling ASIN has only '
            f'{ctx.max_asin_orders_30d:.0f} orders in the period. '
            f'SD retargeting audiences require sufficient product-view traffic to be effective. '
            f'Consider waiting until top ASINs reach ≥50 orders/month before launching SD.'
        )

    # S094 — Portfolio governance: no managed portfolios or budget caps
    if 'S094' in flags:
        texts['S094'] = (
            f'{ctx.portfolio_count} portfolios active. '
            f'{ctx.managed_portfolio_count} managed. '
            f'{ctx.portfolios_with_budget_cap} have budget caps. '
            f'Portfolio governance needs to be tightened — apply budget caps and enable management.'
        )

    # S095 — Campaign-level ACoS overrides active while above constraint
    if 'S095' in flags:
        texts['S095'] = (
            f'Campaign-level ACoS overrides are active while ACoS is above constraint ({ctx.acos_actual:.0%} vs {ctx.acos_constraint:.0f}%). '
            f'Each override should be intentional and documented — review for stale or unintended overrides.'
        )

    # S096 — Product-level ACoS overrides active while above constraint
    if 'S096' in flags:
        texts['S096'] = (
            f'Product-level ACoS overrides are active while account ACoS is above constraint. '
            f'Review product overrides and confirm each is intentional.'
        )

    # S097 — VCPM overuse on products without Buy Box
    if 'S097' in flags:
        texts['S097'] = (
            f'VCPM spend represents {pct(ctx.vcpm_spend_pct)} of total SD spend. '
            f'VCPM on products without consistent Buy Box ownership wastes impressions.'
        )

    # S098 — Tagging and segmentation gap (BA+SPT+ATM, no OP)
    if 'S098' in flags:
        texts['S098'] = (
            f'BA ({dollar(ctx.spend_ba)}), SPT ({dollar(ctx.spend_spt)}), ATM ({dollar(ctx.spend_atm)}) all active '
            f'but no OP product-target campaigns detected. Product-page coverage is missing.'
        )

    # S100 — SBV naming convention not followed
    if 'S100' in flags:
        texts['S100'] = (
            f'SBV campaigns active but not all follow the SBV_ naming convention. '
            f'Non-standard naming reduces governance clarity and reporting consistency.'
        )

    # S101 — Campaigns outside portfolio
    if 'S101' in flags:
        texts['S101'] = (
            f'{ctx.campaigns_not_in_portfolio} campaign(s) not assigned to any portfolio. '
            f'All active campaigns should be assigned to a portfolio for governance and reporting.'
        )

    # S102 — Best-seller visibility gaps
    if 'S102' in flags:
        texts['S102'] = (
            f'{ctx.top_seller_type_gaps} of {ctx.tier1_asin_count} top-selling ASIN(s) (Tier 10–30) '
            f'are missing ≥2 of the required campaign types (ATM, BAK, OP). '
            f'Best-seller visibility is incomplete — launch missing campaign types for these ASINs.'
        )

    # S104 — SB active but SBV missing (historical / manual flag)
    if 'S104' in flags:
        texts['S104'] = (
            f'SB active ({ctx.sb_impressions:,} impressions) but SBV spend is $0. '
            f'Launch SBV product-targeting and branded campaigns.'
        )

    # S105 — Proven targets (top 30 search terms) not utilised in BAK
    if 'S105' in flags:
        texts['S105'] = (
            f'{ctx.unconverted_top_terms} high-converting search term(s) in the top 30 '
            f'(≥3 orders, ≥10% CVR) are not captured in BAK campaigns. '
            f'BAK is only {pct(ctx.pct_bak)} of total spend. '
            f'These proven terms should be promoted to exact match BAK campaigns.'
        )

    # S106 — ProSuite AMC audiences not deployed
    if 'S106' in flags:
        texts['S106'] = (
            f'{ctx.total_campaign_count} campaigns active but no ProSuite AMC audiences applied. '
            f'The account has sufficient scale for audience-based targeting. '
            f'Test Amazon native audiences (NTB, ATC, SNS) via ProSuite AMC.'
        )

    # S107 — Inefficient ASIN spend (high spend, high ACoS, low orders)
    if 'S107' in flags:
        texts['S107'] = (
            f'{ctx.inefficient_asin_count} ASIN(s) with spend >$100, '
            f'ACoS >{ctx.acos_constraint * 1.5:.0f}%, and fewer than 5 orders in the period. '
            f'Spend on these ASINs is not converting — reduce or reallocate to top performers.'
        )

    # S108 — SB active but SBV missing (gated: SB well established, > 5% of spend)
    if 'S108' in flags:
        texts['S108'] = (
            f'SB is active ({ctx.sb_impressions:,} impressions, {pct(ctx.pct_sb)} of spend) '
            f'but SBV spend is $0. '
            f'SBV is the natural next step when SB is established — launch video campaigns '
            f'on the same category targets as existing SB.'
        )

    # S111 — No recurring sales strategy, sales flat or declining
    if 'S111' in flags:
        yoy_note = f'Ad sales declined {abs(ctx.yoy_ad_sales):.0%} YoY. ' if ctx.yoy_ad_sales < 0 else 'Ad sales are flat YoY. '
        texts['S111'] = (
            yoy_note
            + f'No recurring sales strategy (SnS) is active. '
            + f'For repurchasable products, a recurring purchase strategy should be reviewed with the client.'
        )

    # S112 — Sales declining while spend is growing
    if 'S112' in flags:
        texts['S112'] = (
            f'Ad sales declined {abs(ctx.yoy_ad_sales):.0%} YoY while spend increased {ctx.mom_spend_change:.0%} MoM. '
            f'More budget going in, less revenue coming out. '
            f'Budget levels and campaign scope must be reviewed before the next cycle.'
        )

    # S113 — Budget constraint alignment deviation
    if 'S113' in flags:
        if ctx.monthly_budget > 0:
            deviation = abs(ctx.total_spend - ctx.monthly_budget) / ctx.monthly_budget
            texts['S113'] = (
                f'Documented monthly budget: {dollar(ctx.monthly_budget)}. '
                f'Actual spend: {dollar(ctx.total_spend)} — a {deviation:.0%} deviation. '
                f'Budget delivery should be reviewed and aligned with the client.'
            )
        else:
            texts['S113'] = (
                f'No monthly budget is documented in Salesforce. '
                f'Budget constraint alignment cannot be validated automatically.'
            )

    # S117 — Subscribe & Save not active
    if 'S117' in flags:
        texts['S117'] = (
            f'Subscribe & Save is not active. '
            + (f'YoY ad sales: {ctx.yoy_ad_sales:+.0%}. ' if ctx.yoy_ad_sales != 0 else '')
            + f'For repurchasable products, SnS should be evaluated as a retention and growth lever.'
        )

    # S120 — Promo portfolio budget pacing issue
    if 'S120' in flags:
        texts['S120'] = (
            f'{ctx.promo_asin_count} ASIN(s) in active promo. '
            + (f'Promo cost rate averaging {pct(ctx.promo_cost_rate)}. ' if ctx.promo_cost_rate > 0 else '')
            + f'Portfolio budgets should be reviewed to prevent intraday depletion.'
        )

    # S122 — GGS committed but SD < 5% of spend
    if 'S122' in flags:
        sd_note = (
            f'SD spend: {dollar(ctx.spend_sd)} ({ctx.pct_sd:.0%} of total). '
            if ctx.spend_sd > 0 else 'SD spend: $0. '
        )
        texts['S122'] = (
            f'GGS status: {ctx.ggs_status}. {sd_note}'
            f'SD campaigns need to reach at least 5% of total spend to satisfy the GGS commitment.'
        )

    # S123 — SD active but no remarketing (SD_FLEX / SD_AUDI)
    if 'S123' in flags:
        texts['S123'] = (
            f'SD active ({dollar(ctx.spend_sd)}) but no SD_FLEX or SD_AUDI remarketing campaigns. '
            f'Product-view remarketing is not running.'
        )

    # S124 — SD active but no ATC retargeting (GGS section duplicate)
    if 'S124' in flags:
        texts['S124'] = (
            f'SD active ({dollar(ctx.spend_sd)}) but no ATC retargeting in place. '
            f'Add-to-cart retargeting via ProSuite AMC is not activated.'
        )

    # S126 — Branded + NB both > 20% in BA auto layer (promo not relevant here)
    if 'S126' in flags:
        texts['S126'] = (
            f'Branded spend is {ctx.branded_spend_pct:.0%} and non-branded is {ctx.non_branded_spend_pct:.0%} '
            f'of total search term spend — both significant inside the same auto campaign layer. '
            f'Branded and non-branded targeting should be split into separate campaigns for independent bid control.'
        )

    return texts


# ── grade calculator ──────────────────────────────────────────────────────────

def _calculate_grade(flags: dict[str, str]) -> tuple[str, str]:
    """Returns (grade_label, interpretation_text)."""
    n_flag    = sum(1 for v in flags.values() if v == 'FLAG')
    n_partial = sum(1 for v in flags.values() if v == 'PARTIAL')

    if n_flag == 0 and n_partial == 0:
        grade = 'Compliant'
        interp = (
            'The account reflects a well-defined strategic direction with no major gaps identified. '
            'Few or no changes are required — the current campaign structure, targeting approach, '
            'and client alignment are consistent with the account\'s objectives and roadmap.'
        )
    elif n_flag == 0 and n_partial <= 3:
        grade = 'Needs Review'
        interp = (
            f'The account has {n_partial} area(s) that require attention. '
            'No critical gaps were found, but several strategic items should be reviewed '
            'before the next client interaction to avoid them becoming larger issues.'
        )
    elif n_flag <= 2:
        grade = 'Needs Improvement'
        interp = (
            f'The account has {n_flag} critical gap(s) and {n_partial} item(s) needing attention. '
            'Action is required. Review the flagged controls below and align with the client '
            'or internal team on a clear plan before the next review cycle.'
        )
    else:
        grade = 'Non-Compliant'
        interp = (
            f'The account has {n_flag} critical strategic gaps. '
            'Significant structural or strategic work is required. '
            'Prioritise the flagged controls and escalate where client alignment is needed.'
        )

    return grade, interp


# ── main writer ───────────────────────────────────────────────────────────────

def write_strategy(pre_analysis_path: str, template_path: str, output_dir: str) -> str:

    # ── read context ─────────────────────────────────────────────────────────
    ctx = read_strategy_context(pre_analysis_path)

    # ── also read raw tabs for Questionnaire + ChildASIN ─────────────────────
    pa = openpyxl.load_workbook(pre_analysis_path, data_only=True, read_only=True)

    d55    = _tab_to_dict(pa['55_Salesforce_Consolidated_PreA'])
    d38    = _latest_record(_tab_to_records_full(pa['38_Client_Success_Insights_Repo']))
    gong_r = _tab_to_records(pa['37_Gong_Call_Insights_for_Sales'])
    gong   = gong_r[0] if gong_r else {}

    asin_records = _tab_to_records(pa['14_Campaign_Performance_by_Adve'])
    cat_records  = _tab_to_records(pa['22_Catalogue_Details'])
    cat_by_asin  = {r['asin']: r for r in cat_records if r.get('asin')}

    _d54_all      = _tab_to_records_full(pa['54_Project_Dataset_on_SF'])
    _d54_filtered = _filter_by_advertiser(_d54_all, ctx.profile_id)
    d54           = _latest_record(_d54_filtered)

    pa.close()

    # ── compute auto-flags ───────────────────────────────────────────────────
    flags = _compute_flags(ctx)
    grade, interp = _calculate_grade(flags)

    # ── load template ────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(template_path, keep_vba=True)

    # ════════════════════════════════════════════════════════════════════════════
    # TAB 1 — Questionnaire Survey - AMZ
    # ════════════════════════════════════════════════════════════════════════════
    ws1 = wb['Questionaire Survey - AMZ']

    def w1(coord, value):
        ws1[coord] = value

    w1('C6', ctx.member_id)
    w1('F6', ctx.profile_id)
    w1('J6', _safe(d55.get('CSP_Last_Modified_By')))

    w1('F7', ctx.profile_id)
    w1('J7', _safe(d55.get('Projected_Project_MRR__c')))

    w1('C8', _safe(d55.get('Account_Name')))
    ld = d55.get('Launch_Date__c')
    w1('F8', ld.strftime('%Y-%m-%d') if hasattr(ld, 'strftime') else _safe(ld))
    if hasattr(ld, 'strftime'):
        months = (datetime.now().year - ld.year) * 12 + (datetime.now().month - ld.month)
        w1('J8', f'{months} months')
    else:
        w1('J8', _safe(d38.get('Customer_Age_Months__c')))

    w1('C9', _safe(d55.get('Customer_Age_Months__c') or d38.get('Customer_Age_Months__c')))
    w1('F9', _safe(d38.get('Repeat_Purchase_Behavior__c')))
    w1('J9', _safe(d55.get('CSM_Churn_Risk__c')))

    w1('C10', _safe(d38.get('Commodity_Products_or_Branded_Products__c')))
    w1('F10', _safe(d55.get('Vertical__c')))
    w1('J10', _safe(d55.get('Contract_Term__c')))

    w1('C11', _safe(d38.get('Average_Order_Value__c') or d54.get('Average_Order_Value__c')))
    w1('F11', _safe(d55.get('Services_Sold__c')))
    w1('J11', _safe(d55.get('MRR__c')))

    # Row 12 — Director SF user ID not readable; write Active Products
    w1('C12', '')
    w1('F12', _safe(d55.get('Active_Products__c')))

    # Row 13
    w1('F13', _safe(d38.get('Customer_Feedback__c')))

    # Row 15
    w1('C15', _safe(d55.get('Current_Challenges__c')))
    w1('F15', _safe(d55.get('Primary_Objective__c')))
    w1('J15', _safe(d55.get('ACOS_Constraint__c')))

    # Row 16
    w1('C16', _safe(d55.get('Primary_Objective_Additional_Context__c')))
    w1('F16', _safe(d55.get('Primary_Spend_KPI__c')))
    w1('J16', _safe(d38.get('Customer_Acquisition_Cost_Target__c')))

    # Row 17
    w1('C17', _safe(d55.get('Top_Priority__c')))
    w1('J17', _safe(d55.get('TACOS_Constraint__c')))

    # Row 18
    w1('C18', _safe(d55.get('Second_Priority__c')))
    w1('F18', _safe(d54.get('CS_Notes__c')))
    w1('J18', _safe(d55.get('daily_target_spend__c')))

    # Row 19
    w1('C19', _safe(d55.get('Biggest_Expansion_Opportunity__c')))
    w1('F19', _safe(d55.get('Near_Term_3_Month_Considerations__c')))
    w1('J19', _safe(d55.get('Target_ROAS__c')))


    # CJM Salesforce strategy steps — fixed row positions in template
    stage_rows = {1: (24, 25), 2: (27, 28), 3: (30, 31), 4: (33, 34)}
    for s, (r_a, r_i) in stage_rows.items():
        w1(f'C{r_a}', _safe(d55.get(f'AdoptionOrUpsellS{s}__c')))
        w1(f'G{r_a}', _safe(d55.get(f'StrategyS{s}__c')))
        w1(f'J{r_a}', _safe(d55.get(f'StatusS{s}__c')))
        intro = d55.get(f'ExecutionDateS{s}__c')
        w1(f'C{r_i}', intro.strftime('%Y-%m-%d') if hasattr(intro, 'strftime') else _safe(intro))


    # Gong
    w1('C41', _safe(gong.get('Gong__Call_Brief__c') or d55.get('Call_Brief')))
    w1('C42', _safe(gong.get('Gong__Call_Key_Points__c') or d55.get('Key_Points')))
    w1('C43', _safe(gong.get('Gong__Call_Highlights_Next_Steps__c') or d55.get('Highlights_Next_Steps')))

    # ════════════════════════════════════════════════════════════════════════════
    # TAB — STRATEGY OVERVIEW (original tab — col G = Flag, col C = What We Saw)
    # Rows match _SID_TO_ROW exactly (both tabs share the same row numbers).
    # ════════════════════════════════════════════════════════════════════════════
    ws_old = wb['STRATEGY OVERVIEW']
    dynamic_what = _build_what_we_saw(ctx, flags)

    # Reset all AUTO rows to OK before writing — prevents stale template values
    for row_idx in range(2, 135):
        if ws_old.cell(row=row_idx, column=6).value == 'AUTO':
            ws_old.cell(row=row_idx, column=7, value='OK')

    for sid, level in flags.items():
        row_num = _SID_TO_ROW.get(sid)
        if row_num:
            ws_old.cell(row=row_num, column=7, value=level)   # col G = Flag

    for sid, text in dynamic_what.items():
        row_num = _SID_TO_ROW.get(sid)
        if row_num:
            ws_old.cell(row=row_num, column=3, value=text)    # col C = What We Saw

    # ════════════════════════════════════════════════════════════════════════════
    # TAB — New Strategy Overview
    # Col 5  (E) = Auto Review (AUTO/MANUAL — static in template, not overwritten)
    # Col 6  (F) = STATUS written by agent: FLAG / PARTIAL / OK
    # Col 10 (J) = What We Saw — dynamic text built from real account numbers
    # ════════════════════════════════════════════════════════════════════════════
    ws_ov = wb['New Strategy Overview']

    # Reset all AUTO rows to OK before writing — prevents stale template values
    for row_idx in range(2, 127):
        if ws_ov.cell(row=row_idx, column=5).value == 'AUTO':
            ws_ov.cell(row=row_idx, column=6, value='OK')

    # Write STATUS (col 6) for every control that fired
    for sid, level in flags.items():
        row_num = _SID_TO_ROW.get(sid)
        if row_num:
            ws_ov.cell(row=row_num, column=6, value=level)

    # Reuse dynamic_what already computed above
    for sid, text in dynamic_what.items():
        row_num = _SID_TO_ROW.get(sid)
        if row_num:
            ws_ov.cell(row=row_num, column=10, value=text)

    # ════════════════════════════════════════════════════════════════════════════
    # TAB — Account Strategy _Analysis
    # Header writes only — findings population stays manual per design decision
    # ════════════════════════════════════════════════════════════════════════════
    ws2 = wb['Account Strategy _Analysis']
    ws2['A1'] = f'{ctx.account_label} — Account Strategy Analysis'
    ws2['B3'] = f'Account: {ctx.account_label} | Tenant ID: {ctx.tenant_id} | Account ID: {ctx.profile_id}'
    ws2['B4'] = ctx.date_range
    ws2['B5'] = ctx.downloaded

    # TAB 3 — ChildASIN View
    # ════════════════════════════════════════════════════════════════════════════
    ws3 = wb['ChildASIN View']

    import math as _math

    # Build column → index map from header row 2
    header_row2 = list(ws3.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    col_idx = {str(h).strip(): i for i, h in enumerate(header_row2, 1) if h}

    # col14_map: template header label → field key in asin_records (tab 14)
    col14_map = {
        'Parent ASIN':       'ParentASIN',
        'ASIN':              'asin',
        'Total Sales':       'TotalSales',
        'Ad Spend':          'AdSpend',
        'TACoS':             'TACoS',
        'Ad Sales':          'AdSales',
        'Ads Units Ordered': 'Orders',
        'ACoS':              'ACoS',
        'Clicks':            'Clicks',
        'Tier':              'Tier',
        'ATM_Spend':         'ATM_Spend',
        'BA_Spend':          'BA_Spend',
        'Manual_Q1_Spend':   'Manual_Q1_Spend',
        'BAK_Spend':         'BAK_Spend',
        'OP_Spend':          'OP_Spend',
        'SPT_Spend':         'SPT_Spend',
        'CAT_SP_Spend':      'CAT_SP_Spend',
        'WATM_Spend':        'WATM_Spend',
        'SB_Spend':          'SB_Spend',
        'SBV_Spend':         'SBV_Spend',
        'SD_Spend':          'SD_Spend',
        'Imported_Spend':    'Imported_Spend',
        'NonQuartile_Spend': 'NonQuartile_Spend',
        'AOV':               'AOV',
        'TAG 1':             'Tag1',
        'TAG 2':             'Tag2',
        'TAG 3':             'Tag3',
        'TAG 4':             'Tag4',
        'TAG 5':             'Tag5',
    }

    # col22_map: template header → field key in cat_by_asin (tab 22)
    col22_map = {
        'PriceTier':  'PriceTier',
        'Brand':      'Brand',
        'Department': 'Department',
        'Category':   'Category',
    }

    data_row = 3
    for row_idx, rec in enumerate(asin_records, data_row):
        asin = _safe(rec.get('asin', ''))
        cat  = cat_by_asin.get(asin, {})

        for label, src in col14_map.items():
            ci = col_idx.get(label)
            if ci:
                ws3.cell(row=row_idx, column=ci, value=rec.get(src))

        for label, src in col22_map.items():
            ci = col_idx.get(label)
            if ci:
                ws3.cell(row=row_idx, column=ci, value=cat.get(src))

        # Computed columns
        def _sf(v):
            try: return float(v) if v is not None else 0.0
            except: return 0.0
        tot_s = _sf(rec.get('TotalSales'))
        ad_s  = _sf(rec.get('AdSales'))
        aov   = _sf(rec.get('AOV'))

        ci_units = col_idx.get('Total Units Ordered')
        if ci_units and aov > 0:
            ws3.cell(row=row_idx, column=ci_units, value=_math.ceil(tot_s / aov))

        ci_adsales_pct = col_idx.get('Ad Sales (%)')
        if ci_adsales_pct and tot_s > 0:
            ws3.cell(row=row_idx, column=ci_adsales_pct, value=round(ad_s / tot_s, 4))

        ci_org_pct = col_idx.get('Organic Sales (%)')
        if ci_org_pct and tot_s > 0:
            ws3.cell(row=row_idx, column=ci_org_pct, value=round(1 - (ad_s / tot_s), 4))

        ci_bb = col_idx.get('Buy Box%')
        if ci_bb:
            raw_bb = rec.get('Weighted_BuyBoxPercentage')
            if raw_bb is not None:
                ws3.cell(row=row_idx, column=ci_bb, value=round(_sf(raw_bb) / 100, 4))

        # Quartile One / Quartile Bulk formula columns (28/29)
        ci_q1 = col_idx.get('Quartile One')
        ci_qb = col_idx.get('Quartile Bulk')
        if ci_q1:
            ws3.cell(row=row_idx, column=ci_q1,
                     value=f'=SUM(O{row_idx}+Q{row_idx}+S{row_idx})')
        if ci_qb:
            ws3.cell(row=row_idx, column=ci_qb,
                     value=f'=SUM(P{row_idx}+R{row_idx}+T{row_idx}+U{row_idx}+V{row_idx}+W{row_idx}+X{row_idx}+Y{row_idx}+Z{row_idx}+AA{row_idx})')

    # ── save output ──────────────────────────────────────────────────────────
    import re as _re
    safe_label = _re.sub(r'[^\w\s\-]', '', ctx.account_label).strip().replace(' ', '_')
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    fname = f'{safe_label} - Account Strategy Analysis - {ts}.xlsm'
    fpath = os.path.join(output_dir, fname)
    wb.save(fpath)
    wb.close()
    return fpath
