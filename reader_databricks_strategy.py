"""
reader_databricks_strategy.py
─────────────────────────────
Reads the Databricks pre-analysis workbook and returns a StrategyContext
object with all signals needed by rules_engine_strategy.py.

Tabs consumed
─────────────
02  Date Range KPIs          → account-level ACoS, TACoS, CPC, spend, sales
08  Campaign Report          → campaign names, subtypes, spend, portfolio assignment
10  Campaigns Grouped by QT  → spend % per campaign subtype (ATM, BAK, BA, SPT, WATM, etc.)
14  Campaign Perf by Child ASIN → slow mover detection (<3 orders), ATM/BA overlap, SPT on Tier 100
15  Campaign Perf by Parent   → parent ASIN count (single-ASIN account detection)
24  ACoS Changes History     → change frequency + direction in last 30 days
25  Portfolio Insights       → portfolio names, IsManaged, has budget cap
33  RBO Configuration        → whether any RBO rules exist
34  Product Level ACoS       → whether product-level ACoS overrides exist
35  Campaign Level ACoS      → whether campaign-level ACoS overrides exist
36  Account Out of Budget    → whether account hit OOB in period
38  Client Success Insights  → ACoS constraint, account details
42  Amazon GGS/Domo          → GGS status, SD impressions, SB impressions
50  Promo Management Trends  → promo activity (discount, coupon, deal)
55  Salesforce Consolidated  → account name, launch date, MRR

All values are stored as plain Python types — no DataFrames escape this module.
Callers receive a single StrategyContext dataclass.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Optional

import pandas as pd


# ── helpers ──────────────────────────────────────────────────────────────────

def _find_header_row(ws, max_scan: int = 10) -> Optional[int]:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        non_empty = [c for c in row if c is not None]
        if len(non_empty) > 3:
            return i
    return None


def _tab_to_records(ws) -> list[dict]:
    hr = _find_header_row(ws)
    if hr is None:
        return []
    headers = None
    records = []
    for row in ws.iter_rows(min_row=hr, values_only=True):
        if headers is None:
            headers = list(row)
            continue
        if not any(row):
            continue
        rec = {headers[j]: row[j] for j in range(len(headers)) if headers[j] is not None}
        records.append(rec)
    return records


def _tab_to_dict(ws) -> dict:
    """Single-data-row tab → dict (first data row only)."""
    records = _tab_to_records(ws)
    return records[0] if records else {}


def _latest_record(records: list[dict], modstamp_key: str = 'SystemModstamp') -> dict:
    if not records:
        return {}
    if len(records) == 1:
        return records[0]
    col = next(
        (k for k in records[0].keys()
         if re.sub(r'[\s_]', '', k).lower() in ('systemmodstamp', 'modstamp')),
        None
    )
    if col:
        try:
            return max(records, key=lambda r: pd.to_datetime(r.get(col), errors='coerce') or pd.NaT)
        except Exception:
            pass
    return records[0]


def _no_data(ws) -> bool:
    """Returns True if tab has 'NO DATA AVAILABLE' message."""
    for row in ws.iter_rows(min_row=5, max_row=10, values_only=True):
        for cell in row:
            if cell and 'NO DATA' in str(cell).upper():
                return True
    return False


def _safe_float(val, default: float = 0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _safe_str(val, default: str = '') -> str:
    return str(val).strip() if val is not None else default


# ── context dataclass ─────────────────────────────────────────────────────────

@dataclass
class StrategyContext:
    # ── identity ──────────────────────────────────────────────────────────────
    account_label: str = ''
    tenant_id: str = ''
    profile_id: str = ''
    member_id: str = ''
    date_range: str = ''
    downloaded: str = ''

    # ── account-level KPIs (tab 02) ───────────────────────────────────────────
    acos_actual: float = 0.0          # e.g. 0.64 = 64%
    tacos_actual: float = 0.0
    cpc_current: float = 0.0
    cpc_last_year: float = 0.0
    total_spend: float = 0.0
    total_sales: float = 0.0
    ad_sales: float = 0.0
    yoy_ad_sales: float = 0.0         # e.g. -0.47 = -47%

    # ── constraint (tab 38) ───────────────────────────────────────────────────
    acos_constraint: float = 0.0      # e.g. 25 = 25%

    # ── ACoS change history (tab 24) ─────────────────────────────────────────
    acos_changes_30d: int = 0         # number of changes in last 30 days
    acos_direction: str = 'stable'    # 'decreasing', 'increasing', 'mixed'
    acos_current_target: float = 0.0  # most recent IACoS target
    acos_gap_to_constraint: float = 0.0  # current_target - constraint (in pp)

    # ── campaign type spend mix (tab 10) ─────────────────────────────────────
    # Each value is % of total spend (0.0–1.0)
    pct_imported: float = 0.0
    pct_non_quartile: float = 0.0
    pct_atm: float = 0.0
    pct_ba: float = 0.0
    pct_bak: float = 0.0
    pct_spt: float = 0.0
    pct_watm: float = 0.0
    pct_sb: float = 0.0
    pct_sbv: float = 0.0
    pct_sd: float = 0.0
    pct_br: float = 0.0
    pct_ow: float = 0.0
    pct_op: float = 0.0

    # absolute spend per type
    spend_imported: float = 0.0
    spend_non_quartile: float = 0.0
    spend_atm: float = 0.0
    spend_ba: float = 0.0
    spend_bak: float = 0.0
    spend_spt: float = 0.0
    spend_watm: float = 0.0
    spend_sb: float = 0.0
    spend_sbv: float = 0.0
    spend_sd: float = 0.0

    # ── campaign names (tab 08) ───────────────────────────────────────────────
    campaign_names: list[str] = field(default_factory=list)
    campaigns_not_in_portfolio: int = 0
    has_cat_sp: bool = False           # CAT_SP_ naming convention
    has_cat_non_standard: bool = False # CAT_ but not CAT_SP_
    has_sbv: bool = False
    has_sd: bool = False
    has_watm: bool = False
    has_catchall: bool = False
    ba_campaign_count: int = 0
    unmanaged_campaign_count: int = 0  # tab 31
    # NEW: additional campaign presence signals
    has_bak: bool = False              # any BAK campaign active
    has_op: bool = False               # any OP (product target) campaign active
    has_sd_prd: bool = False           # any SD_PRD campaign active
    has_vcpm: bool = False             # any VCPM campaign active
    watm_campaign_count: int = 0       # number of distinct WATM campaigns
    sbv_naming_compliant: bool = True  # all SBV campaigns use SBV_ prefix

    # ── portfolio (tab 25) ────────────────────────────────────────────────────
    portfolio_count: int = 0
    managed_portfolio_count: int = 0
    portfolios_with_budget_cap: int = 0
    portfolio_names: list[str] = field(default_factory=list)

    # ── signals from presence/absence tabs ───────────────────────────────────
    has_rbo: bool = False              # tab 33 has data
    has_product_acos_overrides: bool = False   # tab 34
    has_campaign_acos_overrides: bool = False  # tab 35
    has_oob: bool = False              # tab 36

    # ── GGS / display (tab 42) ───────────────────────────────────────────────
    ggs_status: str = 'No'             # 'Yes' or 'No'
    sd_impressions: int = 0
    sb_impressions: int = 0

    # ── promo (tab 50) ────────────────────────────────────────────────────────
    has_active_promo: bool = False     # any PromotionDiscount > 0 in period
    promo_asin_count: int = 0          # number of ASINs with active promo
    promo_cost_rate: float = 0.0       # avg PromoCostRate_pct over last 4 weeks

    # ── Pro Suite audiences (tab 51) ─────────────────────────────────────────
    has_prosuite_audiences: bool = False      # any campaign has HasAudience=True
    prosuite_audience_spend_pct: float = 0.0  # share of total spend with audience
    prosuite_active: bool = False             # tab 51 has real data rows (not NO DATA) = ProSuite enabled

    # ── BA campaign orders (tab 08) ───────────────────────────────────────────
    ba_orders_30d: float = 0.0               # total orders across BA campaigns in period

    # ── SnS and Promo Management (tab 50) ────────────────────────────────────
    has_sns_active: bool = False       # SnS subscriptions active (ActiveSubscriptions > 0)
    has_promo_portfolio: bool = False  # portfolio named SD_QTL_AMZ or Promo_ exists (GGS gate)

    # ── TACoS constraint (tab 38) ─────────────────────────────────────────────
    tacos_constraint: float = 0.0      # 0 = not documented

    # ── ASIN tiers (tab 15) ───────────────────────────────────────────────────
    tier1_asin_count: int = 0         # TIER 10–30
    tier1_with_atm: int = 0           # Tier1 ASINs that have ATM spend

    # ── Tab 14 ASIN-level derived signals ─────────────────────────────────────
    slow_movers_with_ba: int = 0       # ASINs with <3 orders AND BA_Spend > 0
    slow_mover_asins_with_ba: list[str] = field(default_factory=list)   # ASIN IDs matching above
    slow_mover_asins_with_atm: list[str] = field(default_factory=list)  # ASINs with <3 orders AND ATM_Spend > 0
    tier100_with_spt_asins: list[str] = field(default_factory=list)     # Tier 100 ASINs with SPT_Spend > 0
    max_asin_orders_30d: float = 0.0   # highest order count for any single ASIN in period
    atm_ba_overlap_count: int = 0      # ASINs with ATM_Spend > 0 AND BA_Spend > 0 AND Orders > 80
    atm_ba_overlap_asins: list[str] = field(default_factory=list)       # ASIN IDs + orders for S012 what_we_saw
    spt_slow_mover_pct: float = 0.0    # pct of SPT spend on ASINs with <3 orders
    spt_avg_acos: float = 0.0          # spend-weighted avg ACoS across SPT campaigns (tab 08)
    bak_campaigns: list[dict] = field(default_factory=list)             # [{name, spend, pct_of_total, acos}]
    catalog_asin_count: int = 0        # total ASINs in tab 14 (catalog size)
    spending_asin_count: int = 0       # ASINs with AdSpend > 0 in period
    low_order_campaign_count: int = 0  # campaigns with 1-3 orders in 30d

    # ── Tab 08 campaign-type performance ─────────────────────────────────────
    # avg ACoS by campaign naming prefix — for S053-S056, S064, S066
    atm_avg_acos: float = 0.0          # avg ACoS of ATM_ campaigns
    br_avg_acos: float = 0.0           # avg ACoS of BR_ campaigns
    ph_avg_acos: float = 0.0           # avg ACoS of PH_ campaigns
    ow_avg_acos: float = 0.0           # avg ACoS of OW_ campaigns
    br_campaign_count: int = 0         # count of BR_ campaigns
    ow_campaign_count: int = 0         # count of OW_ campaigns
    ph_campaign_count: int = 0         # count of PH_ campaigns
    has_both_watm_and_catchall: bool = False  # both WATM and CatchAll active simultaneously
    bak_name_overlaps_ba: bool = False  # at least one BAK campaign name matches a BA campaign name

    # ── Campaign-type outperforming signals (tab 08) ──────────────────────────
    sd_flex_avg_acos: float = 0.0      # avg ACoS of SD_FLEX_ campaigns
    sd_audi_avg_acos: float = 0.0      # avg ACoS of SD_AUDI_ campaigns
    sd_prd_avg_acos: float = 0.0       # avg ACoS of SD_PRD_ campaigns
    sd_flex_vcpm_pct: float = 0.0      # share of SD_FLEX spend on VCPM campaigns
    sd_audi_vcpm_pct: float = 0.0      # share of SD_AUDI spend on VCPM campaigns
    sd_prd_vcpm_pct: float = 0.0       # share of SD_PRD spend on VCPM campaigns
    sb_avg_acos: float = 0.0           # avg ACoS of SB_ campaigns
    sbv_avg_acos: float = 0.0          # avg ACoS of SBV_ campaigns
    ow_avg_acos: float = 0.0           # avg ACoS of OW_ (exact match) campaigns
    op_avg_acos: float = 0.0           # avg ACoS of OP_ (product target) campaigns
    op_campaign_count: int = 0         # count of OP_ campaigns
    op_campaigns_with_spend: int = 0   # OP campaigns that had spend in the period
    catchall_orders: float = 0.0       # total orders from CatchAll campaigns
    catsp_avg_acos: float = 0.0        # avg ACoS of CAT_SP_ campaigns

    # ── ACoS change cadence (tab 24) ──────────────────────────────────────────
    days_since_last_acos_change: int = 999  # days since most recent portal ACoS change
    # NEW: single-ASIN account flag
    parent_asin_count: int = 0        # unique parent ASINs in tab 15

    # ── CPC trend ─────────────────────────────────────────────────────────────
    # NEW: derived from tab 02 cpc_current vs cpc_last_year
    cpc_yoy_change_pct: float = 0.0   # positive = increased, negative = decreased

    # ── Portfolio coverage (tab 08) ───────────────────────────────────────────
    campaigns_in_portfolio_pct: float = 0.0   # share of campaigns assigned to a portfolio
    total_campaign_count: int = 0              # total campaign count from tab 08

    # ── Search term category mix (tab 12) ─────────────────────────────────────
    branded_spend_pct: float = 0.0      # Branded row Spend_Pct
    branded_acos: float = 0.0           # Branded row acos
    branded_cpc: float = 0.0            # Branded row cpc
    non_branded_spend_pct: float = 0.0  # Non Branded row Spend_Pct
    non_branded_acos: float = 0.0       # Non Branded row acos
    non_branded_cpc: float = 0.0        # Non Branded row cpc
    vcpm_spend_pct: float = 0.0         # VCPM row Spend_Pct (of total search term spend)

    # ── Monthly trend signals (tab 04 — L24M) ────────────────────────────────
    tacos_trend: str = 'stable'         # 'increasing', 'decreasing', 'stable'
    tacos_trend_pp: float = 0.0         # total pp change over last 3 months
    mom_spend_change: float = 0.0       # MoM spend change last month (decimal)
    mom_sales_change: float = 0.0       # MoM total sales change last month (decimal)
    l3m_tacos_avg: float = 0.0          # average TACoS over last 3 full months

    # ── Salesforce account attributes (tab 38) ────────────────────────────────
    primary_objective: str = ''          # e.g. 'Growth' | 'Profit Maximization (Efficiency, Margin, TACOS)' | etc.
    primary_spend_kpi: str = ''          # 'ACOS' | 'ROAS' | 'TACOS'
    repeat_purchase: str = ''            # 'High' | 'Medium' | 'Low'
    commodity_or_brand: str = ''         # 'Commodity' | 'Brand'
    sales_concentration: str = ''        # 'Low Concentration' | 'Medium Concentration' | 'High Concentration'
    tacos_constraint_documented: bool = False  # True only when tab 38 has a real TACoS_Constraint__c value

    # ── Tab 18 — category performance signals ────────────────────────────────
    qualifying_category_count: int = 0   # categories with AsinCount>=30 AND TotalSalesPct>=5% (CAT_SP gate)

    # ── Structural signals for new controls S126–S136 ────────────────────────
    # S126 — branded + NB both significant in same auto layer
    branded_nb_mixed_in_ba: bool = False  # branded_spend_pct > 0.20 AND non_branded_spend_pct > 0.20

    # S127 — auto-to-manual ratio (tab 10 derived)
    auto_spend_pct: float = 0.0           # (ATM + BA + WATM) as pct of total
    manual_exact_pct: float = 0.0         # BAK pct (primary manual exact layer)

    # S128 — BAK harvest stalled (tab 10 derived)
    bak_underfed: bool = False            # pct_bak > 0 but pct_bak < pct_ba * 0.10

    # S129 — own product page undefended (tab 10 + tab 08 derived)
    has_ow: bool = False                  # any OW_ campaign in campaign names
    ow_campaign_count: int = 0            # count of OW_ campaigns

    # S130 — BR discovery layer
    has_br: bool = False                  # any BR_ campaign active with spend

    # S131 — OW own-page coverage (already has_ow above)

    # S133 — BAK branded/NB mixed signal (tab 12 + tab 10 derived)
    bak_branded_nb_mixed: bool = False    # branded_spend_pct > 0.40 AND non_branded_spend_pct > 0.20

    # S134 — TACoS/ACoS divergence (already in trend fields above)

    # S136 — CatchAll graduation overdue (catchall_orders already exists)

    # ── Tab 17 top search terms (S105) ───────────────────────────────────────
    unconverted_top_terms: int = 0        # top-30 terms with orders≥3 AND CVR≥10% (proxy for not-yet-in-BAK)

    # ── Tab 14 inefficient ASIN spend (S107) ─────────────────────────────────
    inefficient_asin_count: int = 0       # ASINs: AdSpend>$100 AND ACoS>1.5x constraint AND orders<5
    paused_sb_count: int = 0             # SB campaigns with state=paused AND spend > 0
    paused_sbv_count: int = 0            # SBV campaigns with state=paused AND spend > 0

    # ── Tab 15 top-seller type coverage (S102) ────────────────────────────────
    top_seller_type_gaps: int = 0         # Tier 10-30 ASINs missing ≥2 of (ATM, BAK, OP)

    # ── Tab 08 BAK inefficiency signals (S053) ────────────────────────────────
    inefficient_bak_count: int = 0        # BAK campaigns: spend>$200 AND ACoS>1.5x constraint AND orders<5

    # ── Tab 08/10 BR inefficiency signals (S057) ─────────────────────────────
    br_inefficiency_flag: bool = False    # pct_br>0.15 AND br_avg_acos>1.5x constraint AND above_acos

    # ── Tab 38 monthly budget (S113) ─────────────────────────────────────────
    monthly_budget: float = 0.0           # Monthly_Budget__c from tab 38 (0 = not documented)

    # ── S039 — category segmentation gate ────────────────────────────────────
    categories_above_10pct: int = 0       # categories contributing >10% of total sales (tab 18)

    # ── S053/S054/S055 — campaign-level ACoS by type (tab 08) ────────────────
    # SP type (Sponsored Products)
    sp_worst_campaign_name: str = ''      # campaign with highest ACoS among SP-type
    sp_worst_campaign_acos: float = 0.0   # its ACoS (decimal)
    sp_campaigns_above_threshold: int = 0 # count of SP campaigns above threshold
    # SB type (Sponsored Brands)
    sb_worst_campaign_name: str = ''
    sb_worst_campaign_acos: float = 0.0
    sb_campaigns_above_threshold: int = 0
    # SD type (Sponsored Display)
    sd_worst_campaign_name: str = ''
    sd_worst_campaign_acos: float = 0.0
    sd_campaigns_above_threshold: int = 0

    # ── S101 — tagging and segmentation gap (tab 14) ─────────────────────────
    tags: list[str] = field(default_factory=list)  # all unique tag values (Tag1–Tag5) from tab 14

    # ── S109 — inefficient ASIN spend (revised logic) ────────────────────────
    # Fires when ASIN has spend AND (zero sales OR ACoS > 2× constraint)
    # inefficient_asin_count already exists above — reusing it with new logic
    inefficient_asin_names: list[str] = field(default_factory=list)  # ASIN IDs for what_we_saw

    # ── S087/S088/S089 — bulk campaign type spend shares ─────────────────────
    pct_cat_sp: float = 0.0               # CAT_SP_ spend as fraction of total

    # ── OPD — defensive product targeting (split from OP by platform update) ──
    spend_opd: float = 0.0             # OPD spend from tab 10
    pct_opd: float = 0.0               # OPD share of total spend (tab 10 Perc_Spend)
    has_opd: bool = False              # any OPD campaigns on tab 10
    opd_campaign_count: int = 0        # count of OPD campaigns (tab 08)
    opd_avg_acos: float = 0.0          # avg ACoS of OPD campaigns (tab 08)

    # ── SP layer mix (tab 10) — Pod Playbook framework ────────────────────────
    # Granular = ATM+BR+OP+OW+PH · Bulk = BA+BAK+CAT_SP · Defensive = WATM+SPT+OPD+SD_SPT
    gran_spend_pct: float = 0.0        # granular layer share of SP layer spend
    bulk_spend_pct: float = 0.0        # bulk layer share of SP layer spend
    def_spend_pct: float = 0.0         # defensive layer share of SP layer spend
    sp_layer_spend: float = 0.0        # total spend across the 12 layer subtypes
    gran_campaign_count: int = 0       # granular campaigns with spend in period (tab 08)
    gran_median_orders: float = 0.0    # median orders per granular campaign in period (tab 08)

    # ── Pod mapping (tab 43) ──────────────────────────────────────────────────
    main_category: str = ''            # Amazon main category from tab 43 'Your Category'

    # ── S035 best-seller spend concentration (tab 14) ─────────────────────────
    tier1_sales_pct: float = 0.0       # Tier 10-30 share of total sales (TotalSales)
    tier1_core_spend_pct: float = 0.0  # Tier 10-30 ATM+BA+BAK spend share of total AdSpend


# ── main reader ───────────────────────────────────────────────────────────────

def read_strategy_context(pre_analysis_path: str) -> StrategyContext:
    import openpyxl
    pa = openpyxl.load_workbook(pre_analysis_path, data_only=True, read_only=True)
    ctx = StrategyContext()

    # ── identity (tab 01) ────────────────────────────────────────────────────
    ws01 = pa['01_Advertiser_Name']
    account_str = date_range = downloaded = ''
    for row in ws01.iter_rows(min_row=1, max_row=4, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                if 'Account:' in cell:
                    account_str = cell
                elif 'Date Range:' in cell:
                    date_range = cell.replace('Date Range: ', '').strip()
                elif 'Downloaded:' in cell:
                    downloaded = cell.replace('Downloaded: ', '').strip()
    m = re.match(
        r'Account:\s*(.+?)\s*\|\s*Tenant ID:\s*(\S+)\s*\|\s*Account ID:\s*(\S+)',
        account_str
    )
    if m:
        ctx.account_label = m.group(1).strip()
        ctx.tenant_id     = m.group(2).strip()
        ctx.profile_id    = m.group(3).strip()
        ctx.member_id     = ctx.account_label.split(' - ')[0].strip()
    ctx.date_range  = date_range
    ctx.downloaded  = downloaded

    # ── tab 02 — account KPIs ────────────────────────────────────────────────
    d02 = _tab_to_dict(pa['02_Date_Range_KPIs__Date_Range_'])
    ctx.acos_actual   = _safe_float(d02.get('ACoS'))
    ctx.tacos_actual  = _safe_float(d02.get('TACoS'))
    ctx.cpc_current   = _safe_float(d02.get('CPC'))
    ctx.cpc_last_year = _safe_float(d02.get('LastYear_CPC'))
    ctx.total_spend   = _safe_float(d02.get('AdSpend'))
    ctx.total_sales   = _safe_float(d02.get('TotalSales'))
    ctx.ad_sales      = _safe_float(d02.get('AdSales'))
    ctx.yoy_ad_sales  = _safe_float(d02.get('YoY_AdSales'))

    # CPC YoY change — only compute when both values are present and non-zero
    if ctx.cpc_last_year > 0 and ctx.cpc_current > 0:
        ctx.cpc_yoy_change_pct = (ctx.cpc_current - ctx.cpc_last_year) / ctx.cpc_last_year

    # ── tab 38 — constraint + Salesforce account attributes ─────────────────
    d38_all = _tab_to_records(pa['38_Client_Success_Insights_Repo'])
    d38 = _latest_record(d38_all)
    ctx.acos_constraint  = _safe_float(d38.get('ACOS_Constraint__c'))
    ctx.tacos_constraint = _safe_float(d38.get('TACoS_Constraint__c'))
    ctx.tacos_constraint_documented = ctx.tacos_constraint > 0
    # Fallback: if TACoS constraint is not documented, use actual TACoS as reference
    if ctx.tacos_constraint == 0.0 and ctx.tacos_actual > 0:
        ctx.tacos_constraint = ctx.tacos_actual * 100
    # Salesforce account attributes
    ctx.primary_objective  = _safe_str(d38.get('Primary_Objective__c'))
    ctx.primary_spend_kpi  = _safe_str(d38.get('Primary_Spend_KPI__c'))
    ctx.repeat_purchase    = _safe_str(d38.get('Repeat_Purchase_Behavior__c'))
    ctx.commodity_or_brand = _safe_str(d38.get('Commodity_Products_or_Branded_Products__c'))
    ctx.sales_concentration = _safe_str(d38.get('Sales_Concentration__c'))
    # Monthly budget — for S113 budget alignment check
    ctx.monthly_budget = _safe_float(d38.get('Monthly_Budget__c'))

    # ── tab 24 — ACoS change history ─────────────────────────────────────────
    changes = _tab_to_records(pa['24_Account_ACoS_Changes_History'])
    cutoff = datetime.now() - timedelta(days=30)
    recent = []
    for r in changes:
        dt = r.get('Change_Date')
        if dt and hasattr(dt, 'date') and dt >= cutoff:
            recent.append(r)
    ctx.acos_changes_30d = len(recent)

    if changes:
        newest = changes[0]  # already sorted newest-first by Databricks
        ctx.acos_current_target = _safe_float(newest.get('IACoS_Percent'))

        # direction: look at last 5 changes
        last5 = changes[:5]
        if len(last5) >= 2:
            deltas = []
            for r in last5:
                old = _safe_float(r.get('Old_IACoS_Target'))
                new = _safe_float(r.get('IACoS_Percent'))
                deltas.append(new - old)
            n_dec = sum(1 for d in deltas if d < 0)
            n_inc = sum(1 for d in deltas if d > 0)
            if n_dec >= 4:
                ctx.acos_direction = 'decreasing'
            elif n_inc >= 4:
                ctx.acos_direction = 'increasing'
            else:
                ctx.acos_direction = 'mixed'

    ctx.acos_gap_to_constraint = (
        ctx.acos_current_target - ctx.acos_constraint
        if ctx.acos_constraint > 0 else 0.0
    )

    # ── tab 10 — campaign type mix ───────────────────────────────────────────
    subtypes = _tab_to_records(pa['10_Campaigns_Grouped_by_QT_Camp'])
    subtype_map: dict[str, dict] = {}
    for r in subtypes:
        st = _safe_str(r.get('CampaignSubType')).upper()
        subtype_map[st] = r

    def _pct(key: str) -> float:
        return _safe_float(subtype_map.get(key.upper(), {}).get('Perc_Spend'))

    def _spend(key: str) -> float:
        return _safe_float(subtype_map.get(key.upper(), {}).get('Spend'))

    ctx.pct_imported      = _pct('Imported')
    ctx.pct_non_quartile  = _pct('Non-Quartile')
    ctx.pct_atm           = _pct('ATM')
    ctx.pct_ba            = _pct('BA')
    ctx.pct_bak           = _pct('BAK')
    ctx.pct_spt           = _pct('SPT')
    ctx.pct_watm          = _pct('WATM')
    ctx.pct_sb            = _pct('SB')
    ctx.pct_sbv           = _pct('SBV')
    # pct_sd / spend_sd: tab 10 breaks SD into subtypes (SD_AUDI, SD_PRD, SD_FLEX, SD_SPT).
    # There is no single 'SD' row — sum all SD-prefixed subtypes explicitly.
    _sd_keys = [k for k in subtype_map if k.startswith('SD')]
    ctx.pct_sd   = sum(_safe_float(subtype_map[k].get('Perc_Spend')) for k in _sd_keys)
    ctx.pct_br            = _pct('BR')
    ctx.pct_ow            = _pct('OW')
    ctx.pct_op            = _pct('OP')

    ctx.spend_imported     = _spend('Imported')
    ctx.spend_non_quartile = _spend('Non-Quartile')
    ctx.spend_atm          = _spend('ATM')
    ctx.spend_ba           = _spend('BA')
    ctx.spend_bak          = _spend('BAK')
    ctx.spend_spt          = _spend('SPT')
    ctx.spend_watm         = _spend('WATM')
    ctx.spend_sb           = _spend('SB')
    ctx.spend_sbv          = _spend('SBV')
    ctx.spend_sd           = sum(_safe_float(subtype_map[k].get('Spend')) for k in _sd_keys)

    # OPD — defensive product targeting, distinct CampaignSubType since platform update
    ctx.pct_opd   = _pct('OPD')
    ctx.spend_opd = _spend('OPD')
    ctx.has_opd   = 'OPD' in subtype_map

    # SP layer mix — Pod Playbook framework
    _gran_keys = ('ATM', 'BR', 'OP', 'OW', 'PH')
    _bulk_keys = ('BA', 'BAK', 'CAT_SP')
    _def_keys  = ('WATM', 'SPT', 'OPD', 'SD_SPT')
    _gran_spend = sum(_spend(k) for k in _gran_keys)
    _bulk_spend = sum(_spend(k) for k in _bulk_keys)
    _def_spend  = sum(_spend(k) for k in _def_keys)
    _layer_total = _gran_spend + _bulk_spend + _def_spend
    ctx.sp_layer_spend = _layer_total
    if _layer_total > 0:
        ctx.gran_spend_pct = _gran_spend / _layer_total
        ctx.bulk_spend_pct = _bulk_spend / _layer_total
        ctx.def_spend_pct  = _def_spend  / _layer_total

    # ── tab 08 — campaign names ───────────────────────────────────────────────
    camp_records = _tab_to_records(pa['08_Campaign_Report'])
    names = [_safe_str(r.get('CampaignName')) for r in camp_records if r.get('CampaignName')]
    ctx.campaign_names = names

    ctx.campaigns_not_in_portfolio = sum(
        1 for r in camp_records
        if _safe_str(r.get('PortfolioName')).startswith('Campaign Not in Portfolio')
    )
    # ── campaign presence signals — use CampaignSubType (column E) as primary source ──
    # SubType values from Databricks: ATM, BA, BAK, BR, CAT_SP, OP, OW, PH, SB, SBV,
    #   SD_AUDI, SD_PRD, SD_FLEX, SD_SPT, SPT, WATM, Non-Quartile, Imported
    # CatchAll is NOT a SubType — it shows as Non-Quartile; detect via name pattern
    #   scoped to Non-Quartile rows only to avoid false matches on Quartile campaigns.

    def _subtype_eq(val: str) -> bool:
        return any(_safe_str(r.get('CampaignSubType')).upper() == val for r in camp_records)

    def _subtype_count(val: str) -> int:
        return sum(1 for r in camp_records if _safe_str(r.get('CampaignSubType')).upper() == val)

    ctx.has_cat_sp           = _subtype_eq('CAT_SP')
    ctx.has_cat_non_standard = (
        any(re.search(r'\bCAT_', n, re.IGNORECASE) for n in names) and not ctx.has_cat_sp
    )
    ctx.has_sbv              = _subtype_eq('SBV')
    ctx.has_sd               = any(
        _safe_str(r.get('CampaignSubType')).upper().startswith('SD')
        for r in camp_records
    )
    ctx.has_watm             = _subtype_eq('WATM')
    ctx.watm_campaign_count  = _subtype_count('WATM')

    # CatchAll: Non-Quartile rows only, then name pattern on CampaignName
    non_qt_names = [
        _safe_str(r.get('CampaignName'))
        for r in camp_records
        if _safe_str(r.get('CampaignSubType')).upper() == 'NON-QUARTILE'
    ]
    ctx.has_catchall = any(re.search(r'catch.?all', n, re.IGNORECASE) for n in non_qt_names)

    ctx.ba_campaign_count = _subtype_count('BA')
    ctx.has_bak           = _subtype_eq('BAK')
    ctx.has_op            = _subtype_eq('OP')
    ctx.has_sd_prd        = _subtype_eq('SD_PRD')
    ctx.has_vcpm          = any(re.search(r'\bVCPM\b', n, re.IGNORECASE) for n in names)

    # BR and OW via SubType
    ctx.has_br            = _subtype_eq('BR')
    ctx.has_ow            = _subtype_eq('OW')
    ctx.ow_campaign_count = _subtype_count('OW')

    # SBV naming compliance: all SBV campaigns should start with SBV_
    sbv_names = [n for n in names if re.search(r'\bSBV\b', n, re.IGNORECASE)]
    ctx.sbv_naming_compliant = all(
        re.match(r'SBV_', n, re.IGNORECASE) for n in sbv_names
    ) if sbv_names else True

    # ── tab 25 — portfolios ───────────────────────────────────────────────────
    port_records = _tab_to_records(pa['25_Portfolio_Insights_and_Confi'])
    ctx.portfolio_count    = len(port_records)

    def _is_true(val) -> bool:
        """Normalise booleans — openpyxl returns True/False, calamine may return 1/0 or strings."""
        if val is True or val == 1:
            return True
        if isinstance(val, str) and val.strip().lower() in ('true', '1', 'yes'):
            return True
        return False

    ctx.managed_portfolio_count = sum(
        1 for r in port_records if _is_true(r.get('IsManaged'))
    )
    ctx.portfolios_with_budget_cap = sum(
        1 for r in port_records if _is_true(r.get('IsBudgetCap'))
    )
    ctx.portfolio_names = [_safe_str(r.get('Portfolio_Name')) for r in port_records]
    # GGS gate: detect SD portfolio commitment.
    # Portfolio names vary: 'SD_QTL_AMZ', 'SD QT AMZ', 'SD QTL', etc.
    ctx.has_promo_portfolio = any(
        ('SD' in str(n).upper() and ('QT' in str(n).upper() or 'AMZ' in str(n).upper()))
        or 'PROMO' in str(n).upper()
        for n in ctx.portfolio_names
    )

    # ── tabs 33/34/35/36 — presence/absence signals ───────────────────────────
    ctx.has_rbo                    = not _no_data(pa['33_RBO_Configuration_Insights'])
    ctx.has_product_acos_overrides = not _no_data(pa['34_Product_Level_ACoS'])
    ctx.has_campaign_acos_overrides = not _no_data(pa['35_Campaign_Level_ACoS'])
    ctx.has_oob                    = not _no_data(pa['36_Account_Out_of_Budget'])

    # ── tab 42 — GGS ─────────────────────────────────────────────────────────
    ggs_records = _tab_to_records(pa['42_Amazon_GGS_Domo'])
    if ggs_records:
        # Column name confirmed: 'Amazon GGS' — add fallbacks for resilience
        _ggs_col = next(
            (k for k in (ggs_records[0] or {}).keys()
             if re.sub(r'[\s_]', '', str(k)).lower() in ('amazongs', 'ggsstatus', 'ggs')),
            'Amazon GGS'
        )
        ggs_vals = [_safe_str(r.get(_ggs_col)) for r in ggs_records]
        ctx.ggs_status = 'Yes' if any(v == 'Yes' for v in ggs_vals) else 'No'
        ctx.sd_impressions = sum(
            int(_safe_float(r.get('Impressions')))
            for r in ggs_records
            if _safe_str(r.get('CampaignType')) == 'Sponsored Display'
        )
        ctx.sb_impressions = sum(
            int(_safe_float(r.get('Impressions')))
            for r in ggs_records
            if _safe_str(r.get('CampaignType')) == 'Sponsored Brands'
        )

    # ── tab 43 — cohort main category (pod mapping source) ────────────────────
    try:
        _t43_key = next((n for n in pa.sheetnames if str(n).startswith('43_')), None)
        if _t43_key:
            _t43_records = _tab_to_records(pa[_t43_key])
            if _t43_records:
                _cat_col = next(
                    (k for k in _t43_records[0].keys()
                     if re.sub(r'[\s_]', '', str(k)).lower() == 'yourcategory'),
                    None)
                if _cat_col:
                    _cat_val = _safe_str(_t43_records[0].get(_cat_col))
                    if _cat_val and _cat_val.upper() not in ('NAN', 'NONE'):
                        ctx.main_category = _cat_val
    except Exception:
        pass

    # ── tab 50 — promo ────────────────────────────────────────────────────────
    promo_records = _tab_to_records(pa['50_Promo_Management___Account_T'])
    ctx.has_active_promo = any(
        _safe_float(r.get('PromotionDiscount')) > 0
        for r in promo_records
    )
    # Count ASINs with active promo and average promo cost rate over last 4 weeks
    if promo_records:
        last4 = promo_records[-4:]
        ctx.promo_asin_count = max(
            int(_safe_float(r.get('ActivePromoASINs'))) for r in last4
        )
        rates = [_safe_float(r.get('PromoCostRate_pct')) for r in last4
                 if _safe_float(r.get('PromoCostRate_pct')) > 0]
        ctx.promo_cost_rate = sum(rates) / len(rates) if rates else 0.0
        # SnS signal: any row with ActiveSubscriptions > 0
        ctx.has_sns_active = any(
            _safe_float(r.get('ActiveSubscriptions')) > 0
            for r in promo_records
        )

    # ── tab 51 — Pro Suite audience performance ───────────────────────────────
    try:
        prosuite_records = _tab_to_records(pa['51_Pro_Suite__Audience_Performa'])
        if prosuite_records and not any(
            'NO DATA' in str(list(r.values())).upper() for r in prosuite_records[:1]
        ):
            ctx.prosuite_active = True
            total_ps_spend = sum(_safe_float(r.get('TotalSpend')) for r in prosuite_records)
            audience_spend = sum(
                _safe_float(r.get('TotalSpend'))
                for r in prosuite_records
                if r.get('HasAudience') in (True, 'True', 'true', 1, '1')
            )
            ctx.has_prosuite_audiences = audience_spend > 0
            ctx.prosuite_audience_spend_pct = (
                audience_spend / total_ps_spend if total_ps_spend > 0 else 0.0
            )
    except Exception:
        pass  # tab may not exist in older exports — safe default is False / 0.0

    # ── tab 15 — ASIN tiers ───────────────────────────────────────────────────
    asin_records = _tab_to_records(pa['15_Campaign_Performance_by_PARE'])

    # ATM-qualifying threshold: ≥1.5 orders/day × 30 days = 45 orders
    # Tier labels (TIER 10/20/30) are unreliable — accounts can have TIER 50/80
    # ASINs with 20+ orders/day that clearly qualify for ATM.
    # Use Orders column directly as the source of truth.
    _ATM_ORDERS_THRESHOLD = 45
    tier1 = [
        r for r in asin_records
        if _safe_float(r.get('Orders')) >= _ATM_ORDERS_THRESHOLD
    ]
    ctx.tier1_asin_count = len(tier1)
    ctx.tier1_with_atm   = sum(
        1 for r in tier1 if _safe_float(r.get('ATM_Spend')) > 0
    )

    # S035 — best-seller spend concentration.
    # Tier 10-30 share of total sales vs their combined ATM+BA+BAK spend share.
    _total_sales_14 = sum(_safe_float(r.get('TotalSales')) for r in asin_records)
    _total_spend_14 = sum(_safe_float(r.get('AdSpend')) for r in asin_records)
    if _total_sales_14 > 0:
        _t1_sales = sum(_safe_float(r.get('TotalSales')) for r in tier1)
        ctx.tier1_sales_pct = _t1_sales / _total_sales_14
    if _total_spend_14 > 0:
        _t1_core_spend = sum(
            _safe_float(r.get('ATM_Spend'))
            + _safe_float(r.get('BA_Spend'))
            + _safe_float(r.get('BAK_Spend'))
            for r in tier1
        )
        ctx.tier1_core_spend_pct = _t1_core_spend / _total_spend_14
    # NEW: count unique parent ASINs across all tiers to detect single-ASIN accounts
    parent_asins = set(
        _safe_str(r.get('ParentASIN'))
        for r in asin_records
        if not pd.isna(r.get('ParentASIN')) and _safe_str(r.get('ParentASIN'))
    )
    ctx.parent_asin_count = len(parent_asins)

    # Top-seller type coverage gaps — for S102
    # Check each Tier 10-30 ASIN: missing ≥2 of (ATM_Spend, BAK_Spend, OP_Spend) → counts as gap
    gap_count = 0
    for r in tier1:
        missing = sum([
            _safe_float(r.get('ATM_Spend')) == 0,
            _safe_float(r.get('BAK_Spend')) == 0,
            _safe_float(r.get('OP_Spend'))  == 0,
        ])
        if missing >= 2:
            gap_count += 1
    ctx.top_seller_type_gaps = gap_count

    # ── tab 14 — ASIN-level derived signals ─────────────────────────────────────
    try:
        import pandas as _pd14
        xl14 = _pd14.ExcelFile(pre_analysis_path, engine='calamine')
        # Detect header row dynamically — same 4-row metadata pattern as other tabs
        df14_raw = xl14.parse('14_Campaign_Performance_by_Adve', header=None)
        hdr_row14 = None
        for idx in range(min(10, len(df14_raw))):
            if df14_raw.iloc[idx].notna().sum() > 3:
                hdr_row14 = idx
                break
        if hdr_row14 is None:
            raise ValueError("Tab 14 header not found")
        df14 = xl14.parse('14_Campaign_Performance_by_Adve', header=hdr_row14)
        df14 = df14.loc[:, ~df14.columns.astype(str).str.match(r'^Unnamed')]
        df14 = df14.dropna(subset=['asin']) if 'asin' in df14.columns else df14

        if not df14.empty and 'asin' in df14.columns:
            tier_col     = 'Tier' if 'Tier' in df14.columns else None
            ba_col       = 'BA_Spend' if 'BA_Spend' in df14.columns else None
            atm_col      = 'ATM_Spend' if 'ATM_Spend' in df14.columns else None
            spt_col      = 'SPT_Spend' if 'SPT_Spend' in df14.columns else None
            orders_col   = 'Orders' if 'Orders' in df14.columns else None
            spend_col    = 'AdSpend' if 'AdSpend' in df14.columns else None

            has_total_sales = 'TotalSales' in df14.columns
            has_aov = 'AOV' in df14.columns

            # max orders per ASIN — for S014 (no top seller check)
            # Uses total orders proxy (TotalSales/AOV) when available; fallback to ad orders.
            if orders_col:
                if has_total_sales and has_aov:
                    aov_vals_max = df14['AOV'].fillna(0).replace(0, float('nan'))
                    total_orders_proxy = (df14['TotalSales'].fillna(0) / aov_vals_max).fillna(df14[orders_col].fillna(0))
                    ctx.max_asin_orders_30d = float(total_orders_proxy.max())
                else:
                    ctx.max_asin_orders_30d = float(df14[orders_col].fillna(0).max())

            # Global slow mover definition:
            # Use total orders proxy = TotalSales / AOV when both columns exist.
            # This prevents ad-only slow movers from being flagged when the ASIN
            # sells well organically (e.g. 1 ad order but 15 total orders).
            # Fallback: use ad Orders directly when TotalSales or AOV is missing.
            if orders_col and has_total_sales and has_aov:
                aov_vals = df14['AOV'].fillna(0).replace(0, float('nan'))
                est_total_orders = (df14['TotalSales'].fillna(0) / aov_vals).fillna(0)
                is_slow_mover = est_total_orders < 3
            elif orders_col:
                is_slow_mover = df14[orders_col].fillna(0) < 3
            else:
                is_slow_mover = pd.Series([False] * len(df14), index=df14.index)

            # Slow movers with BA spend — for S010/S011/S037
            if ba_col is not None:
                has_ba = df14[ba_col].fillna(0) > 0
                slow_ba_mask = is_slow_mover & has_ba
                ctx.slow_movers_with_ba = int(slow_ba_mask.sum())
                ctx.slow_mover_asins_with_ba = list(df14.loc[slow_ba_mask, 'asin'].astype(str))

            # Slow movers with ATM spend — informational, used in what_we_saw
            if atm_col is not None:
                has_atm = df14[atm_col].fillna(0) > 0
                ctx.slow_mover_asins_with_atm = list(
                    df14.loc[is_slow_mover & has_atm, 'asin'].astype(str)
                )

            # Slow movers with SPT spend — for S031
            if spt_col is not None:
                has_spt = df14[spt_col].fillna(0) > 0
                spt_total  = df14[spt_col].fillna(0).sum()
                spt_slow   = df14.loc[is_slow_mover, spt_col].fillna(0).sum()
                ctx.spt_slow_mover_pct = spt_slow / spt_total if spt_total > 0 else 0.0

            # Low-velocity ASINs with SPT spend — for S032
            # Uses the same is_slow_mover definition (total orders proxy when available)
            if spt_col is not None:
                has_spt_spend   = df14[spt_col].fillna(0) > 0
                ctx.tier100_with_spt_asins = list(
                    df14.loc[is_slow_mover & has_spt_spend, 'asin'].astype(str)
                )

            # ATM + BA overlap on high-velocity ASINs (>80 orders) — for S012/S013
            # Uses total orders proxy (TotalSales/AOV) so organically-selling ASINs
            # are correctly classified as high-velocity even with low ad order counts.
            if atm_col is not None and ba_col is not None and orders_col is not None:
                has_atm_spend  = df14[atm_col].fillna(0) > 0
                has_ba_spend   = df14[ba_col].fillna(0) > 0
                if has_total_sales and has_aov:
                    aov_vals_ov = df14['AOV'].fillna(0).replace(0, float('nan'))
                    overlap_orders = (df14['TotalSales'].fillna(0) / aov_vals_ov).fillna(df14[orders_col].fillna(0))
                else:
                    overlap_orders = df14[orders_col].fillna(0)
                high_velocity  = overlap_orders > 80
                overlap_mask   = has_atm_spend & has_ba_spend & high_velocity
                ctx.atm_ba_overlap_count = int(overlap_mask.sum())
                if ctx.atm_ba_overlap_count > 0:
                    ctx.atm_ba_overlap_asins = [
                        f"{row['asin']} ({int(overlap_orders.loc[idx])} orders)"
                        for idx, row in df14.loc[overlap_mask].iterrows()
                    ]

            # Catalog vs spending ASIN counts — for S022
            ctx.catalog_asin_count  = int(df14['asin'].dropna().nunique())
            if spend_col is not None:
                ctx.spending_asin_count = int(
                    df14[df14[spend_col].fillna(0) > 0]['asin'].dropna().nunique()
                )

            # Inefficient ASIN detection — S109 (revised logic)
            # Fires when: AdSpend > 0 AND (no sales = ACoS is 0/null OR ACoS > 2× constraint)
            acos_col14 = 'ACoS' if 'ACoS' in df14.columns else None
            sales_col14 = 'AdSales' if 'AdSales' in df14.columns else ('TotalSales' if 'TotalSales' in df14.columns else None)
            if spend_col is not None and ctx.acos_constraint > 0:
                has_spend14 = df14[spend_col].fillna(0) > 0
                if acos_col14 is not None:
                    threshold14 = (ctx.acos_constraint / 100) * 2.0
                    acos_num14 = _pd14.to_numeric(df14[acos_col14], errors='coerce')
                    # No sales = acos is NaN or 0 when spend > 0
                    no_sales = has_spend14 & (acos_num14.isna() | (acos_num14 == 0))
                    above_2x = has_spend14 & (acos_num14.fillna(0) > threshold14)
                    ineff14 = no_sales | above_2x
                    ctx.inefficient_asin_count = int(ineff14.sum())
                    if ctx.inefficient_asin_count > 0 and 'asin' in df14.columns:
                        ctx.inefficient_asin_names = list(
                            df14.loc[ineff14, 'asin'].astype(str).dropna().unique()[:10]
                        )
                else:
                    # Fallback: any ASIN with spend but no orders
                    if orders_col is not None:
                        ineff14 = has_spend14 & (df14[orders_col].fillna(0) == 0)
                        ctx.inefficient_asin_count = int(ineff14.sum())

            # Tags extraction — for S101 (Tagging and Segmentation Gap)
            tag_cols = [c for c in df14.columns if re.match(r'^Tag\d+$', str(c), re.IGNORECASE)]
            if tag_cols:
                tag_values = set()
                for tc in tag_cols:
                    tag_values.update(
                        df14[tc].dropna().astype(str).str.strip()
                        .loc[lambda s: s != ''].unique()
                    )
                ctx.tags = [t for t in tag_values if t and t.lower() not in {'nan', 'none', ''}]
    except Exception:
        pass

    # ── tab 24 — days since last ACoS change ─────────────────────────────────────
    try:
        import pandas as _pd24
        from datetime import datetime as _dt
        xl24 = _pd24.ExcelFile(pre_analysis_path, engine='calamine')
        df24_raw = xl24.parse('24_Account_ACoS_Changes_History', header=None)
        hdr_row24 = None
        for idx in range(min(10, len(df24_raw))):
            if df24_raw.iloc[idx].notna().sum() > 3:
                hdr_row24 = idx
                break
        if hdr_row24 is None:
            raise ValueError("Tab 24 header not found")
        df24 = xl24.parse('24_Account_ACoS_Changes_History', header=hdr_row24)
        df24 = df24.loc[:, ~df24.columns.astype(str).str.match(r'^Unnamed')]
        if not df24.empty and 'Change_Date' in df24.columns:
            dates = _pd24.to_datetime(df24['Change_Date'], errors='coerce').dropna()
            if not dates.empty:
                last_change = dates.max()
                if ctx.downloaded is not None:
                    ref = _pd24.Timestamp(ctx.downloaded)
                elif ctx.ref_date is not None:
                    ref = _pd24.Timestamp(ctx.ref_date)
                else:
                    ref = _pd24.Timestamp(_dt.now())
                ctx.days_since_last_acos_change = max(0, int((ref - last_change).days))
    except Exception:
        pass

    # ── tab 08 — campaign-type ACoS and structural signals ──────────────────────
    try:
        import pandas as _pd08
        xl08 = _pd08.ExcelFile(pre_analysis_path, engine='calamine')
        df08 = xl08.parse('08_Campaign_Report', header=5)
        df08 = df08.loc[:, ~df08.columns.astype(str).str.match(r'^Unnamed')]

        if not df08.empty and 'CampaignName' in df08.columns:
            df08['_name'] = df08['CampaignName'].astype(str).str.strip()
            df08['_name_up'] = df08['_name'].str.upper()
            acos_col08 = '_ACOS' if '_ACOS' in df08.columns else ('ACoS' if 'ACoS' in df08.columns else None)
            sub_col08  = 'CampaignSubType' if 'CampaignSubType' in df08.columns else None

            # Helper: avg ACoS by CampaignSubType value (primary) or name prefix (fallback)
            def _subtype_acos(subtype_val: str, prefix_fallback: str) -> float:
                if sub_col08:
                    mask = df08[sub_col08].astype(str).str.upper() == subtype_val.upper()
                else:
                    mask = df08['_name_up'].str.startswith(prefix_fallback)
                if acos_col08 and mask.any():
                    vals = _pd08.to_numeric(df08.loc[mask, acos_col08], errors='coerce').dropna()
                    return float(vals.mean()) if not vals.empty else 0.0
                return 0.0

            def _subtype_count08(subtype_val: str, prefix_fallback: str) -> int:
                if sub_col08:
                    return int((df08[sub_col08].astype(str).str.upper() == subtype_val.upper()).sum())
                return int(df08['_name_up'].str.startswith(prefix_fallback).sum())

            ctx.atm_avg_acos      = _subtype_acos('ATM',    'ATM_')
            ctx.br_avg_acos       = _subtype_acos('BR',     'BR_')
            ctx.ph_avg_acos       = _subtype_acos('PH',     'PH_')
            ctx.ow_avg_acos       = _subtype_acos('OW',     'OW_')
            ctx.br_campaign_count = _subtype_count08('BR',  'BR_')
            ctx.ow_campaign_count = _subtype_count08('OW',  'OW_')
            ctx.ph_campaign_count = _subtype_count08('PH',  'PH_')
            ctx.op_campaign_count = _subtype_count08('OP',  'OP_')
            ctx.opd_campaign_count = _subtype_count08('OPD', 'OPD_')

            # Resolve spend and orders columns — used by multiple blocks below
            orders_col08 = 'Orders' if 'Orders' in df08.columns else None
            spend_col08  = 'Spend'  if 'Spend'  in df08.columns else None

            # OP campaigns with actual spend — for S075
            # S075 should only fire when OP campaigns with spend exist but are underdeveloped,
            # not when no OP campaigns exist at all (that's a framework gap, not a strategy one).
            if sub_col08 and spend_col08:
                _op_mask = df08[sub_col08].astype(str).str.upper() == 'OP'
                _op_spend_num = _pd08.to_numeric(df08.get(spend_col08, 0), errors='coerce').fillna(0)
                ctx.op_campaigns_with_spend = int((_op_mask & (_op_spend_num > 0)).sum())
            if orders_col08:
                lo = df08[orders_col08].fillna(0)
                if sub_col08:
                    auto_mask = df08[sub_col08].astype(str).str.upper().isin(['ATM', 'WATM'])
                else:
                    auto_mask = (
                        df08['_name_up'].str.startswith('ATM_') |
                        df08['_name_up'].str.startswith('WATM_')
                    )
                ctx.low_order_campaign_count = int(((lo >= 1) & (lo <= 3) & ~auto_mask).sum())

            # SPT spend-weighted avg ACoS — for S031
            if sub_col08 and acos_col08 and spend_col08:
                spt_mask = df08[sub_col08].astype(str).str.upper() == 'SPT'
                if spt_mask.any():
                    spt_df   = df08.loc[spt_mask].copy()
                    spt_acos = _pd08.to_numeric(spt_df[acos_col08], errors='coerce').fillna(0)
                    spt_spnd = _pd08.to_numeric(spt_df[spend_col08], errors='coerce').fillna(0)
                    total_spt_spend = spt_spnd.sum()
                    if total_spt_spend > 0:
                        ctx.spt_avg_acos = float((spt_acos * spt_spnd).sum() / total_spt_spend)

            # BAK campaigns list — for what_we_saw
            if sub_col08 and acos_col08 and spend_col08:
                bak_mask = df08[sub_col08].astype(str).str.upper() == 'BAK'
                if bak_mask.any():
                    bak_df = df08.loc[bak_mask].copy()
                    total_spend_acct = float(
                        _pd08.to_numeric(df08[spend_col08], errors='coerce').fillna(0).sum()
                    )
                    for _, row in bak_df.iterrows():
                        sp = float(_pd08.to_numeric(row.get(spend_col08), errors='coerce') or 0)
                        ac = float(_pd08.to_numeric(row.get(acos_col08), errors='coerce') or 0)
                        ctx.bak_campaigns.append({
                            'name':         str(row.get('CampaignName', '')),
                            'spend':        sp,
                            'pct_of_total': sp / total_spend_acct if total_spend_acct > 0 else 0.0,
                            'acos':         ac,
                        })

            # BA campaign orders — for S045 gate (≥80 orders required)
            orders_col08 = next(
                (c for c in df08.columns if str(c).strip().lower() == 'orders'),
                None
            )
            if sub_col08 and orders_col08:
                ba_mask = df08[sub_col08].astype(str).str.upper() == 'BA'
                if ba_mask.any():
                    ctx.ba_orders_30d = float(
                        _pd08.to_numeric(df08.loc[ba_mask, orders_col08], errors='coerce').fillna(0).sum()
                    )

            # Both WATM and CatchAll active simultaneously — for S084
            # WATM: via SubType; CatchAll: Non-Quartile SubType + name pattern
            if sub_col08:
                has_watm08     = (df08[sub_col08].astype(str).str.upper() == 'WATM').any()
                nq_mask08      = df08[sub_col08].astype(str).str.upper() == 'NON-QUARTILE'
                has_catchall08 = bool(
                    df08.loc[nq_mask08, '_name'].str.lower()
                    .str.contains(r'catch.?all', regex=True, na=False).any()
                )
            else:
                has_watm08     = df08['_name_up'].str.startswith('WATM_').any()
                has_catchall08 = df08['_name'].str.lower().str.contains(r'catch.?all', regex=True, na=False).any()
            ctx.has_both_watm_and_catchall = bool(has_watm08 and has_catchall08)

            # BAK name overlaps BA name — for S038
            if sub_col08:
                ba_names  = set(df08.loc[df08[sub_col08].astype(str).str.upper() == 'BA',  '_name'].tolist())
                bak_names = set(df08.loc[df08[sub_col08].astype(str).str.upper() == 'BAK', '_name'].tolist())
                def _extract_token(n):
                    parts = str(n).split('_')
                    return parts[2] if len(parts) > 2 else n
                ba_tokens  = {_extract_token(n) for n in ba_names}
                bak_tokens = {_extract_token(n) for n in bak_names}
                ctx.bak_name_overlaps_ba = bool(ba_tokens & bak_tokens)

            # Campaign-type avg ACoS for outperforming signals — SubType primary, prefix fallback
            ctx.sd_flex_avg_acos = _subtype_acos('SD_FLEX', 'SD_FLEX_')
            ctx.sd_audi_avg_acos = _subtype_acos('SD_AUDI', 'SD_AUDI_')
            ctx.sd_prd_avg_acos  = _subtype_acos('SD_PRD',  'SD_PRD_')
            ctx.sb_avg_acos      = _subtype_acos('SB',      'SB_')
            ctx.sbv_avg_acos     = _subtype_acos('SBV',     'SBV_')
            ctx.op_avg_acos      = _subtype_acos('OP',      'OP_')
            ctx.catsp_avg_acos   = _subtype_acos('CAT_SP',  'CAT_SP_')
            ctx.opd_avg_acos     = _subtype_acos('OPD',     'OPD_')

            # Granular conversion density — isolation gate signal (Pod Playbook)
            try:
                if sub_col08:
                    _gran_mask08 = df08[sub_col08].astype(str).str.upper().isin(
                        ('ATM', 'BR', 'OP', 'OW', 'PH'))
                    _ord_col08 = next(
                        (c for c in df08.columns if str(c).strip().lower() == 'orders'), None)
                    if _ord_col08 is not None and spend_col08:
                        _g_spend = _pd08.to_numeric(
                            df08.loc[_gran_mask08, spend_col08], errors='coerce').fillna(0)
                        _g_orders = _pd08.to_numeric(
                            df08.loc[_gran_mask08, _ord_col08], errors='coerce').fillna(0)
                        _active = _g_spend > 0
                        ctx.gran_campaign_count = int(_active.sum())
                        if _active.any():
                            ctx.gran_median_orders = float(_g_orders[_active].median())
            except Exception:
                pass

            # VCPM spend share per SD subtype — used by S063/064/065 to suppress
            # outperforming signals when VCPM campaigns dominate the subtype.
            # VCPM campaigns use impression-based billing; their ACoS is not comparable.
            if sub_col08 and spend_col08:
                spend_num08 = _pd08.to_numeric(df08[spend_col08], errors='coerce').fillna(0)
                for _sd_sub, _attr in [('SD_FLEX', 'sd_flex_vcpm_pct'), ('SD_AUDI', 'sd_audi_vcpm_pct'), ('SD_PRD', 'sd_prd_vcpm_pct')]:
                    _sub_mask = df08[sub_col08].astype(str).str.upper() == _sd_sub
                    _sub_total = spend_num08[_sub_mask].sum()
                    _vcpm_mask = _sub_mask & df08['_name_up'].str.contains('VCPM', na=False)
                    _vcpm_spend = spend_num08[_vcpm_mask].sum()
                    setattr(ctx, _attr, _vcpm_spend / _sub_total if _sub_total > 0 else 0.0)

            # CatchAll orders — Non-Quartile rows only, then name pattern
            if orders_col08:
                if sub_col08:
                    nq_rows = df08[sub_col08].astype(str).str.upper() == 'NON-QUARTILE'
                    ca_mask = nq_rows & df08['_name'].str.lower().str.contains(r'catch.?all', regex=True, na=False)
                else:
                    ca_mask = df08['_name'].str.lower().str.contains(r'catch.?all', regex=True, na=False)
                ctx.catchall_orders = float(df08.loc[ca_mask, orders_col08].fillna(0).sum())

            # Paused SB / SBV campaigns — SubType primary, prefix fallback
            if 'State' in df08.columns and spend_col08:
                paused_mask = df08['State'].astype(str).str.lower() == 'paused'
                spend_num   = _pd08.to_numeric(df08[spend_col08], errors='coerce').fillna(0)
                had_spend   = spend_num > 0
                if sub_col08:
                    sb_mask  = df08[sub_col08].astype(str).str.upper() == 'SB'
                    sbv_mask = df08[sub_col08].astype(str).str.upper() == 'SBV'
                else:
                    sb_mask  = df08['_name_up'].str.startswith('SB_') & ~df08['_name_up'].str.startswith('SBV_')
                    sbv_mask = df08['_name_up'].str.startswith('SBV_')
                ctx.paused_sb_count  = int((paused_mask & had_spend & sb_mask).sum())
                ctx.paused_sbv_count = int((paused_mask & had_spend & sbv_mask).sum())

            # BAK inefficiency — spend>$200, ACoS>1.5x constraint, orders<5
            if sub_col08 and acos_col08 and spend_col08 and orders_col08:
                bak_mask2 = df08[sub_col08].astype(str).str.upper() == 'BAK'
                if bak_mask2.any() and ctx.acos_constraint > 0:
                    bak_sp   = _pd08.to_numeric(df08[spend_col08],  errors='coerce').fillna(0)
                    bak_ac   = _pd08.to_numeric(df08[acos_col08],   errors='coerce').fillna(0)
                    bak_ord  = _pd08.to_numeric(df08[orders_col08], errors='coerce').fillna(0)
                    threshold = (ctx.acos_constraint / 100) * 1.5
                    ineff_mask = bak_mask2 & (bak_sp > 200) & (bak_ac > threshold) & (bak_ord < 5)
                    ctx.inefficient_bak_count = int(ineff_mask.sum())

            # Campaign-level ACoS by type — for S053 (SP), S054 (SB), S055 (SD)
            # "worst offending" = highest ACoS campaign above the constraint threshold
            if acos_col08 and spend_col08 and ctx.acos_constraint > 0:
                acos_num = _pd08.to_numeric(df08[acos_col08], errors='coerce').fillna(0)
                spend_num08 = _pd08.to_numeric(df08[spend_col08], errors='coerce').fillna(0)
                has_spend_mask = spend_num08 > 0

                def _worst_campaign(type_mask):
                    m = type_mask & has_spend_mask
                    if not m.any():
                        return '', 0.0, 0
                    subset = df08.loc[m].copy()
                    subset['_acos_n'] = _pd08.to_numeric(subset[acos_col08], errors='coerce').fillna(0)
                    threshold_35 = (ctx.acos_constraint / 100) * 1.35
                    threshold_20 = (ctx.acos_constraint / 100) * 1.20
                    above = subset[subset['_acos_n'] > threshold_20]
                    if above.empty:
                        return '', 0.0, 0
                    worst = above.loc[above['_acos_n'].idxmax()]
                    count_above = int((above['_acos_n'] > threshold_20).sum())
                    return (
                        str(worst.get('CampaignName', '')),
                        float(worst['_acos_n']),
                        count_above,
                    )

                # SP / SB / SD type masks — SubType primary, prefix fallback
                if sub_col08:
                    sp_mask   = ~df08[sub_col08].astype(str).str.upper().isin(['SB', 'SBV', 'SD_FLEX', 'SD_AUDI', 'SD_PRD', 'SD_SPT'])
                    sb_mask_t = df08[sub_col08].astype(str).str.upper() == 'SB'
                    sd_mask_t = df08[sub_col08].astype(str).str.upper().str.startswith('SD')
                else:
                    sp_mask   = ~(
                        df08['_name_up'].str.startswith('SB_') |
                        df08['_name_up'].str.startswith('SBV_') |
                        df08['_name_up'].str.startswith('SD_')
                    )
                    sb_mask_t = df08['_name_up'].str.startswith('SB_') & ~df08['_name_up'].str.startswith('SBV_')
                    sd_mask_t = df08['_name_up'].str.startswith('SD_')

                ctx.sp_worst_campaign_name, ctx.sp_worst_campaign_acos, ctx.sp_campaigns_above_threshold = _worst_campaign(sp_mask)
                ctx.sb_worst_campaign_name, ctx.sb_worst_campaign_acos, ctx.sb_campaigns_above_threshold = _worst_campaign(sb_mask_t)
                ctx.sd_worst_campaign_name, ctx.sd_worst_campaign_acos, ctx.sd_campaigns_above_threshold = _worst_campaign(sd_mask_t)

            # CAT_SP spend share — for S087
            if spend_col08:
                total_sp_08 = float(_pd08.to_numeric(df08[spend_col08], errors='coerce').fillna(0).sum())
                if sub_col08:
                    catsp_rows = df08[sub_col08].astype(str).str.upper() == 'CAT_SP'
                else:
                    catsp_rows = df08['_name_up'].str.startswith('CAT_SP_')
                cat_sp_spend = float(_pd08.to_numeric(
                    df08.loc[catsp_rows, spend_col08], errors='coerce'
                ).fillna(0).sum())
                ctx.pct_cat_sp = cat_sp_spend / total_sp_08 if total_sp_08 > 0 else 0.0
    except Exception:
        pass

    # ── tab 08 — portfolio coverage ratio ─────────────────────────────────────
    # Reuse camp_records already parsed above — avoids double-reading the tab
    ctx.total_campaign_count = len(camp_records)
    if ctx.total_campaign_count > 0:
        in_port = sum(
            1 for r in camp_records
            if not _safe_str(r.get('PortfolioName')).startswith('Campaign Not in Portfolio')
            and _safe_str(r.get('PortfolioName'))
        )
        ctx.campaigns_in_portfolio_pct = in_port / ctx.total_campaign_count

    # ── tab 12 — branded vs non-branded keyword mix ────────────────────────────
    try:
        search_cat_records = _tab_to_records(pa['12_Search_Terms_by_Category'])

        def _get_ci(rec: dict, key: str):
            """Case-insensitive dict lookup."""
            key_l = key.lower()
            for k, v in rec.items():
                if str(k).lower() == key_l:
                    return v
            return None

        for r in search_cat_records:
            cat = _safe_str(r.get('KeywordCategory')).lower().strip()
            if cat == 'branded':
                ctx.branded_spend_pct  = _safe_float(_get_ci(r, 'Spend_Pct'))
                ctx.branded_acos       = _safe_float(_get_ci(r, 'acos'))
                ctx.branded_cpc        = _safe_float(_get_ci(r, 'cpc'))
            elif cat in ('non branded', 'non-branded', 'nonbranded'):
                ctx.non_branded_spend_pct = _safe_float(_get_ci(r, 'Spend_Pct'))
                ctx.non_branded_acos      = _safe_float(_get_ci(r, 'acos'))
                ctx.non_branded_cpc       = _safe_float(_get_ci(r, 'cpc'))
            elif cat == 'vcpm':
                ctx.vcpm_spend_pct = _safe_float(_get_ci(r, 'Spend_Pct'))
    except Exception:
        pass

    # ── tab 18 — category performance (CAT_SP qualification gate + S039) ───────
    try:
        cat18_records = _tab_to_records(pa['18_Performance_by_Category'])
        ctx.qualifying_category_count = sum(
            1 for r in cat18_records
            if _safe_float(r.get('AsinCount')) >= 30
            and _safe_float(r.get('TotalSalesPct')) >= 0.05
        )
        # S039: count categories each contributing >10% of total sales
        ctx.categories_above_10pct = sum(
            1 for r in cat18_records
            if _safe_float(r.get('TotalSalesPct')) >= 0.10
        )
    except Exception:
        pass

    # ── tab 17 — top 30 search terms (S105 proxy) ────────────────────────────
    try:
        top17_records = _tab_to_records(pa['17_Top_30_Search_Terms'])
        ctx.unconverted_top_terms = sum(
            1 for r in top17_records
            if _safe_float(r.get('orders')) >= 3
            and _safe_float(r.get('cr')) >= 0.10
        )
    except Exception:
        pass

    # ── derived structural signals for new controls ───────────────────────────
    # S126 — branded + NB both > 20% of search term spend in the same auto bucket
    ctx.branded_nb_mixed_in_ba = (
        ctx.branded_spend_pct > 0.20 and ctx.non_branded_spend_pct > 0.20
    )
    # S127 — auto spend dominance vs manual exact
    ctx.auto_spend_pct   = ctx.pct_atm + ctx.pct_ba + ctx.pct_watm
    ctx.manual_exact_pct = ctx.pct_bak
    # S128 — BAK harvest stalled: BAK exists but is <10% of BA spend
    ctx.bak_underfed = (
        ctx.pct_bak > 0 and ctx.pct_ba > 0
        and ctx.pct_bak < ctx.pct_ba * 0.10
    )
    # S133 — BAK bucket carries both heavy branded and NB (not split by type)
    ctx.bak_branded_nb_mixed = (
        ctx.branded_spend_pct > 0.40 and ctx.non_branded_spend_pct > 0.20
        and ctx.pct_bak > 0
    )
    # S057 — BR strategy too broad: BR consuming >15% of spend at poor efficiency
    ctx.br_inefficiency_flag = (
        ctx.pct_br > 0.15
        and ctx.acos_constraint > 0
        and ctx.br_avg_acos > (ctx.acos_constraint / 100) * 1.5
        and ctx.acos_actual * 100 > ctx.acos_constraint
    )

    # ── tab 04 — L24M monthly trend signals ──────────────────────────────────
    try:
        import pandas as _pd04
        xl04 = _pd04.ExcelFile(pre_analysis_path, engine='calamine')
        # Detect header row dynamically — avoid hardcoded row offsets
        df04_raw = xl04.parse('04_L24M_Monthly_Performance_Sum', header=None)
        hdr_row04 = None
        for idx in range(min(10, len(df04_raw))):
            if df04_raw.iloc[idx].notna().sum() > 3:
                hdr_row04 = idx
                break
        if hdr_row04 is None:
            raise ValueError("Tab 04 header not found")
        df04 = xl04.parse('04_L24M_Monthly_Performance_Sum', header=hdr_row04)
        df04 = df04.loc[:, ~df04.columns.astype(str).str.match(r'^Unnamed')]

        # Confirmed columns: Month, TotalSales, AdSpend, TACoS, AdSales, OrganicSales, ACoS, CPC, CR
        month_col04 = 'Month'      if 'Month'      in df04.columns else None
        tacos_col04 = 'TACoS'      if 'TACoS'      in df04.columns else None
        spend_col04 = 'AdSpend'    if 'AdSpend'    in df04.columns else None
        sales_col04 = 'TotalSales' if 'TotalSales' in df04.columns else None

        if month_col04 and tacos_col04:
            df04 = df04.dropna(subset=[month_col04])
            df04['_month_ts'] = _pd04.to_datetime(df04[month_col04], errors='coerce')
            df04 = df04.dropna(subset=['_month_ts']).sort_values('_month_ts')

            tacos_series = _pd04.to_numeric(df04[tacos_col04], errors='coerce')
            valid_tacos  = tacos_series.dropna().tolist()

            if len(valid_tacos) >= 3:
                last3_tacos = valid_tacos[-3:]
                ctx.l3m_tacos_avg = sum(last3_tacos) / len(last3_tacos)
                pp_change = (last3_tacos[-1] - last3_tacos[0]) * 100
                ctx.tacos_trend_pp = round(pp_change, 2)

                if pp_change > 1.5:
                    ctx.tacos_trend = 'increasing'
                elif pp_change < -1.5:
                    ctx.tacos_trend = 'decreasing'
                else:
                    ctx.tacos_trend = 'stable'

                if spend_col04:
                    spend_vals = _pd04.to_numeric(df04[spend_col04], errors='coerce').dropna().tolist()
                    if len(spend_vals) >= 2 and spend_vals[-2] > 0:
                        ctx.mom_spend_change = (spend_vals[-1] - spend_vals[-2]) / spend_vals[-2]

                if sales_col04:
                    sales_vals = _pd04.to_numeric(df04[sales_col04], errors='coerce').dropna().tolist()
                    if len(sales_vals) >= 2 and sales_vals[-2] > 0:
                        ctx.mom_sales_change = (sales_vals[-1] - sales_vals[-2]) / sales_vals[-2]
    except Exception:
        pass

    pa.close()
    return ctx
